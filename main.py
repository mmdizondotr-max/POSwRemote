import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import os
import datetime
import json
import sys
import shutil
import threading
import time
import random
import socket
import queue
import secrets
import re
import zipfile
import tempfile
from typing import Optional, List, Dict, Any, Tuple, Union, Set
from PIL import Image, ImageTk
from style_manager import StyleManager, PASTEL_GREEN_BG, PASTEL_GREEN_ACCENT, PASTEL_GREEN_FG

# pyinstaller --onefile --noconsole --splash splash_image3.png main.py


# --- CONFIGURATION CONSTANTS ---

APP_DATA_DIR = os.path.join(os.environ.get("APPDATA", os.path.expanduser("~")), "MMD_POS_System")

BACKUP_DIR = os.path.join(APP_DATA_DIR, "backups")

# Ensure app data directories exist
if not os.path.exists(APP_DATA_DIR):
    os.makedirs(APP_DATA_DIR)
if not os.path.exists(BACKUP_DIR):
    os.makedirs(BACKUP_DIR)

RECEIPT_FOLDER = "receipts"
INVENTORY_FOLDER = "inventoryreceipts"
SUMMARY_FOLDER = "summaryreceipts"
CORRECTION_FOLDER = "correctionreceipts"
SALESREPORT_FOLDER = "salesreport"
DATA_FILE = "products.xlsx"
CONFIG_FILE = os.path.join(APP_DATA_DIR, "config.json")
LEDGER_FILE = os.path.join(APP_DATA_DIR, "ledger.json")
APP_TITLE = "MMD Inventory Tracker v16.14"  # Refactored Version

# --- EMAIL CONFIGURATION ---
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
SENDER_EMAIL = "mmdpos.diz@gmail.com"
SENDER_PASSWORD = "wvdg kkka myuk inve"

# Ensure folders exist
for folder in [RECEIPT_FOLDER, INVENTORY_FOLDER, SUMMARY_FOLDER, CORRECTION_FOLDER, SALESREPORT_FOLDER]:
    if not os.path.exists(folder):
        os.makedirs(folder)

# --- UTILS ---
TIME_OFFSET = datetime.timedelta(seconds=0)

def _sync_time_background():
    global TIME_OFFSET
    try:
        import ntplib
        client = ntplib.NTPClient()
        response = client.request('ntp.pagasa.dost.gov.ph', version=3, timeout=5)
        
        # NTP response is in UTC timestamp, convert to PH Time (UTC+8)
        ntp_utc = datetime.datetime.fromtimestamp(response.tx_time, tz=datetime.timezone.utc)
        ntp_ph_time = ntp_utc.astimezone(datetime.timezone(datetime.timedelta(hours=8))).replace(tzinfo=None)
        
        # To compute offset, we need the local time at the moment the response was received
        local_time = get_current_time() - TIME_OFFSET # Because get_current_time() already applies it
        TIME_OFFSET = ntp_ph_time - local_time
        print(f"Time synced with PH official time. Offset: {TIME_OFFSET}")
    except Exception as e:
        print(f"Time sync failed: {e}")

# Start the background sync thread
threading.Thread(target=_sync_time_background, daemon=True).start()

def get_current_time() -> datetime.datetime:
    """Returns the official PH government time if online, otherwise local PC time."""
    return datetime.datetime.now() + TIME_OFFSET

def truncate_product_name(name: str) -> str:
    if len(name) > 30:
        return name[:15] + name[-15:]
    return name

# --- DEPENDENCY CONTAINER ---
class AppModules:
    """
    Holds references to lazily imported modules to allow strict typing hints
    and organized access without global pollution.
    """
    pd: Any = None
    canvas: Any = None
    letter: Any = None
    inch: Any = None
    PdfReader: Any = None
    ntplib: Any = None
    Flask: Any = None
    request: Any = None
    jsonify: Any = None
    render_template_string: Any = None
    qrcode: Any = None
    smtplib: Any = None
    ssl: Any = None
    MIMEText: Any = None
    MIMEMultipart: Any = None
    MIMEBase: Any = None
    encoders: Any = None

    @classmethod
    def is_loaded(cls) -> bool:
        return cls.pd is not None

# --- DATA STRUCTURES ---
class TransactionItem(Dict):
    """Type hint helper for transaction items"""
    name: str
    price: float
    qty: int
    category: str
    subtotal: Optional[float]
    # ... other fields allow dynamic keys

# --- DATA MANAGER ---
class DataManager:
    """
    Handles all data persistence, calculation, and product management.
    Separates logic from UI.
    """
    def __init__(self, modules: AppModules):
        self.mod = modules
        self.products_df: Any = None  # Pandas DataFrame
        self.ledger: List[Dict] = []
        self.product_history: List[Dict] = []
        self.summary_count: int = 0
        self.shortcuts_asked: bool = False
        self.business_name = "iMD Pharmacy"
        self.startup_stats: Dict = {}
        self.offline_receipts: List[Dict] = []
        self.last_bi_date: str = ""
        self._ledger_lock = threading.Lock()

        # Caches
        self.stock_cache: Dict[str, Dict] = {}
        self.name_lookup_cache: Dict[str, Dict] = {}
        self.display_name_map: Dict[str, str] = {}  # Full Name -> Smart Display Name
        self.config: Dict = {}
        self.date_fmt = "%Y-%m-%d %H:%M:%S"

        self.load_config()
        self.create_rolling_backup()
        self.load_ledger()
        self.load_products()
        self.refresh_stock_cache()

    def load_config(self) -> None:
        default = {
            "startup": False,
            "splash_img": "",
            "cached_business_name": "iMD Pharmacy",
            "previous_products": [],
            "recipient_email": "",
            "last_bi_date": "",
            "last_email_sync": "",
            "known_categories": []
        }
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, 'r') as f:
                    self.config = json.load(f)
            except:
                self.config = default
        else:
            self.config = default

        # Cleanup legacy config keys
        for key in ["splash_img", "touch_mode"]:
            if key in self.config:
                del self.config[key]

    def save_config(self) -> None:
        try:
            with open(CONFIG_FILE, 'w') as f:
                json.dump(self.config, f)
        except Exception as e:
            print(f"Config Save Error: {e}")

    def load_ledger(self) -> None:
        _loaded = False
        if os.path.exists(LEDGER_FILE):
            try:
                with open(LEDGER_FILE, 'r') as f:
                    data = json.load(f)
                    if isinstance(data, list):
                        self.ledger = data
                        self.summary_count = 0
                        self.shortcuts_asked = False
                        self.product_history = []
                        self.offline_receipts = []
                    elif isinstance(data, dict):
                        self.ledger = data.get("transactions", [])
                        self.summary_count = data.get("summary_count", 0)
                        self.shortcuts_asked = data.get("shortcuts_asked", False)
                        self.product_history = data.get("product_history", [])
                        self.offline_receipts = data.get("offline_receipts", [])
                        self.last_bi_date = data.get("last_bi_date", "")
                    _loaded = True
            except:
                self.ledger = []
                self.product_history = []
                self.offline_receipts = []

        # Re-save to ensure the ledger is stored in minified format
        if _loaded:
            self.save_ledger()

    def create_rolling_backup(self) -> None:
        """Creates a rolling backup of the ledger file in a background thread."""
        if not os.path.exists(LEDGER_FILE):
            return

        thread = threading.Thread(target=self._perform_backup, daemon=True)
        thread.start()

    def _perform_backup(self):
        """The actual backup logic to be run in a thread."""
        try:
            with self._ledger_lock:
                if not os.path.exists(LEDGER_FILE):
                    return

                timestamp = get_current_time().strftime("%Y%m%d-%H%M%S")
                backup_name = f"ledger_backup_{timestamp}.json"
                backup_path = os.path.join(BACKUP_DIR, backup_name)
                shutil.copy2(LEDGER_FILE, backup_path)

            # Cleanup can happen outside the lock
            backups = [
                os.path.join(BACKUP_DIR, f)
                for f in os.listdir(BACKUP_DIR)
                if f.startswith("ledger_backup_") and f.endswith(".json")
            ]
            backups.sort(key=os.path.getctime)

            while len(backups) > 10:
                os.remove(backups.pop(0))

        except Exception as e:
            print(f"Backup Error: {e}")

    def save_ledger(self) -> None:
        try:
            data = {
                "transactions": self.ledger,
                "summary_count": self.summary_count,
                "shortcuts_asked": self.shortcuts_asked,
                "product_history": self.product_history,
                "offline_receipts": self.offline_receipts,
                "last_bi_date": self.last_bi_date
            }

            with self._ledger_lock:
                # Atomic Write
                temp_file = LEDGER_FILE + ".tmp"
                with open(temp_file, 'w') as f:
                    json.dump(data, f, separators=(',', ':'))
                    f.flush()
                    os.fsync(f.fileno())
                os.replace(temp_file, LEDGER_FILE)

        except Exception as e:
            messagebox.showerror("Save Error", f"Could not save database: {e}")

    def load_products(self) -> None:
        pd = self.mod.pd
        req_cols = ["Business Name", "Product Category", "Product Name", "Price"]

        if not os.path.exists(DATA_FILE):
            df = pd.DataFrame(columns=req_cols)
            df.loc[0] = ["iMD Pharmacy", "General", "Sample Product", 100.00]
            try:
                df.to_excel(DATA_FILE, index=False)
            except:
                pass

        raw_df = pd.DataFrame()
        try:
            raw_df = pd.read_excel(DATA_FILE)
            raw_df.columns = raw_df.columns.str.strip()
        except Exception as e:
            messagebox.showerror("Load Error", f"Error reading Excel: {e}")
            self.products_df = pd.DataFrame(columns=req_cols)
            return

        # --- Data Cleanup & Normalization ---
        def clean_text(text_val, is_product_name=False):
            if pd.isna(text_val): return ""
            s = str(text_val).upper()
            s = s.replace("'", "")       # Remove apostrophes
            s = s.replace("\n", " ")     # Single line
            s = re.sub(r'\s+', ' ', s)   # Remove double spaces
            if is_product_name:
                # Remove spaces between numbers and common units
                s = re.sub(r'(\d+)\s+(MG|G|KG|ML|L|OZ|LB|CM|M|MM|PCS)\b', r'\1\2', s)
            return s.strip()

        cleaned_count = 0
        if 'Product Name' in raw_df.columns:
            def clean_and_count(x):
                nonlocal cleaned_count
                orig = str(x)
                new = clean_text(x, is_product_name=True)
                # Count if effectively changed (ignoring NaN vs "" diff if original was nan)
                if pd.notna(x) and orig != new:
                    cleaned_count += 1
                return new
            raw_df['Product Name'] = raw_df['Product Name'].apply(clean_and_count)

        if 'Product Category' in raw_df.columns:
            raw_df['Product Category'] = raw_df['Product Category'].apply(clean_text)

        if 'Remarks' not in raw_df.columns:
            raw_df['Remarks'] = ""

        # Business Name Logic & Cleanup
        if "Business Name" in raw_df.columns:
            # Find first non-empty value
            first_val = "iMD Pharmacy"
            found_idx = -1

            # Search for first valid string
            for idx, val in enumerate(raw_df["Business Name"]):
                s_val = str(val).strip()
                if s_val and s_val.lower() != "nan":
                    first_val = s_val
                    found_idx = idx
                    break

            # If nothing found, try config or keep default
            if found_idx == -1:
                first_val = self.config.get("cached_business_name", "iMD Pharmacy")

            self.business_name = first_val
            self.config["cached_business_name"] = first_val

        valid_products = []
        seen_names = set()
        rejected_details = []

        # Temporary lists for sorting
        valid_rows_list = []
        error_rows_list = []

        for index, row in raw_df.iterrows():
            cat = str(row.get('Product Category', ''))
            name = str(row.get('Product Name', ''))
            raw_price = row.get('Price', 0)

            try:
                price = float(raw_price)
            except:
                price = 0.0

            rejection_reason = None
            if price <= 0 or pd.isna(raw_price): rejection_reason = "Price <= 0"
            elif not cat or cat == 'NAN': rejection_reason = "Invalid Category"
            elif not name or name == 'NAN': rejection_reason = "Invalid Name"
            elif name in seen_names: rejection_reason = "Duplicate Name"

            # Construct row dict for reconstruction
            row_dict = row.to_dict()

            if rejection_reason is None:
                seen_names.add(name)

                # Ensure Business Name is populated in memory even if empty in file row
                raw_b_name = str(row.get('Business Name', ''))
                if not raw_b_name or raw_b_name.lower() == 'nan':
                    b_name = self.business_name
                else:
                    b_name = raw_b_name

                entry = {
                    "Business Name": b_name,
                    "Product Category": cat,
                    "Product Name": name,
                    "Price": price
                }
                valid_products.append(entry)

                # Populate Lookup Cache
                self.name_lookup_cache[name] = entry
                truncated = truncate_product_name(name)
                self.name_lookup_cache[truncated] = entry

                # Clear remarks
                row_dict['Remarks'] = ""
                valid_rows_list.append(row_dict)
            else:
                rejected_details.append({"name": name, "reason": rejection_reason})
                row_dict['Remarks'] = rejection_reason
                error_rows_list.append(row_dict)

        # Sort Logic: Errors first, then valid. Both sorted by Category -> Name
        def sort_helper(r):
            return (str(r.get('Product Category', '')), str(r.get('Product Name', '')))

        error_rows_list.sort(key=sort_helper)
        valid_rows_list.sort(key=sort_helper)

        final_rows = error_rows_list + valid_rows_list
        raw_df = pd.DataFrame(final_rows)

        try:
            # Ensure column order matches standard
            if not raw_df.empty:
                cols = ["Business Name", "Product Category", "Product Name", "Price", "Remarks"]
                # Keep other columns if they exist
                existing_cols = [c for c in raw_df.columns if c not in cols]
                final_cols = [c for c in cols + existing_cols if not str(c).startswith('_')]
                # Reindex safely
                raw_df = raw_df.reindex(columns=final_cols)

            # --- Business Name Cleanup (Visual in File) ---
            # Ensure only the first row (A2) has the business name, others empty.
            if "Business Name" in raw_df.columns:
                raw_df["Business Name"] = ""
                if not raw_df.empty:
                    raw_df.at[0, "Business Name"] = self.business_name

            raw_df.to_excel(DATA_FILE, index=False)

            # Post-processing: Apply Number Format
            import openpyxl
            wb = openpyxl.load_workbook(DATA_FILE)
            ws = wb.active

            # Find Price column index (1-based)
            price_col_idx = None
            for idx, cell in enumerate(ws[1], 1):
                if cell.value and str(cell.value).strip() == "Price":
                    price_col_idx = idx
                    break

            if price_col_idx:
                for row in ws.iter_rows(min_row=2, min_col=price_col_idx, max_col=price_col_idx):
                    for cell in row:
                        cell.number_format = '0.00'

            wb.save(DATA_FILE)

        except Exception as e:
            print(f"Failed to update products.xlsx: {e}")

        # --- Smart Display Name Resolution ---
        self.resolve_display_names(valid_products)

        # Populate Display Name Map (Full Name -> Smart Name)
        self.display_name_map = {}
        for p in valid_products:
            if '_display_name' in p:
                self.display_name_map[p['Product Name']] = p['_display_name']

        self.products_df = pd.DataFrame(valid_products)

        # --- Product History Versioning ---
        current_list = self.products_df.to_dict('records')

        should_save_history = False
        if not self.product_history:
            should_save_history = True
        else:
            # Compare with latest
            # Simple check: json dumps comparison to ensure deep equality including order if sorted,
            # but list order matters in excel, so direct comparison is fine.
            # However, we must ensure we are comparing compatible structures.
            # 'records' gives list of dicts.
            last_version = self.product_history[-1].get('items', [])
            if current_list != last_version:
                should_save_history = True

        if should_save_history and current_list:
            timestamp = get_current_time().strftime("%Y-%m-%d %H:%M:%S")
            self.product_history.append({
                "timestamp": timestamp,
                "items": current_list
            })
            # Keep only last 4 versions (Current + 3 past)
            if len(self.product_history) > 4:
                self.product_history = self.product_history[-4:]
            self.save_ledger()

        # Category Validation
        all_detected_cats = set(p['Product Category'] for p in valid_products)
        known_cats = self.config.get("known_categories", [])
        
        # Backwards Compatibility: If known_categories is empty, initialize it with current ones
        if not known_cats and all_detected_cats:
            self.config["known_categories"] = sorted(list(all_detected_cats))
            self.save_config()
            known_cats = self.config["known_categories"]

        new_categories = sorted(list(all_detected_cats - set(known_cats)))

        # Stats
        previous_products = set(self.config.get("previous_products", []))
        current_products = set(seen_names)
        self.startup_stats = {
            "total": len(valid_products),
            "new": len(current_products - previous_products),
            "rejected": len(rejected_details),
            "phased_out": len(previous_products - current_products),
            "rejected_details": rejected_details,
            "cleaned_names": cleaned_count,
            "new_categories": new_categories
        }
        self.config["previous_products"] = list(seen_names)
        self.save_config()

    def clear_unconfirmed_categories(self, unconfirmed_cats: List[str]) -> None:
        """
        Removes the category value from products.xlsx for rows that match any of the unconfirmed categories.
        """
        try:
            import openpyxl
            wb = openpyxl.load_workbook(DATA_FILE)
            ws = wb.active
            
            # Find Category column index
            cat_col_idx = None
            for idx, cell in enumerate(ws[1], 1):
                if cell.value and str(cell.value).strip() == "Product Category":
                    cat_col_idx = idx
                    break
            
            if cat_col_idx:
                for row in ws.iter_rows(min_row=2):
                    cell = row[cat_col_idx - 1]
                    if cell.value and str(cell.value).strip() in unconfirmed_cats:
                        cell.value = ""
            
            wb.save(DATA_FILE)
        except Exception as e:
            print(f"Failed to clear unconfirmed categories: {e}")

    def add_transaction(self, t_type: str, filename: str, items: List[Dict],
                        timestamp: Optional[str] = None, ref_type: str = None,
                        ref_filename: str = None, **kwargs) -> None:
        if not timestamp:
            timestamp = get_current_time().strftime('%Y-%m-%d %H:%M:%S')

        transaction = {
            "type": t_type,
            "timestamp": timestamp,
            "filename": filename,
            "items": items
        }
        if ref_type: transaction['ref_type'] = ref_type
        if ref_filename: transaction['ref_filename'] = ref_filename
        
        # Add any extra metadata (like cash_tendered, change)
        transaction.update(kwargs)

        self.create_rolling_backup()
        self.ledger.append(transaction)
        self.save_ledger()
        self.refresh_stock_cache()

    def refresh_stock_cache(self) -> None:
        self.stock_cache, _, _, _, _ = self.calculate_stats(None)

    def calculate_stats(self, period_filter: Optional[Tuple[datetime.datetime, datetime.datetime]]) \
            -> Tuple[Dict, int, int, List[str], List[str]]:
        """
        Calculates inventory stats.
        Optimization: We parse dates inside the loop.
        """
        stats = {}
        in_count = 0
        out_count = 0
        corrections_in_period = []
        
        # Time-sensitive rename relevance logic
        all_renames_ledger = []
        min_ts_by_pname = {}  # Track earliest appearance of any name in the period
        rename_fns_in_period = set()

        if self.ledger is None: return stats, 0, 0, [], []
        for transaction in self.ledger:
            try:
                # Type Check
                t_type = transaction.get('type')
                if not t_type: continue

                # Collect all rename events from ledger (globally)
                if t_type == 'correction' and transaction.get('ref_type') == 'rename':
                    all_renames_ledger.append({
                        'ts': transaction.get('timestamp', '0000-00-00 00:00:00'),
                        'old': transaction.get('old_name', '?'),
                        'new': transaction.get('new_name', '?'),
                        'old_cat': transaction.get('old_cat', ''),
                        'new_cat': transaction.get('new_cat', ''),
                        'fn': transaction.get('filename', '?')
                    })

                # Filter by Period
                is_in_period = True
                if period_filter is not None:
                    ts_str = transaction.get('timestamp')
                    try:
                        dt = datetime.datetime.strptime(str(ts_str or ""), self.date_fmt)
                    except:
                        dt = get_current_time()

                    s, e = period_filter
                    if not (s <= dt <= e):
                        is_in_period = False
                
                if not is_in_period:
                    continue
                
                # Below logic is only for transactions in the period
                ts_period = transaction.get('timestamp', '0000-00-00 00:00:00')

                # Below logic is only for transactions in the period (if filter active)
                if t_type == 'correction':
                    fname = transaction.get('filename', 'Unknown')
                    if transaction.get('ref_type') == 'rename':
                        old = transaction.get('old_name', '?')
                        new = transaction.get('new_name', '?')
                        fname = f"RN:{old}->{new} ({fname})"
                        rename_fns_in_period.add(transaction.get('filename', '?'))
                    else:
                        corrections_in_period.append(fname)

                # Aggregate Logic
                if t_type == 'inventory':
                    in_count += 1
                elif t_type == 'sales':
                    out_count += 1
                elif t_type == 'returns':
                    # Returns add back to inventory, so we count them as an "in" event for activity
                    in_count += 1 
                elif t_type == 'damaged':
                    # Damaged items leave inventory, so we count them as an "out" event for activity
                    out_count += 1

                ref_type = transaction.get('ref_type')

                for item in transaction.get('items', []):
                    name = item.get('name', 'Unknown')
                    base_name = item.get('base_product', name)
                    # Track earliest appearance of both specific name and base product in this period
                    for n_key in {name, base_name}:
                        if n_key not in min_ts_by_pname or ts_period < min_ts_by_pname[n_key]:
                            min_ts_by_pname[n_key] = ts_period
                    qty = int(item.get('qty', 0))
                    price = float(item.get('price', 0))
                    item_cat = item.get('category', 'Uncategorized')

                    base_name = item.get('base_product', name)
                    
                    if base_name not in stats:
                        stats[base_name] = {
                            'name': base_name, 'in': 0, 'out': 0, 
                            'returns': 0, 'damaged': 0, 'returns_amt': 0.0,
                            'sales_lines': [], 'in_lines': [], 
                            'return_lines': [], 'damaged_lines': [], 
                            'latest_cat': item_cat
                        }
                    else:
                        stats[base_name]['latest_cat'] = item_cat

                    if t_type == 'sales':
                        amt = float(item.get('subtotal', 0))
                        stats[base_name]['out'] += qty
                        # Include variant name for display if it's different from base_name
                        display_item = {'name': name, 'price': price, 'qty': qty, 'amt': amt, 'category': item_cat}
                        stats[base_name]['sales_lines'].append(display_item)
                    elif t_type == 'inventory':
                        stats[base_name]['in'] += qty
                        stats[base_name]['in_lines'].append({'price': price, 'qty': qty, 'category': item_cat})
                    elif t_type == 'returns':
                        # Returns add back to inventory
                        stats[base_name]['in'] += qty
                        stats[base_name]['returns'] += qty
                        amt = float(item.get('subtotal', price * qty))
                        stats[base_name]['returns_amt'] += amt
                        stats[base_name]['return_lines'].append({'name': name, 'price': price, 'qty': qty, 'amt': amt, 'category': item_cat})
                    elif t_type == 'damaged':
                        # Damaged removes from inventory
                        stats[base_name]['out'] += qty
                        stats[base_name]['damaged'] += qty
                        stats[base_name]['damaged_lines'].append({'name': name, 'price': price, 'qty': qty, 'category': item_cat})
                    elif t_type == 'correction':
                        if ref_type == 'sales':
                            stats[base_name]['out'] += qty
                            amt = qty * price
                            stats[base_name]['sales_lines'].append({'name': name, 'price': price, 'qty': qty, 'amt': amt, 'category': item_cat})
                        elif ref_type == 'inventory':
                            stats[base_name]['in'] += qty
                            stats[base_name]['in_lines'].append({'price': price, 'qty': qty, 'category': item_cat})
                        elif ref_type == 'returns':
                            stats[base_name]['in'] += qty
                            stats[base_name]['returns'] += qty
                            amt = qty * price
                            stats[base_name]['returns_amt'] += amt
                            stats[base_name]['return_lines'].append({'name': name, 'price': price, 'qty': qty, 'amt': amt, 'category': item_cat})
                        elif ref_type == 'damaged':
                            stats[base_name]['out'] += qty
                            stats[base_name]['damaged'] += qty
                            stats[base_name]['damaged_lines'].append({'name': name, 'price': price, 'qty': qty, 'category': item_cat})

            except Exception:
                continue

        # Post-process: Reverse-chronological historical tracing
        master_renames = []
        relevant_renames_fns = set(rename_fns_in_period)
        
        # We trace history backwards. If a product in the summary is "C", 
        # and we find a rename "B -> C" that happened after "C" appeared (or just generally),
        # then "B" also becomes a relevant name to look for in even older renames.
        
        # 1. Sort renames by timestamp (reverse)
        sorted_renames = sorted(all_renames_ledger, key=lambda x: x['ts'], reverse=True)
        
        # 2. Iterate and propagate
        changed = True
        while changed:
            changed = False
            for r in sorted_renames:
                if r['fn'] in relevant_renames_fns:
                    continue
                
                old_n = r['old']
                new_n = r['new']
                r_ts = r['ts']
                
                # Rule: If the NEW name is in our summary, and the rename happened AFTER it first appeared
                # OR if the OLD name is in our summary and it was renamed later.
                is_relevant = False
                if new_n in min_ts_by_pname and r_ts >= min_ts_by_pname[new_n]:
                    is_relevant = True
                elif old_n in min_ts_by_pname and r_ts >= min_ts_by_pname[old_n]:
                    is_relevant = True
                
                if is_relevant:
                    relevant_renames_fns.add(r['fn'])
                    changed = True
                    # Propagate timestamp to the other side so we can continue tracing
                    if old_n not in min_ts_by_pname or r_ts < min_ts_by_pname[old_n]:
                        min_ts_by_pname[old_n] = r_ts
                    if new_n not in min_ts_by_pname or r_ts < min_ts_by_pname[new_n]:
                        min_ts_by_pname[new_n] = r_ts

        # Final build in ledger order
        added_ids = set()
        for r in all_renames_ledger:
            if r['fn'] in relevant_renames_fns and r['fn'] not in added_ids:
                ts_str_r = str(r['ts']).split(' ', 1)[0]
                rename_str = f"{ts_str_r} {r['old']}->{r['new']}"
                
                # Add Category change if applicable
                o_cat = r.get('old_cat', '')
                n_cat = r.get('new_cat', '')
                if o_cat and n_cat and o_cat != n_cat:
                    rename_str += f" | {o_cat}->{n_cat}"
                
                # Add filename
                rename_str += f" ({r['fn']})"
                
                master_renames.append(rename_str)
                added_ids.add(r['fn'])

        return stats, in_count, out_count, corrections_in_period, master_renames

    def get_product_list(self) -> List[Dict]:
        if self.products_df is None or self.products_df.empty:
            return []
        return self.products_df.to_dict('records')

    def get_product_details_by_str(self, selection_string: str) -> Tuple[str, str, float, str]:
        """Returns (code, name, price, category)"""
        if not selection_string: return "", "", 0.0, "Uncategorized"

        # Try direct lookup in cache (Exact Full Name or Truncated Name)
        if selection_string in self.name_lookup_cache:
            item = self.name_lookup_cache[selection_string]
            return "", str(item['Product Name']), float(item['Price']), str(item['Product Category'])

        # Try parsing "Name (Price)" format
        try:
            name_part = selection_string.rsplit(" (", 1)[0]
        except:
            name_part = selection_string

        if name_part in self.name_lookup_cache:
            item = self.name_lookup_cache[name_part]
            return "", item['Product Name'], float(item['Price']), item['Product Category']

        # Fallback for old items not in current product list but in ledger (Phased Out)
        # We rely on what was passed if possible, or return Phased Out
        return "", "Unknown Item", 0.0, "Phased Out"

    def get_stock_level(self, name: str) -> int:
        st = self.stock_cache.get(name, {'in': 0, 'out': 0})
        return st['in'] - st['out']

    def resolve_display_names(self, product_list: List[Dict]) -> None:
        """
        Generates distinct display names for products to avoid ambiguity in dropdowns.
        Updates 'product_list' in-place with a '_display_name' key.
        """
        # Group by initial 30-char truncation
        groups = {}
        for p in product_list:
            name = p['Product Name']
            trunc = truncate_product_name(name)
            if trunc not in groups: groups[trunc] = []
            groups[trunc].append(p)

        # Limits to try for collision resolution
        limits = [45, 60, 100, 200]

        for key, group in groups.items():
            if len(group) == 1:
                p = group[0]
                p['_display_name'] = key
                self.name_lookup_cache[key] = p
            else:
                # Collision detected: Try to resolve by increasing limit
                current_limit_idx = 0
                resolved = False
                current_group = group

                while not resolved and current_limit_idx < len(limits):
                    limit = limits[current_limit_idx]
                    sub_groups = {}

                    for p in current_group:
                        full_name = p['Product Name']
                        if len(full_name) > limit:
                             half = limit // 2
                             new_name = full_name[:half] + full_name[-half:]
                        else:
                             new_name = full_name

                        if new_name not in sub_groups: sub_groups[new_name] = []
                        sub_groups[new_name].append(p)

                    max_collision_size = max(len(sg) for sg in sub_groups.values())
                    if max_collision_size == 1:
                        resolved = True
                        for name_ver, items in sub_groups.items():
                            p = items[0]
                            p['_display_name'] = name_ver
                            self.name_lookup_cache[name_ver] = p
                    else:
                        current_limit_idx += 1

                if not resolved:
                    # Fallback: Use full name if still colliding (should be distinct per validation)
                    for p in current_group:
                        full = p['Product Name']
                        p['_display_name'] = full
                        self.name_lookup_cache[full] = p

# --- REPORT MANAGER ---
class ReportManager:
    """
    Handles PDF generation.
    """
    def __init__(self, modules: AppModules, business_name: str, session_user: str, data_manager: Optional[DataManager] = None):
        self.mod = modules
        self.business_name = business_name
        self.session_user = session_user
        self.data_manager = data_manager

    def generate_thermal_sales_receipt(self, filepath: str, title: str, date_str: str, items: List[Dict],
                                       cash_tendered: float, change: float, user: Optional[str] = None, **kwargs) -> bool:
        try:
            canvas = self.mod.canvas
            inch = self.mod.inch

            # 57mm is approx 2.24 inches
            width = 2.24 * inch
            # Calculate required height: Base header/footer + line per item
            margin = 0.1 * inch
            content_width = width - (2 * margin)
            
            line_height = 10
            # Precise height: Header(120) + Items(22 each) + Totals(80) + Padding(10)
            footer_h = 80
            if kwargs.get('total_discount', 0.0) > 0: footer_h += 12
            height = 120 + (len(items) * 22) + footer_h + 10
            
            c = canvas.Canvas(filepath, pagesize=(width, height))
            y = height - (0.2 * inch)
            
            # --- Header ---
            c.setFont("Helvetica-Bold", 10)
            c.drawCentredString(width / 2.0, y, str(self.business_name))
            y -= 12
            
            c.setFont("Helvetica", 8)
            c.drawCentredString(width / 2.0, y, APP_TITLE)
            y -= 12
            
            c.setFont("Helvetica-Bold", 9)
            c.drawCentredString(width / 2.0, y, str(title))
            y -= 15
            
            c.setFont("Helvetica", 7)
            c.drawString(margin, y, f"Date: {date_str}")
            y -= 10
            current_user = user if user else self.session_user
            c.drawString(margin, y, f"Cashier: {current_user}")
            y -= 12
            
            c.setFont("Helvetica", 6)
            c.drawCentredString(width / 2.0, y, "For internal use only.")
            y -= 10
            
            # --- Divider ---
            c.setDash(2, 2)
            c.line(margin, y, width - margin, y)
            c.setDash(1, 0)
            y -= 12
            
            # --- Items ---
            c.setFont("Helvetica-Bold", 7)
            c.drawString(margin, y, "ITEM")
            c.drawRightString(width - margin, y, "TOTAL")
            y -= 10
            
            c.setDash(2, 2)
            c.line(margin, y, width - margin, y)
            c.setDash(1, 0)
            y -= 12
            
            c.setFont("Helvetica", 7)
            total_amount = 0.0
            total_qty = 0
            total_discount = kwargs.get('total_discount', 0.0)
            
            for item in items:
                raw_name = item.get('name', 'Unknown')
                display_name = str(raw_name)
                if self.data_manager and raw_name in self.data_manager.display_name_map:
                    display_name = str(self.data_manager.display_name_map[raw_name])
                
                # Truncate if too long (25 chars)
                d_s = str(display_name)
                if len(d_s) > 25:
                    display_name = d_s[:12] + "..." + d_s[-10:]
                    
                qty = float(item.get('qty', 0))
                price = float(item.get('price', 0.0))
                subtotal = float(item.get('subtotal', price * qty))
                
                total_amount += subtotal
                total_qty += int(qty)
                
                # Line 1: Item Name (Left)
                # Indent if variant
                is_variant = item.get('is_variant') or display_name.startswith('(')
                p_x = margin + (5 if is_variant else 0)
                c.drawString(p_x, y, str(display_name))
                y -= 10
                
                # Line 2: Qty x Price (Indented) and Subtotal (Right)
                detail_str = f"  {int(qty)} x {price:.2f}"
                c.drawString(p_x, y, detail_str)
                c.drawRightString(width - margin, y, f"{subtotal:.2f}")
                y -= 12

            # --- Divider ---
            c.setDash(2, 2)
            c.line(margin, y, width - margin, y)
            c.setDash(1, 0)
            y -= 12
            
            # --- Totals ---
            c.setFont("Helvetica-Bold", 8)
            
            if total_discount > 0:
                c.drawString(margin, y, "TOTAL DISCOUNT:")
                c.drawRightString(width - margin, y, f"{total_discount:.2f}")
                y -= 12

            c.drawString(margin, y, "TOTAL DUE:")
            c.drawRightString(width - margin, y, f"{total_amount:.2f}")
            y -= 12
            
            c.setFont("Helvetica", 7)
            c.drawString(margin, y, "CASH TENDERED:")
            c.drawRightString(width - margin, y, f"{float(cash_tendered):.2f}")
            y -= 10
            
            c.drawString(margin, y, "CHANGE:")
            c.drawRightString(width - margin, y, f"{float(change):.2f}")
            y -= 15
            
            c.drawString(margin, y, f"Total Items: {int(total_qty)}")
            y -= 15
            
            # --- Footer ---
            c.setFont("Helvetica-Oblique", 7)
            c.drawCentredString(width / 2.0, y, "Thank you for shopping with us!")
            y -= 10
            c.drawCentredString(width / 2.0, y, "Please come again.")
            
            c.save()
            return True
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Could not create 57mm PDF:\n{e}")
            return False

    def generate_thermal_summary_receipt(self, filepath: str, title: str, date_str: str, period_str: str, 
                                         aggregated_data: List[Dict], per_trans_data: List[Dict], is_per_trans: bool, 
                                         tot_sales: float, total_cash: float, total_change: float, 
                                         in_c: int, out_c: int, corr_list: List[str], 
                                         master_renames: Optional[List[str]] = None, user: Optional[str] = None) -> bool:
        try:
            canvas = self.mod.canvas
            inch = self.mod.inch
            width = 2.24 * inch
            margin = 0.1 * inch
            
            items_list = per_trans_data if is_per_trans else aggregated_data
            
            base_height = 250
            
            # --- Wrapping logic to calculate height and display correctly ---
            max_chars = 32 # Capacity for 8pt Courier on 57mm
            
            total_discount_sum = sum(float(pt.get('total_discount', 0.0)) for pt in (per_trans_data or []) if pt.get('type') == 'sales')

            def wrap_txt(txt, limit):
                s = str(txt)
                if len(s) <= limit: return [s]
                res = [s[:limit]]
                rem = s[limit:]
                while rem:
                    res.append("  " + rem[:limit-2])
                    rem = rem[limit-2:]
                return res

            wrapped_corrs = []
            for cr in (corr_list or []):
                wrapped_corrs.extend(wrap_txt(cr, max_chars))
            
            rename_notes = master_renames or []
            wrapped_renames = []
            for rn in rename_notes:
                wrapped_renames.extend(wrap_txt(rn, max_chars))

            current_user = user if user else self.session_user
            user_line_raw = f"Generated: {date_str} by {current_user}"
            wrapped_user_header = wrap_txt(user_line_raw, 40)

            # 1. Header Area: Overhead(63.4) + UserLines(len*12) + Internal(10) + Divider(12)
            calc_h = 63.4 + (len(wrapped_user_header) * 12) + 22
            
            if is_per_trans:
                # Table Header: 10 + Divider(12) = 22
                calc_h += 22
                for pt in per_trans_data:
                    # Header(10)
                    calc_h += 10
                    if pt.get('type') == 'sales':
                        # Items(22 each)
                        calc_h += (len(pt.get('items', [])) * 22)
                        # Discount(10)
                        if float(pt.get('total_discount', 0.0)) > 0: calc_h += 10
                        # Cash/Change(22)
                        if float(pt.get('cash', 0.0)) > 0 or float(pt.get('change', 0.0)) > 0: calc_h += 22
                    else: # Inventory
                        # Items(12 each)
                        calc_h += (len(pt.get('items', [])) * 12)
                    
                    # Gap(5)
                    calc_h += 5
            else:
                # Table Header: 10 + Divider(12) = 22
                calc_h += 22
                current_cat = None
                for it in aggregated_data:
                    cat = it.get('category', 'Uncategorized')
                    if cat != current_cat:
                        calc_h += 12 # Category label
                        current_cat = cat
                    calc_h += 22 # Name(10) + Details(12)
            
            # 2. Main Footer
            # Divider(12) + Net Sales(12) = 24
            calc_h += 24
            
            # Additional footer labels
            tot_returns_amt = sum(float(i.get('returns_amt', 0)) for i in (aggregated_data or []))
            tot_damaged_qty = sum(int(i.get('damaged', 0)) for i in (aggregated_data or []))
            
            if tot_returns_amt > 0:
                calc_h += 22 # Gross(10) + Returns(10) + Dash(2)
            
            if tot_damaged_qty > 0:
                calc_h += 10 # Damaged line
            
            if total_cash > 0 or total_change != 0:
                calc_h += 35 # Cash(10) + Change(10) + Discount(15)
            
            # Total Qty (In/Out) line
            calc_h += 15
            
            # 3. Audit Records
            if wrapped_corrs:
                # Gap(20) + Header(12) + Lines(10 each) + Gap(5)
                calc_h += 20 + 12 + (len(wrapped_corrs) * 10) + 5
            
            if wrapped_renames:
                # Gap(20) + Header(10) + Lines(9 each) + Gap(10)
                calc_h += 20 + 10 + (len(wrapped_renames) * 9) + 10
            
            # 4. Final Closure (15 used by text + 10 padding)
            calc_h += 25
            
            height = calc_h
            c = canvas.Canvas(filepath, pagesize=(width, height))
            y = height - (0.2 * inch)
            
            c.setFont("Helvetica-Bold", 10)
            c.drawCentredString(width / 2.0, y, str(self.business_name))
            y -= 12
            c.setFont("Helvetica", 8)
            c.drawCentredString(width / 2.0, y, APP_TITLE)
            y -= 12
            c.setFont("Helvetica-Bold", 9)
            c.drawCentredString(width / 2.0, y, str(title))
            y -= 15
            
            c.setFont("Helvetica", 7)
            c.drawString(margin, y, f"Period: {period_str}")
            y -= 10
            
            c.setFont("Helvetica", 7)
            for u_line in wrapped_user_header:
                c.drawString(margin, y, u_line)
                y -= 12
            
            c.setFont("Helvetica", 6)
            c.drawCentredString(width / 2.0, y, "For internal use only.")
            y -= 10
            
            c.setDash(2, 2)
            c.line(margin, y, width - margin, y)
            c.setDash(1, 0)
            y -= 12
            
            total_qty = 0
            
            if is_per_trans:
                c.setFont("Helvetica-Bold", 7)
                c.drawString(margin, y, "TIME / TRANSACTION")
                c.drawRightString(width - margin, y, "SUBTOT")
                y -= 10
                c.setDash(2, 2)
                c.line(margin, y, width - margin, y)
                c.setDash(1, 0)
                y -= 12
                
                c.setFont("Helvetica", 7)
                for pt in items_list:
                    if pt.get('type') == 'sales':
                        header_text = f"{pt['time']} SALES - {pt['filename']}"
                        if len(header_text) > 30: header_text = header_text[:27] + "..."
                        
                        c.setFont("Helvetica-Bold", 7)
                        c.drawString(margin, y, header_text)
                        c.drawRightString(width - margin, y, f"{pt['subtotal']:.2f}")
                        y -= 10
                        
                        c.setFont("Helvetica", 7)
                        for item in pt.get('items', []):
                            name = str(item.get('name', 'Unknown'))
                            if len(name) > 20: name = name[:10] + "..." + name[-7:]
                            
                            qty = item.get('qty', 0)
                            price = item.get('price', 0.0)
                            subtotal = item.get('subtotal', 0.0)
                            total_qty += int(qty)
                            
                            c.drawString(margin + 5, y, name)
                            y -= 10
                            c.drawString(margin + 5, y, f"  {int(qty)} x {price:.2f}")
                            c.drawRightString(width - margin, y, f"{subtotal:.2f}")
                            y -= 12
                            
                        # Show per-transaction discount
                        t_disc = float(pt.get('total_discount', 0.0))
                        if t_disc > 0:
                            c.setFont("Helvetica-Bold", 7)
                            c.setFillColor("green")
                            c.drawString(margin + 5, y, "DISCOUNT:")
                            c.drawRightString(width - margin, y, f"{t_disc:.2f}")
                            y -= 10
                            c.setFillColor("black")
                            c.setFont("Helvetica", 7)
                            
                        cash = pt.get('cash', 0.0)
                        change = pt.get('change', 0.0)
                        if cash > 0 or change > 0:
                            c.setFont("Helvetica-Oblique", 7)
                            c.drawString(margin + 5, y, "CASH TENDERED:")
                            c.drawRightString(width - margin, y, f"{cash:.2f}")
                            y -= 10
                            c.drawString(margin + 5, y, "CHANGE:")
                            c.drawRightString(width - margin, y, f"{change:.2f}")
                            y -= 12
                            c.setFont("Helvetica", 7)
                        
                        y -= 5 # space between trans
                            
                    elif pt.get('type') == 'returns':
                        header_text = f"{pt['time']} RETURN - {pt['filename']}"
                        if len(header_text) > 30: header_text = header_text[:27] + "..."
                        
                        c.setFont("Helvetica-Bold", 7)
                        c.drawString(margin, y, header_text)
                        c.drawRightString(width - margin, y, f"-{pt['subtotal']:.2f}")
                        y -= 10
                        
                        c.setFont("Helvetica", 7)
                        for item in pt.get('items', []):
                            name = str(item.get('name', 'Unknown'))
                            if len(name) > 20: name = name[:10] + "..." + name[-7:]
                            
                            qty = item.get('qty', 0)
                            price = item.get('price', 0.0)
                            subtotal = item.get('subtotal', 0.0)
                            
                            c.drawString(margin + 5, y, name)
                            y -= 10
                            c.drawString(margin + 5, y, f"  {int(qty)} x {price:.2f}")
                            c.drawRightString(width - margin, y, f"-{subtotal:.2f}")
                            y -= 12
                        
                        y -= 5 # space between trans

                    elif pt.get('type') == 'damaged':
                        header_text = f"{pt['time']} DAMAGED - {pt['filename']}"
                        if len(header_text) > 30: header_text = header_text[:27] + "..."
                        
                        c.setFont("Helvetica-Bold", 7)
                        c.drawString(margin, y, header_text)
                        c.drawRightString(width - margin, y, f"Qty:{pt['total_qty']}")
                        y -= 10
                        
                        c.setFont("Helvetica", 7)
                        for item in pt.get('items', []):
                            name = str(item.get('name', 'Unknown'))
                            if len(name) > 20: name = name[:10] + "..." + name[-7:]
                            qty = item.get('qty', 0)
                            c.drawString(margin + 5, y, name)
                            c.drawRightString(width - margin, y, str(int(qty)))
                            y -= 12
                        
                        y -= 5 # space between trans
                        header_text = f"{pt['time']} INVENTORY - {pt['filename']}"
                        if len(header_text) > 30: header_text = header_text[:27] + "..."
                        
                        c.setFont("Helvetica-Bold", 7)
                        c.drawString(margin, y, header_text)
                        c.drawRightString(width - margin, y, f"Qty:{pt['total_qty']}")
                        y -= 10
                        
                        c.setFont("Helvetica", 7)
                        for item in pt.get('items', []):
                            name = str(item.get('name', 'Unknown'))
                            if len(name) > 20: name = name[:10] + "..." + name[-7:]
                            
                            qty = item.get('qty', 0)
                            
                            c.drawString(margin + 5, y, name)
                            c.drawRightString(width - margin, y, str(int(qty)))
                            y -= 12
                        
                        y -= 5 # space between trans
            else:
                c.setFont("Helvetica-Bold", 7)
                c.drawString(margin, y, "ITEM")
                c.drawRightString(width - margin, y, "SALES")
                y -= 10
                c.setDash(2, 2)
                c.line(margin, y, width - margin, y)
                c.setDash(1, 0)
                y -= 12
                
                c.setFont("Helvetica", 7)
                current_cat = None
                for item in items_list:
                    cat = item.get('category', 'Uncategorized')
                    if cat != current_cat:
                        c.setFont("Helvetica-Bold", 8)
                        c.drawString(margin, y, f"[{cat.upper()}]")
                        y -= 12
                        current_cat = cat
                        c.setFont("Helvetica", 7)
                        
                    is_var = item.get('is_variant', False)
                    name = str(item.get('name', 'Unknown'))
                    indent = margin + 8 if is_var else margin
                    # Use more robust truncation to satisfy linter
                    d_name = str(name)
                    if len(d_name) > 25:
                        name = d_name[:12] + "..." + d_name[-10:]
                    else:
                        name = d_name
                    
                    in_v = item.get('in', 0)
                    out_v = item.get('out', 0)
                    stk_v = item.get('remaining', 0) if not is_var else "-"
                    ret_v = item.get('returns', 0)
                    dam_v = item.get('damaged', 0)
                    sales = item.get('sales', 0.0)
                    total_qty += int(out_v)
                    
                    extra = ""
                    if ret_v > 0: extra += f" Ret:{int(ret_v)}"
                    if dam_v > 0: extra += f" Dam:{int(dam_v)}"
                    
                    c.drawString(indent, y, name)
                    y -= 10
                    # Pure In/Out for display: Stock activity minus corrections
                    pure_in = in_v - ret_v
                    pure_out = out_v - dam_v
                    
                    c.drawString(indent, y, f"  In:{int(pure_in)} Out:{int(pure_out)} Stk:{stk_v}{extra}")
                    c.drawRightString(width - margin, y, f"{sales:.2f}")
                    y -= 12
                    
            c.setDash(2, 2)
            c.line(margin, y, width - margin, y)
            c.setDash(1, 0)
            y -= 12
            
            # Calculate extra footer totals
            tot_returns_amt = sum(float(i.get('returns_amt', 0)) for i in (aggregated_data or []))
            tot_damaged_qty = sum(int(i.get('damaged', 0)) for i in (aggregated_data or []))
            gross_sales = tot_sales + tot_returns_amt

            c.setFont("Helvetica", 7)
            if tot_returns_amt > 0:
                c.drawString(margin, y, "TOTAL SALES (GROSS):")
                c.drawRightString(width - margin, y, f"{gross_sales:.2f}")
                y -= 12
                c.drawString(margin, y, "TOTAL RETURNS:")
                c.drawRightString(width - margin, y, f"-{tot_returns_amt:.2f}")
                y -= 14
                c.setDash(2, 2)
                # Separator line between returns and net sales (positioned above Net Sales)
                c.line(margin, y + 10, width - margin, y + 10)
                c.setDash(1, 0)
            
            c.setFont("Helvetica-Bold", 8)
            c.drawString(margin, y, "NET SALES (DAY):")
            c.drawRightString(width - margin, y, f"{tot_sales:.2f}")
            y -= 14
            
            c.setFont("Helvetica", 7)
            if tot_damaged_qty > 0:
                c.drawString(margin, y, "TOTAL DAMAGED ITEMS:")
                c.drawRightString(width - margin, y, str(int(tot_damaged_qty)))
                y -= 10

            if total_cash > 0 or total_change != 0:
                c.drawString(margin, y, "TOTAL CASH TND:")
                c.drawRightString(width - margin, y, f"{total_cash:.2f}")
                y -= 10
                c.drawString(margin, y, "TOTAL CHANGE:")
                c.drawRightString(width - margin, y, f"{total_change:.2f}")
                y -= 10
                # Add total discount for the period
                if total_discount_sum > 0:
                    c.setFont("Helvetica-Bold", 7)
                    c.setFillColor("green")
                    c.drawString(margin, y, "TOTAL DISCOUNT:")
                    c.drawRightString(width - margin, y, f"{total_discount_sum:.2f}")
                    c.setFillColor("black")
                    y -= 15
                else:
                    y -= 5
            
            c.setFont("Helvetica", 6)
            c.drawString(margin, y, f"In: {in_c} | Out: {out_c}")
            y -= 15
            
            if wrapped_corrs:
                y -= 20
                c.setFont("Helvetica-Bold", 8)
                c.drawString(margin, y, "Corrections:")
                y -= 12
                c.setFont("Courier", 8)
                for c_line in wrapped_corrs:
                    c.drawString(margin + 5, y, f"- {c_line}")
                    y -= 10
                y -= 5

            if wrapped_renames:
                y -= 20
                c.setFont("Helvetica-BoldOblique", 7)
                c.drawString(margin, y, "Master Update Record:")
                y -= 10
                c.setFont("Courier", 7)
                for r_line in wrapped_renames:
                    c.drawString(margin + 5, y, f"* {r_line}")
                    y -= 9
                y -= 10

            y -= 15
            c.drawCentredString(width / 2.0, y, "*** END OF RECEIPT ***")

            c.save()
            return True
        except Exception as e:
            messagebox.showerror("Export Error", f"Could not create 57mm Summary PDF:\n{e}")
            return False

    def generate_thermal_inventory_receipt(self, filepath: str, title: str, date_str: str, items: List[Dict], user: Optional[str] = None) -> bool:
        try:
            canvas = self.mod.canvas
            inch = self.mod.inch

            # 57mm is approx 2.24 inches
            width = 2.24 * inch
            # Calculate required height: Base header/footer + line per item
            margin = 0.1 * inch
            content_width = width - (2 * margin)
            
            line_height = 10
            # Precise height: Header(97.4) + TableHeader(22) + Items(len*22) + Footer(37) + Safety(10)
            height = 97.4 + 22 + (len(items) * 22) + 37 + 10
            
            c = canvas.Canvas(filepath, pagesize=(width, height))
            y = height - (0.2 * inch)
            
            # --- Header ---
            c.setFont("Helvetica-Bold", 10)
            c.drawCentredString(width / 2.0, y, str(self.business_name))
            y -= 12
            
            c.setFont("Helvetica", 8)
            c.drawCentredString(width / 2.0, y, APP_TITLE)
            y -= 12
            
            c.setFont("Helvetica-Bold", 9)
            c.drawCentredString(width / 2.0, y, str(title))
            y -= 15
            
            c.setFont("Helvetica", 7)
            c.drawString(margin, y, f"Date: {date_str}")
            y -= 10
            current_user = user if user else self.session_user
            c.drawString(margin, y, f"User: {current_user}")
            y -= 12
            
            c.setFont("Helvetica", 6)
            c.drawCentredString(width / 2.0, y, "For internal use only.")
            y -= 10
            
            # --- Divider ---
            c.setDash(2, 2)
            c.line(margin, y, width - margin, y)
            c.setDash(1, 0)
            y -= 12
            
            # --- Items Header ---
            c.setFont("Helvetica-Bold", 7)
            c.drawString(margin, y, "ITEM")
            c.drawRightString(width - margin, y, "NEW STK")
            y -= 10
            
            c.setDash(2, 2)
            c.line(margin, y, width - margin, y)
            c.setDash(1, 0)
            y -= 12
            
            c.setFont("Helvetica", 7)
            total_added = 0
            
            for item in items:
                raw_name = item.get('name', 'Unknown')
                display_name = str(raw_name)
                if self.data_manager and raw_name in self.data_manager.display_name_map:
                    display_name = str(self.data_manager.display_name_map[raw_name])
                
                # Truncate if too long (25 chars)
                d_s = str(display_name)
                if len(d_s) > 25:
                    display_name = d_s[:12] + "..." + d_s[-10:]
                    
                qty = float(item.get('qty', 0))
                new_stock = float(item.get('new_stock', 0))
                
                total_added += int(qty)
                
                # Line 1: Item Name (Left)
                c.drawString(margin, y, str(display_name))
                y -= 10
                
                # Line 2: Qty Added and New Stock (Right)
                detail_str = f"  Added: {int(qty)}"
                c.drawString(margin, y, detail_str)
                c.drawRightString(width - margin, y, f"{int(new_stock)}")
                y -= 12

            # --- Divider ---
            c.setDash(2, 2)
            c.line(margin, y, width - margin, y)
            c.setDash(1, 0)
            y -= 12
            
            # --- Totals ---
            c.setFont("Helvetica-Bold", 8)
            c.drawString(margin, y, "TOTAL ADDED:")
            c.drawRightString(width - margin, y, f"{int(total_added)}")
            y -= 15
            
            # --- Footer ---
            c.setFont("Helvetica-Oblique", 7)
            c.drawCentredString(width / 2.0, y, "Initial Inventory Recorded.")
            y -= 12
            c.drawCentredString(width / 2.0, y, f"Ref: BEGINNING_INV")
            y -= 10
            c.drawCentredString(width / 2.0, y, f"Counter: {self.data_manager.summary_count:04d}" if (self.data_manager and hasattr(self.data_manager, 'summary_count')) else "")
            
            c.save()
            return True
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Could not create 57mm Inventory PDF:\n{e}")
            return False

    def generate_thermal_beginning_inventory_receipt(self, filepath: str, title: str, date_str: str, items: List[Dict], user: Optional[str] = None) -> bool:
        try:
            canvas = self.mod.canvas
            inch = self.mod.inch

            # 57mm is approx 2.24 inches
            width = 2.24 * inch
            margin = 0.1 * inch
            
            # Adobe Acrobat limit is 200 inches (14,400 pts)
            # We'll use multiple pages if we exceed a safe limit per page (150 inches)
            max_page_height = 150 * inch 
            row_height = 13
            # Overhead: Header(104) + Footer(35) + Safety(10)
            base_overhead = 149

            # Group by category
            def sort_key(x):
                cat = x.get('category', 'Uncategorized')
                if cat == "Phased Out": cat = "zzz_Phased Out"
                return (cat, str(x.get('name', '')))
            sorted_items = sorted(items, key=sort_key)
            
            categories = set(it.get('category', 'Uncategorized') for it in sorted_items)
            num_cats = len(categories)
            
            total_needed = base_overhead + ((len(sorted_items) + num_cats) * row_height)
            first_h = min(total_needed, max_page_height)
            
            c = canvas.Canvas(filepath, pagesize=(width, first_h))
            y = first_h - 15
            
            def draw_header(canv, curr_y, page_num=1):
                canv.setFont("Helvetica-Bold", 11)
                canv.drawCentredString(width / 2.0, curr_y, str(self.business_name))
                curr_y -= 14
                
                canv.setFont("Helvetica", 7)
                canv.drawCentredString(width / 2.0, curr_y, APP_TITLE)
                curr_y -= 10
                
                canv.setFont("Helvetica-Bold", 9)
                title_text = f"{title}" + (f" (Page {page_num})" if page_num > 1 else "")
                canv.drawCentredString(width / 2.0, curr_y, title_text)
                curr_y -= 15
                
                canv.setFont("Helvetica", 7)
                canv.drawString(margin, curr_y, f"Date: {date_str}")
                curr_y -= 10
                current_user = user if user else self.session_user
                canv.drawString(margin, curr_y, f"User: {current_user}")
                curr_y -= 10
                
                canv.setFont("Helvetica", 6)
                canv.drawCentredString(width / 2.0, curr_y, "For internal use only.")
                curr_y -= 10
                
                # --- Divider ---
                canv.setDash(2, 2)
                canv.line(margin, curr_y, width - margin, curr_y)
                canv.setDash(1, 0)
                curr_y -= 10
                
                canv.setFont("Helvetica-Bold", 7)
                canv.drawString(margin, curr_y, "ITEM")
                canv.drawRightString(width - margin, curr_y, "QTY")
                curr_y -= 8
                
                canv.setDash(2, 2)
                canv.line(margin, curr_y, width - margin, curr_y)
                canv.setDash(1, 0)
                curr_y -= 12
                return curr_y

            # Draw initial header
            y = draw_header(c, y, 1)
            
            c.setFont("Helvetica", 7)
            total_qty = 0
            page_count = 1
            current_cat = None
            
            for i, item in enumerate(sorted_items):
                cat = item.get('category', 'Uncategorized')
                if cat != current_cat:
                    # Category Header
                    if y < 40: 
                        c.showPage()
                        page_count += 1
                        rem_items = len(sorted_items) - i
                        # Apprx cats remaining
                        rem_cats = len(set(it.get('category') for it in sorted_items[i:]))
                        next_h = min(max_page_height, 100 + ((rem_items + rem_cats) * row_height) + 80)
                        c.setPageSize((width, next_h))
                        y = next_h - 15
                        y = draw_header(c, y, page_count)
                        c.setFont("Helvetica", 7)
                    
                    c.setFont("Helvetica-Bold", 8)
                    c.drawString(margin, y, f"[{cat.upper()}]")
                    y -= row_height
                    current_cat = cat
                    c.setFont("Helvetica", 7)

                # New Page Check (if near bottom of current page)
                if y < 70: 
                    c.showPage()
                    page_count += 1
                    rem_items = len(sorted_items) - (i+1)
                    rem_cats = len(set(it.get('category') for it in sorted_items[i+1:]))
                    next_h = min(max_page_height, 100 + ((rem_items + rem_cats) * row_height) + 80)
                    c.setPageSize((width, next_h))
                    y = next_h - 15
                    y = draw_header(c, y, page_count)
                    c.setFont("Helvetica", 7)

                raw_name = item.get('name', 'Unknown')
                display_name = str(raw_name)
                if self.data_manager and raw_name in self.data_manager.display_name_map:
                    display_name = str(self.data_manager.display_name_map[raw_name])
                
                # Truncate to fit on one line (approx 28 chars)
                d_s = str(display_name)
                if len(d_s) > 28:
                    display_name = d_s[:14] + "..." + d_s[-11:]
                    
                qty = float(item.get('qty', 0))
                total_qty += int(qty)
                
                # Compact One-Line Layout
                c.drawString(margin, y, str(display_name))
                c.drawRightString(width - margin, y, f"{int(qty)}")
                y -= row_height

            # --- Final Footer ---
            y -= 8
            c.setDash(2, 2)
            c.line(margin, y, width - margin, y)
            c.setDash(1, 0)
            y -= 12
            
            c.setFont("Helvetica-Bold", 8)
            c.drawString(margin, y, "TOTAL ITEMS QTY:")
            c.drawRightString(width - margin, y, f"{int(total_qty)}")
            y -= 15
            
            c.setFont("Helvetica-Oblique", 7)
            c.drawCentredString(width / 2.0, y, "*** END OF REPORT ***")
            y -= 10
            c.drawCentredString(width / 2.0, y, f"Counter: {self.data_manager.summary_count:04d}" if (self.data_manager and hasattr(self.data_manager, 'summary_count')) else "")
            
            c.save()
            return True
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Could not create 57mm BI PDF:\n{e}")
            return False

    def generate_grouped_pdf(self, filepath: str, title: str, date_str: str, items: List[Dict],
                             col_headers: List[str], col_pos: List[float], is_summary: bool = False,
                             extra_info: str = "", subtotal_indices: List[int] = None,
                             is_inventory: bool = False, correction_list: List[str] = None,
                             is_bi: bool = False, canvas_obj: Any = None, user: Optional[str] = None) -> bool:
        try:
            canvas = self.mod.canvas
            letter = self.mod.letter
            inch = self.mod.inch

            if canvas_obj:
                c = canvas_obj
            else:
                c = canvas.Canvas(filepath, pagesize=letter)

            width, height = letter
            y = height - 1 * inch

            # Header
            c.setFont("Helvetica-Bold", 18)
            c.drawString(1 * inch, y, self.business_name)
            y -= 0.35 * inch
            c.setFont("Helvetica-Bold", 14)
            c.drawString(1 * inch, y, title)
            y -= 0.25 * inch
            c.setFont("Helvetica", 9)
            c.drawString(1 * inch, y, APP_TITLE)
            y -= 0.2 * inch
            c.setFont("Helvetica", 10)
            c.drawString(1 * inch, y, f"Date: {date_str}")
            y -= 0.2 * inch
            current_user = user if user else self.session_user
            c.drawString(1 * inch, y, f"User: {current_user}")
            if extra_info:
                y -= 0.2 * inch
                c.drawString(1 * inch, y, extra_info)
            y -= 0.2 * inch
            
            c.setFont("Helvetica", 8)
            c.drawString(1 * inch, y, "For internal use only.")
            y -= 0.4 * inch

            # Columns
            c.setFont("Helvetica-Bold", 9)
            for i, h in enumerate(col_headers):
                c.drawString(col_pos[i] * inch, y, h)
            c.line(1 * inch, y - 5, 7.5 * inch, y - 5)
            y -= 20

            def sort_key(x):
                cat = x.get('category', 'Uncategorized')
                if cat == "Phased Out": cat = "zzz_Phased Out"
                return (cat, x['name'])

            sorted_items = sorted(items, key=sort_key)
            current_cat = None
            cat_sums = [0.0] * len(col_headers)
            grand_sums = [0.0] * len(col_headers)

            for item in sorted_items:
                if y < 1 * inch:
                    c.showPage()
                    y = height - 1 * inch

                cat = item.get('category', 'Uncategorized')
                if cat != current_cat:
                    # Subtotal Logic
                    if current_cat is not None:
                        if not is_inventory and not "qty_final" in item:
                            c.setFont("Helvetica-Bold", 9)
                            c.line(col_pos[-1] * inch - 0.5 * inch, y + 2, 7.5 * inch, y + 2)
                            if subtotal_indices:
                                for idx in subtotal_indices:
                                    if idx < len(col_pos):
                                        val = cat_sums[idx]
                                        is_float = (is_summary and idx in [1, 5]) or (not is_summary and idx in [1, 3])
                                        txt = f"{val:.2f}" if is_float else f"{int(val)}"
                                        c.drawString(col_pos[idx] * inch, y - 10, txt)
                            c.drawString(col_pos[-1] * inch - 0.7 * inch, y - 10, "Subtotal:")
                            y -= 30
                        else:
                            y -= 10

                    c.setFont("Helvetica-Bold", 11)
                    c.setFillColor("blue")
                    c.drawString(1 * inch, y, f"Category: {cat}")
                    c.setFillColor("black")
                    y -= 15
                    current_cat = cat
                    cat_sums = [0.0] * len(col_headers)

                c.setFont("Helvetica", 9)
                row_vals = []
                row_txt = []

                # Resolve Name
                raw_name = item.get('name', 'Unknown')
                display_name = raw_name

                # Smart Name Lookup
                if self.data_manager and raw_name in self.data_manager.display_name_map:
                    display_name = self.data_manager.display_name_map[raw_name]

                # Safety Truncation (40 chars max)
                if len(display_name) > 40:
                    display_name = display_name[:20] + "..." + display_name[-17:]

                # Determine Row Data based on context
                if is_bi:
                    row_txt = [display_name, str(int(item['qty']))]
                    row_vals = [0, item['qty']]
                elif is_summary:
                    price_txt = f"{item['price']:.2f}" if item['price'] > 0 else "-"
                    row_txt = [display_name, price_txt, str(int(item['in'])),
                               str(int(item['out'])), str(int(item['remaining'])), f"{item['sales']:.2f}"]
                    row_vals = [0, 0, item['in'], item['out'], item['remaining'], item['sales']]
                elif "subtotal" in item:
                    row_txt = [display_name, f"{item['price']:.2f}", str(int(item['qty'])),
                               f"{item['subtotal']:.2f}"]
                    row_vals = [0, 0, item['qty'], item['subtotal']]
                elif "new_stock" in item:
                    row_txt = [display_name, f"{item.get('price', 0):.2f}", f"{int(item['qty'])}",
                               str(int(item.get('new_stock', 0)))]
                    row_vals = [0, 0, item['qty'], 0]
                elif "qty_final" in item:
                    row_txt = [display_name, str(int(item['qty_orig'])), f"{int(item['qty']):+}",
                               str(int(item['qty_final']))]
                    row_vals = [0, 0, item['qty'], 0]
                else:
                    row_txt = [display_name, f"{item.get('price', 0):.2f}", f"{int(item['qty'])}"]
                    row_vals = [0, 0, item['qty']]

                for i, txt in enumerate(row_txt):
                    c.drawString(col_pos[i] * inch, y, txt)

                for i, val in enumerate(row_vals):
                    cat_sums[i] += val
                    grand_sums[i] += val

                y -= 15

            # Final Subtotal for last category
            if current_cat is not None and not is_inventory and not is_bi and not "qty_final" in (items[0] if items else {}):
                c.setFont("Helvetica-Bold", 9)
                c.line(col_pos[-1] * inch - 0.5 * inch, y + 2, 7.5 * inch, y + 2)
                if subtotal_indices:
                    for idx in subtotal_indices:
                        if idx < len(col_pos):
                            val = cat_sums[idx]
                            is_float = (is_summary and idx in [1, 5]) or (not is_summary and idx in [1, 3])
                            txt = f"{val:.2f}" if is_float else f"{int(val)}"
                            c.drawString(col_pos[idx] * inch, y - 10, txt)
                c.drawString(col_pos[-1] * inch - 0.7 * inch, y - 10, "Subtotal:")
                y -= 30

            # Grand Total
            if not is_bi:
                c.line(1 * inch, y + 10, 7.5 * inch, y + 10)
                c.setFont("Helvetica-Bold", 12)
                lbl = ""
                if is_summary:
                    lbl = f"TOTAL SALES: {grand_sums[5]:.2f}"
                elif items and "subtotal" in items[0]:
                    lbl = f"TOTAL AMOUNT: {grand_sums[3]:.2f}"
                elif is_inventory:
                    lbl = f"TOTAL ADDED: {int(grand_sums[2])}"
                
                if lbl:
                    c.drawString(4.5 * inch, y, lbl)
                    
                    # Compute total discount for this list
                    t_disc_list = 0.0
                    for it in items:
                        t_disc_list += (it.get('orig_price', it.get('price', 0)) - it.get('price', 0.0)) * it.get('qty', 0)
                    
                    if t_disc_list > 0:
                        y -= 0.2 * inch
                        c.setFont("Helvetica-Bold", 9)
                        c.setFillColor("green")
                        c.drawString(4.5 * inch, y, f"Total Discount: {t_disc_list:.2f}")
                        c.setFillColor("black")

            # Corrections List (Summary only)
            if is_summary and correction_list:
                y -= 40
                if y < 1 * inch:
                    c.showPage()
                    y = height - 1 * inch
                c.setFont("Helvetica-Bold", 10)
                c.drawString(1 * inch, y, "Corrections included in this period:")
                y -= 15
                c.setFont("Helvetica", 9)
                for cf in correction_list:
                    if y < 0.5 * inch:
                        c.showPage()
                        y = height - 1 * inch
                    c.drawString(1.2 * inch, y, f"- {cf}")
                    y -= 12

            if not canvas_obj:
                c.save()
            return True
        except Exception as e:
            messagebox.showerror("PDF Error", str(e))
            return False

    def generate_catchup_report(self, filepath: str, intervals: List[Tuple[datetime.datetime, datetime.datetime]],
                                data_manager: 'DataManager', get_stats_func) -> bool:
        try:
            canvas = self.mod.canvas
            letter = self.mod.letter
            c = canvas.Canvas(filepath, pagesize=letter)

            for idx, (start, end) in enumerate(intervals):
                # Calculate stats for this specific interval
                # We reuse the logic from get_sum_data but we need to inject the period
                # Since get_sum_data is in POSSystem, we will need to replicate aggregation logic here or pass a callback
                # Ideally, we pass a callback that returns the formatted data rows for a given period

                rows, in_c, out_c, corr_list = get_stats_func(period=(start, end))

                title = f"CATCHUP SUMMARY ({idx+1}/3)"
                date_str = f"{start.strftime('%Y-%m-%d %H:%M')} to {end.strftime('%Y-%m-%d %H:%M')}"

                self.generate_grouped_pdf(
                    filepath="", # Not used when canvas_obj is passed
                    title=title,
                    date_str=date_str,
                    items=rows,
                    col_headers=["Product", "Price", "Added", "Sold", "Stock", "Sales"],
                    col_pos=[1.0, 4.5, 5.2, 5.9, 6.6, 7.3],
                    is_summary=True,
                    extra_info=f"Interval {idx+1} | In: {in_c} | Out: {out_c}",
                    subtotal_indices=[2, 3, 5],
                    correction_list=corr_list,
                    canvas_obj=c
                )
                c.showPage()

            c.save()
            return True
        except Exception as e:
            print(f"Catchup Report Error: {e}")
            return False

    def generate_product_rename_report(self, filepath: str, date_str: str, 
                                       old_name: str, new_name: str, 
                                       old_cat: str, new_cat: str, 
                                       in_qty: int, out_qty: int, user: str) -> bool:
        try:
            canvas = self.mod.canvas
            inch = self.mod.inch
            width = 2.24 * inch
            margin = 0.1 * inch
            
            # 57mm Thermal format - Dynamic height
            height = 225
            c = canvas.Canvas(filepath, pagesize=(width, height))
            y = height - (0.2 * inch)
            
            c.setFont("Helvetica-Bold", 10)
            c.drawCentredString(width / 2.0, y, str(self.business_name))
            y -= 12
            c.setFont("Helvetica", 8)
            c.drawCentredString(width / 2.0, y, APP_TITLE)
            y -= 18
            
            c.setFont("Helvetica-Bold", 9)
            c.drawCentredString(width / 2.0, y, "PRODUCT RENAME RECEIPT")
            y -= 15
            
            c.setFont("Helvetica", 7)
            c.drawString(margin, y, f"Date: {date_str}")
            y -= 10
            c.drawString(margin, y, f"User: {user}")
            y -= 15
            
            c.setDash(2, 2)
            c.line(margin, y, width - margin, y)
            c.setDash(1, 0)
            y -= 12
            
            c.setFont("Helvetica-Bold", 7)
            c.drawString(margin, y, "RENAME DETAILS")
            y -= 10
            c.setFont("Helvetica", 7)
            
            # Old vs New Name
            c.setFont("Helvetica-Bold", 7)
            c.drawString(margin + 5, y, "FROM:")
            c.setFont("Helvetica", 7)
            c.drawString(margin + 40, y, f"{old_name}")
            y -= 10
            c.setFont("Helvetica-Bold", 7)
            c.drawString(margin + 5, y, "TO:")
            c.setFont("Helvetica", 7)
            c.drawString(margin + 40, y, f"{new_name}")
            y -= 15
            
            # Old vs New Category
            c.setFont("Helvetica-Bold", 7)
            c.drawString(margin + 5, y, "OLD CAT:")
            c.setFont("Helvetica", 7)
            c.drawString(margin + 40, y, f"{old_cat}")
            y -= 10
            c.setFont("Helvetica-Bold", 7)
            c.drawString(margin + 5, y, "NEW CAT:")
            c.setFont("Helvetica", 7)
            c.drawString(margin + 40, y, f"{new_cat}")
            y -= 15
            
            c.setDash(2, 2)
            c.line(margin, y, width - margin, y)
            c.setDash(1, 0)
            y -= 12
            
            c.setFont("Helvetica-Bold", 7)
            c.drawString(margin, y, "AFFECTED QTYS (AS OF EDIT)")
            y -= 10
            c.setFont("Helvetica", 7)
            
            c.drawString(margin + 5, y, f"Total In: {int(in_qty)}")
            y -= 10
            c.drawString(margin + 5, y, f"Total Out: {int(out_qty)}")
            y -= 10
            c.drawString(margin + 5, y, f"Stock Carryover: {int(in_qty - out_qty)}")
            y -= 25
            
            c.setFont("Helvetica-Oblique", 6)
            c.drawCentredString(width / 2.0, y, "*** END OF CORRECTION ***")
            
            c.save()
            return True
        except Exception as e:
            print(f"Rename Report Error: {e}")
            return False

# --- EMAIL MANAGER ---
class EmailManager:
    """
    Handles background email sending.
    """
    def __init__(self, modules: AppModules):
        self.mod = modules
        self.last_email_time = 0

    def validate_email_format(self, email: str) -> bool:
        regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return re.match(regex, email) is not None

    def send_email_thread(self, recipient: str, subject: str, body: str,
                          attachment_paths: List[str] = None, is_test: bool = False,
                          on_success: Any = None, on_error: Any = None) -> None:
        if not recipient or not recipient.strip():
            recipient = SENDER_EMAIL

        def task():
            try:
                # Local imports for thread safety if needed, or use self.mod
                smtplib = self.mod.smtplib
                ssl = self.mod.ssl
                MIMEText = self.mod.MIMEText
                MIMEMultipart = self.mod.MIMEMultipart
                MIMEBase = self.mod.MIMEBase
                encoders = self.mod.encoders

                msg = MIMEMultipart()
                msg['From'] = SENDER_EMAIL
                msg['To'] = recipient
                msg['Subject'] = subject
                msg.attach(MIMEText(body, 'plain'))

                temp_zips = []
                try:
                    if attachment_paths:
                        for path in attachment_paths:
                            if os.path.exists(path):
                                filename = os.path.basename(path)
                                _, ext = os.path.splitext(filename)

                                if ext.lower() in ('.pdf', '.xlsx'):
                                    # Attach PDFs and XLSX directly (already compressed)
                                    with open(path, "rb") as attachment:
                                        part = MIMEBase("application", "octet-stream")
                                        part.set_payload(attachment.read())
                                    encoders.encode_base64(part)
                                    part.add_header("Content-Disposition", f"attachment; filename= {filename}")
                                    msg.attach(part)
                                else:
                                    # Compress non-PDF/non-XLSX files (e.g. .json) into a zip
                                    zip_filename = filename + ".zip"
                                    temp_zip = os.path.join(tempfile.gettempdir(), zip_filename)
                                    temp_zips.append(temp_zip)
                                    with zipfile.ZipFile(temp_zip, 'w', zipfile.ZIP_DEFLATED) as zf:
                                        zf.write(path, filename)
                                    with open(temp_zip, "rb") as attachment:
                                        part = MIMEBase("application", "octet-stream")
                                        part.set_payload(attachment.read())
                                    encoders.encode_base64(part)
                                    part.add_header("Content-Disposition", f"attachment; filename= {zip_filename}")
                                    msg.attach(part)
                finally:
                    # Clean up temporary zip files
                    for tz in temp_zips:
                        try:
                            if os.path.exists(tz):
                                os.remove(tz)
                        except:
                            pass

                context = ssl.create_default_context()
                with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
                    server.starttls(context=context)
                    server.login(SENDER_EMAIL, SENDER_PASSWORD)
                    server.sendmail(SENDER_EMAIL, recipient, msg.as_string())

                # Execute success callback if provided
                if on_success:
                    try:
                        # Schedule in main thread if it involves UI or critical data that shouldn't race
                        # But for config saving (simple file I/O), running in thread is acceptable
                        # as long as DataManager isn't thread-hostile.
                        # DataManager.save_config is just json.dump, which is atomic enough for this use case.
                        on_success()
                    except Exception as callback_err:
                        print(f"Email callback error: {callback_err}")

                print(f"Email sent successfully to {recipient}")

            except Exception as e:
                err_msg = str(e)
                if on_error: on_error(err_msg)
                print(f"Failed to send email: {err_msg}")

        threading.Thread(target=task, daemon=True).start()

    def trigger_summary_email(self, recipient: str, summary_pdf_path: str, ledger_path: str,
                              business_name: str, count: int, user: str, extra_body: str = "",
                              extra_attachments: List[str] = None, on_success: Any = None, on_error: Any = None) -> None:
        # Remove if not recipient: return to allow defaulting in send_email_thread

        date_str = get_current_time().strftime("%Y-%m-%d")
        subject = f"[{count:04d}] {business_name} - System Report ({date_str})"
        
        body = (
            f"Hello,\n\n"
            f"Here is the requested system report and latest ledger data for {business_name}.\n\n"
            f"--- Details ---\n"
            f"Generated By:  {user}\n"
            f"Email Counter: {count:04d}\n"
            f"Date & Time:   {get_current_time().strftime('%Y-%m-%d %H:%M:%S')}\n"
            f"---------------\n\n"
            f"{extra_body}\n"
            f"Attachments: PDF Report & JSON Database Backup (ledger.json).\n\n"
            f"Best regards,\n"
            f"{APP_TITLE}"
        )

        attachments = [summary_pdf_path, ledger_path]
        if extra_attachments:
            attachments.extend(extra_attachments)

        self.send_email_thread(recipient, subject, body, attachments, on_success=on_success, on_error=on_error)

# --- WEB SERVER THREAD ---
class WebServerThread(threading.Thread):
    def __init__(self, modules: AppModules, task_queue: queue.Queue, port: int,
                 data_manager_provider: Any, token_provider: Any):
        super().__init__()
        self.mod = modules
        self.task_queue = task_queue
        self.port = port
        self.get_data_manager = data_manager_provider
        self.get_token = token_provider
        self.app = self.mod.Flask(__name__)
        self.daemon = True

        import logging
        log = logging.getLogger('werkzeug')
        log.setLevel(logging.ERROR)

        self.setup_routes()

    def setup_routes(self):
        @self.app.route('/')
        def index():
            token = self.mod.request.args.get('token')
            current_valid_token = self.get_token()
            if not token or token != current_valid_token:
                return "<h1>403 Forbidden</h1><p>Invalid or Expired QR Code.</p>", 403
            return self.mod.render_template_string(MOBILE_HTML_TEMPLATE, token=current_valid_token)

        @self.app.route('/get_products')
        def get_products():
            token = self.mod.request.args.get('token')
            if not token or token != self.get_token():
                return self.mod.jsonify({"error": "Unauthorized"}), 403

            dm = self.get_data_manager()
            # dm is the DataManager instance
            prods_df = dm.products_df

            prods = prods_df.to_dict(orient='records')
            cleaned_prods = []
            for p in prods:
                name = p.get('Product Name', 'Unknown')
                current_qty = dm.get_stock_level(name)

                cleaned_prods.append({
                    "name": name,
                    "price": float(p.get('Price', 0)),
                    "category": p.get('Product Category', 'General'),
                    "stock": int(current_qty)
                })
            return self.mod.jsonify({"business": dm.business_name, "products": cleaned_prods})

        @self.app.route('/submit_transaction', methods=['POST'])
        def submit():
            token = self.mod.request.args.get('token')
            if not token or token != self.get_token():
                return self.mod.jsonify({"error": "Unauthorized"}), 403

            data = self.mod.request.json
            mode = data.get('mode')
            items = data.get('items', [])

            # Server-side stock validation
            if mode == 'sales':
                dm = self.get_data_manager()
                for item in items:
                    name = item.get('name')
                    req_qty = int(item.get('qty', 0))
                    avail = dm.get_stock_level(name)
                    if req_qty > avail:
                        return self.mod.jsonify({"status": "error",
                                        "message": f"Stock change detected! Only {int(avail)} remaining for {name}."})

            client_ip = self.mod.request.remote_addr
            self.task_queue.put({"type": "web_transaction", "data": data, "ip": client_ip})
            return self.mod.jsonify({"status": "success", "message": "Queued"})

    def run(self):
        try:
            self.app.run(host='0.0.0.0', port=self.port, threaded=True)
        except Exception as e:
            print(f"Web Server Error: {e}")

# --- MAIN SYSTEM GUI ---
class POSSystem:
    def __init__(self, root: tk.Tk, username: str, splash: Any, modules: AppModules):
        self.root = root
        self.mod = modules
        self.root.title(APP_TITLE)
        self.root.geometry("1100x750")
        self.root.minsize(800, 500)
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self._current_tab_index = None
        try:
            self.root.state('zoomed')
        except:
            self.root.attributes('-zoomed', True)

        # Session
        login_time = get_current_time().strftime("%H%M%S")
        self.session_user = f"{username}-{login_time}"

        # Initialize Managers
        if splash: splash.update_status("Loading Data Manager...")
        self.data_manager = DataManager(self.mod)

        self.report_manager = ReportManager(self.mod, self.data_manager.business_name, self.session_user, self.data_manager)
        self.email_manager = EmailManager(self.mod)

        # Carts & UI State
        self.sales_cart: List[Dict] = []
        self.inventory_cart: List[Dict] = []
        self.returns_cart: List[Dict] = []
        self.damaged_cart: List[Dict] = []
        self.correction_cart: List[Dict] = []
        self.remote_requests: List[Dict] = []
        self.lws_sidebars: Dict[str, ttk.Frame] = {}
        self.current_master_renames: List[str] = []

        # Web Server State
        self.web_queue = queue.Queue()
        self.local_ip = self.get_local_ip()
        self.web_port = self.find_free_port()
        self.connected_devices: Dict[str, int] = {}
        self.session_token = secrets.token_hex(4)
        self.web_thread = None
        self.web_server_running = False

        # Build UI
        if splash: splash.update_status("Building UI...")
        self.style_manager = StyleManager(self.root)
        self.setup_ui()
        self.ensure_recipient_email()
        self.show_startup_report()
        self.check_restored_status()
        self.check_edit_notifications()

        # Tasks
        self.root.after(1000, self.check_beginning_inventory_reminder)
        self.root.after(2000, self.check_shortcuts)
        self.root.after(100, self.process_web_queue)
        self.root.after(3000, self.process_offline_receipts)

    # --- UI SETUP ---
    def setup_ui(self):
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(expand=True, fill='both', padx=2, pady=2)

        self.tab_inventory = ttk.Frame(self.notebook)
        self.tab_pos = ttk.Frame(self.notebook)
        self.tab_returns = ttk.Frame(self.notebook)
        self.tab_damaged = ttk.Frame(self.notebook)
        self.tab_correction = ttk.Frame(self.notebook)
        self.tab_receipts = ttk.Frame(self.notebook)
        self.tab_summary = ttk.Frame(self.notebook)
        self.tab_sales_report = ttk.Frame(self.notebook)
        self.tab_settings = ttk.Frame(self.notebook)

        self.notebook.add(self.tab_inventory, text="Inventory")
        self.notebook.add(self.tab_pos, text="POS Sales")
        self.notebook.add(self.tab_returns, text="Returns")
        self.notebook.add(self.tab_damaged, text="Damaged")
        self.notebook.add(self.tab_correction, text="Correction")
        self.notebook.add(self.tab_receipts, text="Receipts")
        self.notebook.add(self.tab_summary, text="Summary")
        self.notebook.add(self.tab_sales_report, text="Sales Report")
        self.notebook.add(self.tab_settings, text="Settings")

        self.setup_inventory_tab()
        self.setup_pos_tab()
        self.setup_returns_tab()
        self.setup_damaged_tab()
        self.setup_correction_tab()
        self.setup_receipts_tab()
        self.setup_summary_tab()
        self.setup_sales_report_tab()
        self.setup_settings_tab()

        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_change)

    def on_tab_change(self, event):
        try:
            new_tab_index = self.notebook.index(self.notebook.select())
        except tk.TclError:
            return

        if hasattr(self, '_current_tab_index') and self._current_tab_index is not None:
            if new_tab_index != self._current_tab_index:
                if self._current_tab_index == self.notebook.index(self.tab_inventory) and getattr(self, 'inventory_cart', []):
                    if not messagebox.askyesno("Uncommitted Items", "You have uncommitted items in INVENTORY.\nAre you sure you want to leave this tab?", icon='warning'):
                        self.notebook.select(self._current_tab_index)
                        return
                elif self._current_tab_index == self.notebook.index(self.tab_pos) and getattr(self, 'sales_cart', []):
                    if not messagebox.askyesno("Uncommitted Items", "You have uncommitted items in POS (SALES).\nAre you sure you want to leave this tab?", icon='warning'):
                        self.notebook.select(self._current_tab_index)
                        return
                elif self._current_tab_index == self.notebook.index(self.tab_returns) and getattr(self, 'returns_cart', []):
                    if not messagebox.askyesno("Uncommitted Items", "You have uncommitted items in RETURNS.\nAre you sure you want to leave this tab?", icon='warning'):
                        self.notebook.select(self._current_tab_index)
                        return
                elif self._current_tab_index == self.notebook.index(self.tab_damaged) and getattr(self, 'damaged_cart', []):
                    if not messagebox.askyesno("Uncommitted Items", "You have uncommitted items in DAMAGED.\nAre you sure you want to leave this tab?", icon='warning'):
                        self.notebook.select(self._current_tab_index)
                        return


        self._current_tab_index = new_tab_index

        self.data_manager.refresh_stock_cache()

        # Reset specific tab states
        if hasattr(self, 'pos_qty_var'): self.pos_qty_var.set(1)
        if hasattr(self, 'lbl_stock_avail'): self.lbl_stock_avail.config(text="")
        if hasattr(self, 'pos_dropdown'): self.pos_dropdown.set('')

        current_tab = self.notebook.tab(self.notebook.select(), "text").upper()
        if current_tab == 'CORRECTION':
            self.refresh_correction_list()
        elif current_tab == 'SUMMARY':
            if hasattr(self, 'report_type'):
                self.report_type.set('Daily')
                self.on_report_type_change()
        elif current_tab == 'RETURNS':
            self.refresh_returns()
        elif current_tab == 'DAMAGED':
            self.refresh_damaged()


    def on_closing(self):
        has_uncommitted = False
        warning_msg = ""
        if getattr(self, 'inventory_cart', []):
            has_uncommitted = True
            warning_msg += "You have uncommitted items in INVENTORY.\n"
        if getattr(self, 'sales_cart', []):
            has_uncommitted = True
            warning_msg += "You have uncommitted items in POS (SALES).\n"
        if getattr(self, 'returns_cart', []):
            has_uncommitted = True
            warning_msg += "You have uncommitted items in RETURNS.\n"
        if getattr(self, 'damaged_cart', []):
            has_uncommitted = True
            warning_msg += "You have uncommitted items in DAMAGED.\n"


        if has_uncommitted:
            if not messagebox.askyesno("Uncommitted Items", f"{warning_msg}Are you sure you want to exit?", icon='warning'):
                return

        try:
            self.generate_daily_summary_on_close()
        except Exception as e:
            print(f"Failed to generate auto daily summary: {e}")

        self.root.destroy()

    def process_offline_receipts(self):
        if not getattr(self.data_manager, 'offline_receipts', None):
            return

        print("Attempting to send offline catchup email...")
        try:
            from pypdf import PdfMerger
        except ImportError as e:
            print(f"Could not import pypdf for catchup email: {e}")
            return
            
        receipts = self.data_manager.offline_receipts
        if not receipts: return
        
        compiled_pdf_path = os.path.join(APP_DATA_DIR, "catchup_compiled_receipts.pdf")
        
        merger = PdfMerger()
        latest_ledger = None
        
        receipts.sort(key=lambda x: x.get('timestamp', ''))
        
        covered_start = receipts[0].get('timestamp', 'Unknown')
        covered_end = receipts[-1].get('timestamp', 'Unknown')
        
        has_pdfs = False
        for r in receipts:
            for fpath in r.get('files', []):
                if fpath.endswith('.pdf') and os.path.exists(fpath):
                    try:
                        merger.append(fpath)
                        has_pdfs = True
                    except Exception as merge_err:
                        print(f"Failed to merge {fpath}: {merge_err}")
                elif fpath.endswith('.json') and 'ledger.json' in fpath:
                    if os.path.exists(fpath):
                        latest_ledger = fpath
                        
        try:
            if has_pdfs:
                merger.write(compiled_pdf_path)
            merger.close()
        except Exception as e:
            print(f"Failed to write compiled PDF: {e}")
            return
            
        recipient = self.data_manager.config.get("recipient_email", "").strip()
        if not recipient:
            recipient = SENDER_EMAIL
            
        safe_biz_name = "".join(c for c in self.data_manager.business_name if c.isalnum() or c in (' ', '_', '-')).strip()
        subject = f"Catchup Email - {safe_biz_name} - {get_current_time().strftime('%Y%m%d')}"
        
        body = (f"Catchup Email.\n\n"
                f"This email contains items that failed to send previously due to being offline.\n"
                f"Covered Period: {covered_start} to {covered_end}\n\n")
        
        if has_pdfs or latest_ledger:
            body += f"Please find the compiled PDF of all unsent receipts attached." if has_pdfs else ""
            body += f"{' The latest Ledger database is also attached.' if latest_ledger else ''}\n\n"
            
        has_text = False
        for r in receipts:
            if 'body' in r:
                body += f"--- {r.get('subject', 'Offline Log')} ---\n{r['body']}\n\n"
                has_text = True
        
        attachments = []
        if has_pdfs and os.path.exists(compiled_pdf_path):
            attachments.append(compiled_pdf_path)
        if latest_ledger:
            attachments.append(latest_ledger)
            
        if not attachments and not has_text:
            print("No valid attachments or logs found for catchup email. Clearing queue.")
            self.data_manager.offline_receipts.clear()
            self.data_manager.save_ledger()
            return
        
        def on_success():
            self.data_manager.offline_receipts.clear()
            self.data_manager.save_ledger()
            try:
                if os.path.exists(compiled_pdf_path):
                    os.remove(compiled_pdf_path)
            except:
                pass
            print("Catchup email sent and offline queue cleared.")
            
        def on_error(err):
            print(f"Catchup email failed again: {err}")
            
        self.email_manager.send_email_thread(
            recipient, subject, body,
            attachments,
            on_success=on_success, on_error=on_error
        )

    def get_dropdown_values(self) -> List[str]:
        # Helper to get formatted "Name (Price)" list
        if self.data_manager.products_df.empty: return []
        sorted_df = self.data_manager.products_df.sort_values(by=["Product Name"])
        # Use Smart Display Name if available, else fallback
        return sorted_df.apply(lambda x: f"{x.get('_display_name', truncate_product_name(x['Product Name']))} ({x['Price']:.2f})", axis=1).tolist()

    def setup_searchable_combobox(self, combo):
        """Enables typing to filter the combobox values."""
        # Attach a cache to the widget instance to avoid re-fetching on every key stroke
        combo._search_cache = []

        def on_focus(event):
            # Refresh cache when user focuses
            combo._search_cache = self.get_dropdown_values()

        def on_keyrelease(event):
            # Ignore navigation keys
            if event.keysym in ('Up', 'Down', 'Return', 'Tab', 'Left', 'Right', 'Escape'): return

            value = event.widget.get()

            # Lazy load if empty (should have been loaded on focus, but safety check)
            if not combo._search_cache:
                combo._search_cache = self.get_dropdown_values()

            all_values = combo._search_cache

            if value == '':
                combo['values'] = all_values
            else:
                filtered = [item for item in all_values if value.lower() in item.lower()]
                combo['values'] = filtered

        combo.bind('<FocusIn>', on_focus)
        combo.bind('<KeyRelease>', on_keyrelease)

    def show_startup_report(self):
        self.root.update()
        stats = self.data_manager.startup_stats

        # --- New Category Confirmation ---
        new_cats = stats.get('new_categories', [])
        if new_cats:
            if not self._handle_new_categories_confirmation(new_cats):
                return # App already closed or about to close inside helper

        
        phased_out_with_stock = 0
        names_in_excel = set(self.data_manager.products_df['Product Name'].astype(str))
        for name, st in self.data_manager.stock_cache.items():
            stock = st['in'] - st['out']
            if name not in names_in_excel and stock != 0:
                phased_out_with_stock += 1

        po_warning = ""
        if phased_out_with_stock > 0:
            po_warning = f"\n\nWARNING: You have {phased_out_with_stock} phased out product(s) with existing inventory.\nPlease go to the Settings tab to clear them."

        if stats.get('rejected', 0) == 0:
            msg = (f"Business: {self.data_manager.business_name}\n"
                   f"Product Load Summary:\nTotal Loaded: {stats['total']}\n"
                   f"New Products: {stats['new']}\n"
                   f"Cleaned Names: {stats.get('cleaned_names', 0)}\n"
                   f"Rejected (Errors): {stats['rejected']}\n"
                   f"Phased-Out: {stats['phased_out']}"
                   f"{po_warning}")
            messagebox.showinfo("Startup Report", msg)
            return

        # Custom Dialog for Rejections
        win = tk.Toplevel(self.root)
        win.title("Startup Report")
        win.geometry("380x350")

        ttk.Label(win, text=f"Business: {self.data_manager.business_name}", font=("Segoe UI", 11, "bold")).pack(pady=10)

        f = ttk.Frame(win)
        f.pack(pady=5, padx=20, fill="x")

        ttk.Label(f, text=f"Total Loaded: {stats['total']}").pack(anchor="w")
        ttk.Label(f, text=f"New Products: {stats['new']}").pack(anchor="w")
        ttk.Label(f, text=f"Cleaned Names: {stats.get('cleaned_names', 0)}").pack(anchor="w")
        ttk.Label(f, text=f"Rejected (Errors): {stats['rejected']}", foreground="red", font=("Segoe UI", 9, "bold")).pack(anchor="w")
        ttk.Label(f, text=f"Phased-Out: {stats['phased_out']}").pack(anchor="w")
        
        if phased_out_with_stock > 0:
            ttk.Label(f, text=po_warning.strip(), foreground="darkorange", font=("Segoe UI", 9, "bold"), wraplength=320).pack(anchor="w", pady=(10,0))

        def show_rejected():
            r_win = tk.Toplevel(win)
            r_win.title("Rejected Products")
            r_win.geometry("600x400")

            tree = ttk.Treeview(r_win, columns=("name", "reason"), show="headings")
            tree.heading("name", text="Product Name")
            tree.heading("reason", text="Reason")
            tree.column("name", width=350)
            tree.column("reason", width=200)
            tree.pack(fill="both", expand=True, padx=10, pady=10)

            for item in stats.get('rejected_details', []):
                tree.insert("", "end", values=(item['name'], item['reason']))

            ttk.Button(r_win, text="Close", command=r_win.destroy).pack(pady=10)

        ttk.Button(win, text="VIEW REJECTED PRODUCTS", command=show_rejected, style="Danger.TButton").pack(pady=15, ipadx=10)
        ttk.Button(win, text="OK", command=win.destroy).pack(pady=5, ipadx=20)

    def _handle_new_categories_confirmation(self, new_cats: List[str]) -> bool:
        """
        Shows a dialog listing new categories. User must type 'confirm' to proceed.
        Returns True if confirmed, False otherwise.
        """
        result = [False]
        win = tk.Toplevel(self.root)
        win.title("NEW PRODUCT CATEGORIES DETECTED")
        win.geometry("450x350")
        win.grab_set() # Modal
        
        # Center the window
        win.update_idletasks()
        w = win.winfo_width()
        h = win.winfo_height()
        sw = win.winfo_screenwidth()
        sh = win.winfo_screenheight()
        win.geometry(f"+{int(sw/2 - w/2)}+{int(sh/2 - h/2)}")

        ttk.Label(win, text="NEW CATEGORIES FOUND:", font=("Segoe UI", 11, "bold")).pack(pady=(20, 10))
        
        # List categories in a scrollable or text area if many
        list_frame = ttk.Frame(win)
        list_frame.pack(fill="both", expand=True, padx=30)
        
        txt = tk.Text(list_frame, height=5, font=("Consolas", 10), state="normal")
        txt.pack(side="left", fill="both", expand=True)
        for cat in new_cats:
            txt.insert("end", f"• {cat}\n")
        txt.config(state="disabled")
        
        scrollbar = ttk.Scrollbar(list_frame, command=txt.yview)
        scrollbar.pack(side="right", fill="y")
        txt['yscrollcommand'] = scrollbar.set

        ttk.Label(win, text="To accept these categories, type 'confirm' below:", font=("Segoe UI", 9)).pack(pady=(15, 5))
        
        entry_var = tk.StringVar()
        entry = ttk.Entry(win, textvariable=entry_var, justify="center", font=("Consolas", 11))
        entry.pack(pady=5, padx=50, fill="x")
        entry.focus_set()

        def on_confirm(*args):
            if entry_var.get().strip().lower() == "confirm":
                # Save to known categories
                known = self.data_manager.config.get("known_categories", [])
                for cat in new_cats:
                    if cat not in known:
                        known.append(cat)
                self.data_manager.config["known_categories"] = sorted(known)
                self.data_manager.save_config()
                result[0] = True
                win.destroy()
            else:
                messagebox.showwarning("Validation", "Please type 'confirm' exactly to proceed.")

        def on_cancel():
            if messagebox.askyesno("Cancel", "Rejecting new categories will clear them from products.xlsx\nand CLOSE the software. Continue?"):
                self.data_manager.clear_unconfirmed_categories(new_cats)
                self.root.destroy()
                sys.exit(0)

        # Intercept window close button
        win.protocol("WM_DELETE_WINDOW", on_cancel)

        btn_frame = ttk.Frame(win)
        btn_frame.pack(pady=20)
        ttk.Button(btn_frame, text="CONFIRM", command=on_confirm).pack(side="left", padx=10)
        ttk.Button(btn_frame, text="CANCEL / REJECT", command=on_cancel).pack(side="left")

        entry.bind("<Return>", on_confirm)
        
        self.root.wait_window(win)
        return result[0]

    def check_restored_status(self):
        if self.data_manager.config.get('restored_flag'):
            try:
                # Generate Yesterday's Summary
                yesterday = get_current_time() - datetime.timedelta(days=1)
                self.generate_daily_summary_on_close(target_date=yesterday)
                
                # Generate Today's BI (Beginning Inventory) - Now with email enabled to notify user after restore
                self.generate_beginning_inventory_report(silent=False)
                
                # Clear flag
                self.data_manager.config['restored_flag'] = False
                self.data_manager.save_config()
                
                messagebox.showinfo("Post-Restore Synchronization", 
                                    "A new Beginning Inventory (BI) report and yesterday's daily summary "
                                    "have been successfully generated following the data restoration.")
            except Exception as e:
                messagebox.showerror("Error", f"Post-restore report generation failed: {e}")

    def check_edit_notifications(self):
        """Checks for any pending product edit notifications stored in config."""
        notifications = self.data_manager.config.get("edit_notifications", [])
        if notifications:
            for note in notifications:
                msg = (f"PRODUCT CHANGE NOTIFICATION\n"
                       f"---------------------------\n"
                       f"A product has been updated in the system:\n\n"
                       f"Name: {note['old_name']} -> {note['new_name']}\n"
                       f"Category: {note['old_cat']} -> {note['new_cat']}\n"
                       f"Affected In: {int(note['in_qty'])}\n"
                       f"Affected Out: {int(note['out_qty'])}\n\n"
                       f"All historical records have been updated.")
                messagebox.showinfo("Product Change Notification", msg)
            
            # Clear notifications after showing
            self.data_manager.config["edit_notifications"] = []
            self.data_manager.save_config()

    # --- INVENTORY TAB ---
    def setup_inventory_tab(self):
        self.tab_inventory.config(style="Inventory.TFrame")
        self.setup_lws_sidebar(self.tab_inventory, "inventory")

        main_content = ttk.Frame(self.tab_inventory, style="Inventory.TFrame")
        main_content.pack(side="left", fill="both", expand=True)

        f = ttk.LabelFrame(main_content, text="Stock In", style="Inventory.TLabelframe")
        f.pack(fill="x", padx=5, pady=5)

        top_bar = ttk.Frame(f, style="Inventory.TFrame")
        top_bar.pack(fill="x", padx=5, pady=5)

        self.inv_prod_var = tk.StringVar()
        self.inv_dropdown = ttk.Combobox(top_bar, textvariable=self.inv_prod_var, width=45)
        self.inv_dropdown['values'] = self.get_dropdown_values()
        self.setup_searchable_combobox(self.inv_dropdown)
        self.inv_dropdown.pack(side="left", padx=5)

        ttk.Label(top_bar, text="Qty:", style="Inventory.TLabel").pack(side="left")
        self.inv_qty_var = tk.IntVar(value=1)
        ttk.Entry(top_bar, textvariable=self.inv_qty_var, width=5).pack(side="left", padx=5)
        ttk.Button(top_bar, text="Add", command=self.add_inv).pack(side="left", padx=10)

        tree_frame = ttk.Frame(main_content, style="Inventory.TFrame")
        tree_frame.pack(fill="both", expand=True, padx=5, pady=5)

        scrollbar = ttk.Scrollbar(tree_frame)
        scrollbar.pack(side="right", fill="y")

        self.inv_tree = ttk.Treeview(tree_frame, columns=("cat", "name", "price", "qty"), show="headings",
                                     yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.inv_tree.yview)
        self.inv_tree.heading("cat", text="Category")
        self.inv_tree.heading("name", text="Product")
        self.inv_tree.heading("price", text="Price")
        self.inv_tree.heading("qty", text="Qty")
        self.inv_tree.pack(fill="both", expand=True)

        b = ttk.Frame(main_content, style="Inventory.TFrame")
        b.pack(fill="x", padx=5, pady=10)
        ttk.Button(b, text="COMMIT STOCK", command=self.commit_inv, style="Accent.TButton").pack(side="right", ipadx=10)
        ttk.Button(b, text="Import from XLSX", command=self.import_stock_xlsx).pack(side="right", padx=5)
        ttk.Button(b, text="Clear", command=self.clear_inv).pack(side="right", padx=5)
        ttk.Button(b, text="Del Line", command=self.del_inv_line).pack(side="right", padx=5)

    def add_inv(self):
        sel, qty = self.inv_prod_var.get(), self.inv_qty_var.get()
        if not sel or qty <= 0: return
        _, name, price, cat = self.data_manager.get_product_details_by_str(sel)

        found = False
        for i in self.inventory_cart:
            if i['name'] == name and i['price'] == price:
                i['qty'] += qty
                found = True
                break
        if not found:
            self.inventory_cart.append({"code": "", "name": name, "price": price, "qty": qty, "category": cat})
        self.refresh_inv()

    def refresh_inv(self):
        for i in self.inv_tree.get_children(): self.inv_tree.delete(i)
        for i in sorted(self.inventory_cart, key=lambda x: (x['category'], x['name'])):
            self.inv_tree.insert("", "end", values=(i['category'], i['name'], f"{i['price']:.2f}", i['qty']))

    def del_inv_line(self):
        if not self.inv_tree.selection(): return
        name = self.inv_tree.item(self.inv_tree.selection()[0])['values'][1]
        self.inventory_cart = [i for i in self.inventory_cart if str(i['name']) != str(name)]
        self.refresh_inv()

    def clear_inv(self):
        self.inventory_cart = []
        self.refresh_inv()

    def commit_inv(self):
        if not self.inventory_cart: return
        now = get_current_time()
        date_str = now.strftime('%Y-%m-%d %H:%M:%S')
        fname = f"Inventory_{now.strftime('%Y%m%d-%H%M%S')}.pdf"

        # Calculate new stocks for receipt
        p_items = []
        for i in self.inventory_cart:
            curr_stock = self.data_manager.get_stock_level(i['name'])
            new_stock = curr_stock + i['qty']
            x = i.copy()
            x['new_stock'] = new_stock
            p_items.append(x)

        success = self.report_manager.generate_thermal_inventory_receipt(
            os.path.join(INVENTORY_FOLDER, fname),
            "INVENTORY RECEIPT", date_str, p_items
        )

        if success:
            self.data_manager.add_transaction("inventory", fname, self.inventory_cart, date_str, user=self.session_user)
            self.clear_inv()
            messagebox.showinfo("Success", f"Stock Added. Receipt: {fname}")

    # --- POS (SALES) TAB ---
    def setup_pos_tab(self):
        self.tab_pos.config(style="Sales.TFrame")
        self.setup_lws_sidebar(self.tab_pos, "sales")

        main_content = ttk.Frame(self.tab_pos, style="Sales.TFrame")
        main_content.pack(side="left", fill="both", expand=True)

        f = ttk.LabelFrame(main_content, text="Sale")
        f.pack(fill="x", padx=5, pady=5)

        input_row = ttk.Frame(f)
        input_row.pack(fill="x", padx=5, pady=5)

        self.pos_prod_var = tk.StringVar()
        self.pos_dropdown = ttk.Combobox(input_row, textvariable=self.pos_prod_var, width=45)
        self.pos_dropdown['values'] = self.get_dropdown_values()
        self.setup_searchable_combobox(self.pos_dropdown)
        self.pos_dropdown.pack(side="left", padx=5)
        self.pos_dropdown.bind("<<ComboboxSelected>>", self.on_pos_sel)

        ttk.Label(input_row, text="Qty:").pack(side="left")
        self.pos_qty_var = tk.IntVar(value=1)
        ttk.Entry(input_row, textvariable=self.pos_qty_var, width=5).pack(side="left", padx=2)
        ttk.Button(input_row, text="ADD", command=self.add_pos).pack(side="left", padx=10)

        self.lbl_stock_avail = ttk.Label(input_row, text="", foreground="blue", font=("Segoe UI", 9, "bold"))
        self.lbl_stock_avail.pack(side="left", padx=10)

        tree_frame = ttk.Frame(main_content)
        tree_frame.pack(fill="both", expand=True, padx=5, pady=5)

        scrollbar = ttk.Scrollbar(tree_frame)
        scrollbar.pack(side="right", fill="y")

        self.pos_tree = ttk.Treeview(tree_frame, columns=("cat", "name", "price", "qty", "sub", "promo"),
                                     show="headings", yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.pos_tree.yview)

        self.pos_tree.heading("cat", text="Cat")
        self.pos_tree.heading("name", text="Product")
        self.pos_tree.heading("price", text="Price")
        self.pos_tree.heading("qty", text="Qty")
        self.pos_tree.heading("sub", text="Sub")
        self.pos_tree.heading("promo", text="")
        self.pos_tree.column("cat", width=60)
        self.pos_tree.column("name", width=250)
        self.pos_tree.column("price", width=60)
        self.pos_tree.column("qty", width=40)
        self.pos_tree.column("sub", width=70)
        self.pos_tree.column("promo", width=35, anchor="center")
        self.pos_tree.pack(fill="both", expand=True)
        self.pos_tree.bind("<ButtonRelease-1>", self.on_pos_tree_click)
        self.pos_tree.tag_configure('variant', foreground='#555555')

        b = ttk.Frame(main_content)
        b.pack(fill="x", padx=5, pady=10)
        
        sums_frame = ttk.Frame(b)
        sums_frame.pack(side="left")
        
        self.lbl_pos_total = ttk.Label(sums_frame, text="Total: 0.00", font=("Segoe UI", 14, "bold"), foreground="#d32f2f")
        self.lbl_pos_total.pack(side="top", anchor="w")
        
        self.lbl_pos_discount = ttk.Label(sums_frame, text="Discount: 0.00", font=("Segoe UI", 10, "bold"), foreground="#2e7d32")
        self.lbl_pos_discount.pack(side="top", anchor="w")

        ttk.Button(b, text="CHECKOUT", command=self.checkout, style="Accent.TButton").pack(side="right", ipadx=20)
        ttk.Button(b, text="Clear", command=self.clear_pos).pack(side="right", padx=5)
        ttk.Button(b, text="Del", command=self.del_pos_line).pack(side="right", padx=5)

    def on_pos_sel(self, e):
        sel = self.pos_prod_var.get()
        if not sel: self.lbl_stock_avail.config(text=""); return

        _, name, _, _ = self.data_manager.get_product_details_by_str(sel)
        real_inv = self.data_manager.get_stock_level(name)
        
        # Count all occurrences of this base product (including variants)
        in_cart = 0
        for i in self.sales_cart:
            if i.get('base_product') == name or i['name'] == name:
                in_cart += i['qty']

        self.lbl_stock_avail.config(text=f"Stk: {int(real_inv - in_cart)}")

    def on_pos_tree_click(self, event):
        region = self.pos_tree.identify_region(event.x, event.y)
        if region == "cell":
            column = self.pos_tree.identify_column(event.x)
            if column == "#6": # Promo Column
                item_id = self.pos_tree.identify_row(event.y)
                if item_id:
                    self.show_promo_options(item_id)

    def show_promo_options(self, item_id):
        # Find index in sales_cart
        item_vals = self.pos_tree.item(item_id, 'values')
        name = item_vals[1].strip() # Might be indented
        
        tags = self.pos_tree.item(item_id, 'tags')
        if not tags: return
        cart_idx = -1
        for t in tags:
            if t.startswith('idx_'):
                cart_idx = int(t.split('_')[1])
                break
        
        if cart_idx == -1 or cart_idx >= len(self.sales_cart): return
        
        item = self.sales_cart[cart_idx]
        
        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label=f"X% Discount", command=lambda: self.ask_discount_promo(cart_idx))
        menu.add_command(label=f"Buy Y Take Z", command=lambda: self.ask_bytz_promo(cart_idx))
        menu.add_separator()
        menu.add_command(label="Remove Promo", command=lambda: self.remove_promo(cart_idx))
        
        # Position menu near cursor
        menu.post(self.root.winfo_pointerx(), self.root.winfo_pointery())

    def ask_discount_promo(self, idx):
        item = self.sales_cart[idx]
        if item.get('is_variant') and 'OFF)' not in item['name']:
            messagebox.showwarning("Warning", "Cannot apply discount to this variant.")
            return
            
        percent = simpledialog.askinteger("Discount", "Enter Discount Percentage (1-99):", minvalue=1, maxvalue=99, parent=self.root)
        if percent is None: return
        
        orig_name = item.get('base_product', item['name'])
        orig_price = item.get('orig_price', item['price'])
        
        new_price = orig_price * (1 - (percent / 100.0))
        new_name = f"({percent}%OFF){orig_name}"
        
        # Update item
        item['name'] = new_name
        item['price'] = new_price
        item['subtotal'] = new_price * item['qty']
        item['is_variant'] = True
        item['base_product'] = orig_name
        item['orig_price'] = orig_price
        
        self.refresh_pos()

    def ask_bytz_promo(self, idx):
        item = self.sales_cart[idx]
        # BYTZ logic requires knowing the total quantity of the base product
        base_name = item.get('base_product', item['name'])
        
        y = simpledialog.askinteger("Buy Y", "Enter 'Y' (Buy Y):", minvalue=1, parent=self.root)
        if y is None: return
        z = simpledialog.askinteger("Take Z", "Enter 'Z' (Take Z):", minvalue=1, parent=self.root)
        if z is None: return
        
        self.apply_bytz_logic(base_name, y, z)
        self.refresh_pos()

    def apply_bytz_logic(self, base_name, y, z):
        # 1. Collect total qty of this base product
        total_qty = 0
        cat = "General"
        orig_price = 0
        
        new_cart = []
        for item in self.sales_cart:
            if item.get('base_product', item['name']) == base_name:
                total_qty += item['qty']
                cat = item['category']
                orig_price = item.get('orig_price', item['price'])
            else:
                new_cart.append(item)
        
        if total_qty == 0: return # Should not happen
        
        # 2. Calculate Split
        set_size = y + z
        num_sets = total_qty // set_size
        remainder = total_qty % set_size
        
        free_qty = num_sets * z
        paid_qty = (num_sets * y) + remainder
        
        # 3. Add items back to cart
        if paid_qty > 0:
            new_cart.append({
                "code": "", "name": base_name, "price": orig_price, 
                "qty": paid_qty, "subtotal": orig_price * paid_qty, 
                "category": cat, "base_product": base_name, "orig_price": orig_price
            })
        
        if free_qty > 0:
            variant_name = f"(B{y}T{z}){base_name}"
            new_cart.append({
                "code": "", "name": variant_name, "price": 0.0, 
                "qty": free_qty, "subtotal": 0.0, 
                "category": cat, "is_variant": True, "base_product": base_name, "orig_price": orig_price
            })
            
        self.sales_cart = new_cart

    def remove_promo(self, idx):
        item = self.sales_cart[idx]
        base_name = item.get('base_product', item['name'])
        orig_price = item.get('orig_price', item['price'])
        cat = item['category']
        
        # Restore all items of this base product to normal
        total_qty = 0
        new_cart = []
        for it in self.sales_cart:
            if it.get('base_product', it['name']) == base_name:
                total_qty += it['qty']
            else:
                new_cart.append(it)
        
        new_cart.append({
            "code": "", "name": base_name, "price": orig_price, 
            "qty": total_qty, "subtotal": orig_price * total_qty, 
            "category": cat, "base_product": base_name, "orig_price": orig_price
        })
        self.sales_cart = new_cart
        self.refresh_pos()

    def add_pos(self):
        sel, qty = self.pos_prod_var.get(), self.pos_qty_var.get()
        if not sel or qty <= 0: return
        _, name, price, cat = self.data_manager.get_product_details_by_str(sel)

        real_inv = self.data_manager.get_stock_level(name)
        
        # Count all occurrences of this base product
        cart_q = 0
        for i in self.sales_cart:
            if i.get('base_product', i['name']) == name:
                cart_q += i['qty']

        if (cart_q + qty) > real_inv:
            messagebox.showerror("Stock", f"Low Stock!\nAvail: {int(real_inv)}")
            return

        # Check if there's an existing BYTZ promo for this product
        has_bytz = False
        promo_params = (0, 0)
        for i in self.sales_cart:
            if i.get('base_product', i['name']) == name and i['name'].startswith('(B') and 'T' in i['name']:
                # Extract Y and Z from name like (B1T1)...
                import re
                m = re.match(r'\(B(\d+)T(\d+)\)', i['name'])
                if m:
                    has_bytz = True
                    promo_params = (int(m.group(1)), int(m.group(2)))
                    break
        
        if has_bytz:
            # Re-apply BYTZ logic after adding qty to total
            # Find base product item (if exists) or add it
            found = False
            for i in self.sales_cart:
                if i['name'] == name:
                    i['qty'] += qty
                    i['subtotal'] = i['price'] * i['qty']
                    found = True
                    break
            if not found:
                # Add temporary base item to be merged by BYTZ logic
                self.sales_cart.append({
                    "code": "", "name": name, "price": price, "qty": qty, 
                    "subtotal": price * qty, "category": cat, 
                    "base_product": name, "orig_price": price
                })
            self.apply_bytz_logic(name, promo_params[0], promo_params[1])
        else:
            # Normal add or merge into non-promo item
            found = False
            for i in self.sales_cart:
                if i['name'] == name:
                    i['qty'] += qty
                    i['subtotal'] = i['price'] * i['qty']
                    found = True
                    break
            if not found:
                self.sales_cart.append({
                    "code": "", "name": name, "price": price, "qty": qty, 
                    "subtotal": price * qty, "category": cat, 
                    "base_product": name, "orig_price": price
                })

        self.refresh_pos()
        self.on_pos_sel(None)

    def refresh_pos(self):
        for i in self.pos_tree.get_children(): self.pos_tree.delete(i)
        tot = 0
        total_discount = 0.0
        
        # Sorting logic: Variants follow parents
        def get_sort_key(item):
            base = item.get('base_product', item['name'])
            cat = item['category']
            is_var = item.get('is_variant', False)
            return (cat, base, is_var, item['name'])
            
        sorted_cart = sorted(self.sales_cart, key=get_sort_key)
        
        for i in sorted_cart:
            actual_idx = self.sales_cart.index(i)
            
            display_name = i['name']
            tags = (f'idx_{actual_idx}',)
            if i.get('is_variant'):
                display_name = "  " + display_name
                tags += ('variant',)
                
            orig_p = i.get('orig_price', i['price'])
            disc = (orig_p - i['price']) * i['qty']
            total_discount += disc
            
            self.pos_tree.insert("", "end", values=(i['category'], display_name, f"{i['price']:.2f}", i['qty'],
                                                    f"{i['subtotal']:.2f}", "\u25BC"),
                                 tags=tags)
            tot += i['subtotal']
            
        self.lbl_pos_total.config(text=f"Total Due: {tot:.2f}")
        self.lbl_pos_discount.config(text=f"Total Discount: {total_discount:.2f}")

    def del_pos_line(self):
        if not self.pos_tree.selection(): return
        item_id = self.pos_tree.selection()[0]
        tags = self.pos_tree.item(item_id, 'tags')
        
        cart_idx = -1
        for t in tags:
            if t.startswith('idx_'):
                cart_idx = int(t.split('_')[1])
                break
                
        if cart_idx != -1:
            self.sales_cart.pop(cart_idx)
            self.refresh_pos()
            self.on_pos_sel(None)

    def clear_pos(self):
        self.sales_cart = []
        self.refresh_pos()
        self.on_pos_sel(None)

    def checkout(self):
        if not self.sales_cart: return
        
        tot = sum([i['subtotal'] for i in self.sales_cart])
        
        # Calculate discount for display
        total_discount = 0.0
        for i in self.sales_cart:
            orig_p = i.get('orig_price', i['price'])
            total_discount += (orig_p - i['price']) * i['qty']
        
        # Payment Dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("Checkout Payment")
        dialog.geometry("350x300")
        dialog.grab_set()
        dialog.resizable(False, False)
        
        f = ttk.Frame(dialog, padding=20)
        f.pack(fill="both", expand=True)
        
        if total_discount > 0:
            ttk.Label(f, text=f"Discount Applied: {total_discount:.2f}", font=("Segoe UI", 10), foreground="#2e7d32").pack()
            
        ttk.Label(f, text=f"Total Due: {tot:.2f}", font=("Segoe UI", 16, "bold")).pack(pady=10)
        
        ttk.Label(f, text="Cash Tendered:", font=("Segoe UI", 12)).pack(pady=5)
        cash_var = tk.StringVar()
        cash_entry = ttk.Entry(f, textvariable=cash_var, font=("Segoe UI", 14), justify="center")
        cash_entry.pack(pady=5)
        cash_entry.focus_set()
        
        lbl_change = ttk.Label(f, text="Change: 0.00", font=("Segoe UI", 14, "bold"), foreground="green")
        lbl_change.pack(pady=10)
        
        def update_change(*args):
            try:
                cash = float(cash_var.get())
                change = cash - tot
                if change >= 0:
                    lbl_change.config(text=f"Change: {change:.2f}", foreground="green")
                else:
                    lbl_change.config(text=f"Change: {change:.2f}", foreground="red")
            except ValueError:
                lbl_change.config(text="Change: 0.00", foreground="green")
                
        cash_var.trace_add("write", update_change)
        
        def on_confirm(*args):
            try:
                cash = float(cash_var.get())
            except ValueError:
                messagebox.showerror("Error", "Please enter a valid cash amount.", parent=dialog)
                return
                
            if cash < tot:
                messagebox.showerror("Error", "Cash tendered is less than total amount.", parent=dialog)
                return
                
            change = cash - tot
            dialog.destroy()
            self._finalize_checkout(cash, change)
            
        btn_frame = ttk.Frame(f)
        btn_frame.pack(fill="both", expand=True, pady=10)
        
        # Use classic styling config to guarantee correct display. The 'Accent.TButton' will work with clam theme.
        ttk.Button(btn_frame, text="Cancel", command=dialog.destroy).pack(side="left", expand=True, fill="x", padx=10, ipady=5)
        ttk.Button(btn_frame, text="Confirm\n(Proceed)", style="Accent.TButton", command=on_confirm).pack(side="left", expand=True, fill="x", padx=10, ipady=5)

        
        dialog.bind('<Return>', on_confirm)
        dialog.bind('<Escape>', lambda e: dialog.destroy())

    def _finalize_checkout(self, cash_tendered: float, change: float):
        now = get_current_time()
        date_str = now.strftime('%Y-%m-%d %H:%M:%S')
        fname = f"{now.strftime('%Y%m%d-%H%M%S')}.pdf"

        # Calculate Total Discount for the transaction
        total_discount = 0.0
        for i in self.sales_cart:
            orig_p = i.get('orig_price', i['price'])
            total_discount += (orig_p - i['price']) * i['qty']

        # Use new 57mm thermal receipt generator
        success = self.report_manager.generate_thermal_sales_receipt(
            os.path.join(RECEIPT_FOLDER, fname),
            "SALES RECEIPT", date_str, self.sales_cart,
            cash_tendered, change, total_discount=total_discount
        )

        if success:
            self.data_manager.add_transaction("sales", fname, self.sales_cart, date_str, 
                                              cash_tendered=cash_tendered, change=change, 
                                              total_discount=total_discount, user=self.session_user)
            self.clear_pos()
            messagebox.showinfo("Success", f"Saved: {fname}")

    # --- RETURNS TAB ---
    def setup_returns_tab(self):
        self.tab_returns.config(style="Sales.TFrame")
        self.setup_lws_sidebar(self.tab_returns, "returns")

        main_content = ttk.Frame(self.tab_returns, style="Sales.TFrame")
        main_content.pack(side="left", fill="both", expand=True)

        f = ttk.LabelFrame(main_content, text="Return Items")
        f.pack(fill="x", padx=5, pady=5)

        input_row = ttk.Frame(f)
        input_row.pack(fill="x", padx=5, pady=5)

        self.ret_prod_var = tk.StringVar()
        self.ret_dropdown = ttk.Combobox(input_row, textvariable=self.ret_prod_var, width=45)
        self.ret_dropdown['values'] = self.get_dropdown_values()
        self.setup_searchable_combobox(self.ret_dropdown)
        self.ret_dropdown.pack(side="left", padx=5)

        ttk.Label(input_row, text="Qty:").pack(side="left")
        self.ret_qty_var = tk.IntVar(value=1)
        ttk.Entry(input_row, textvariable=self.ret_qty_var, width=5).pack(side="left", padx=2)
        ttk.Button(input_row, text="ADD", command=self.add_return).pack(side="left", padx=10)

        self.lbl_ret_sold_7d = ttk.Label(input_row, text="", foreground="blue", font=("Segoe UI", 9, "bold"))
        self.lbl_ret_sold_7d.pack(side="left", padx=10)

        self.ret_dropdown.bind("<<ComboboxSelected>>", self.on_ret_sel)

        tree_frame = ttk.Frame(main_content)
        tree_frame.pack(fill="both", expand=True, padx=5, pady=5)

        scrollbar = ttk.Scrollbar(tree_frame)
        scrollbar.pack(side="right", fill="y")

        self.ret_tree = ttk.Treeview(tree_frame, columns=("cat", "name", "price", "qty", "sub", "promo"),
                                     show="headings", yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.ret_tree.yview)

        self.ret_tree.heading("cat", text="Cat")
        self.ret_tree.heading("name", text="Product")
        self.ret_tree.heading("price", text="Price")
        self.ret_tree.heading("qty", text="Qty")
        self.ret_tree.heading("sub", text="Sub")
        self.ret_tree.heading("promo", text="[EDIT]")
        self.ret_tree.column("cat", width=60)
        self.ret_tree.column("name", width=250)
        self.ret_tree.column("price", width=60)
        self.ret_tree.column("qty", width=40)
        self.ret_tree.column("sub", width=70)
        self.ret_tree.column("promo", width=35, anchor="center")
        self.ret_tree.pack(fill="both", expand=True)
        self.ret_tree.bind("<ButtonRelease-1>", self.on_ret_tree_click)
        self.ret_tree.tag_configure('variant', foreground='#555555')

        b = ttk.Frame(main_content)
        b.pack(fill="x", padx=5, pady=10)
        
        self.lbl_ret_total = ttk.Label(b, text="Total Refund: 0.00", font=("Segoe UI", 14, "bold"), foreground="#d32f2f")
        self.lbl_ret_total.pack(side="left", anchor="w")

        ttk.Button(b, text="RETURN", command=self.commit_returns, style="Accent.TButton").pack(side="right", ipadx=20)
        ttk.Button(b, text="Clear", command=self.clear_returns).pack(side="right", padx=5)
        ttk.Button(b, text="Del", command=self.del_returns_line).pack(side="right", padx=5)

    def on_ret_sel(self, e=None):
        sel = self.ret_prod_var.get()
        if not sel: self.lbl_ret_sold_7d.config(text=""); return

        _, name, _, _ = self.data_manager.get_product_details_by_str(sel)
        
        # Calculate items sold in last 7 days
        sold_7d = 0
        now = get_current_time()
        start_date = now - datetime.timedelta(days=7)
        
        for trans in self.data_manager.ledger:
            if trans.get('type') == 'sales':
                try:
                    ts = datetime.datetime.strptime(trans.get('timestamp', ''), "%Y-%m-%d %H:%M:%S")
                    if ts >= start_date:
                        for item in trans.get('items', []):
                            if item.get('name') == name or item.get('base_product') == name:
                                sold_7d += item.get('qty', 0)
                except:
                    continue
        
        self.lbl_ret_sold_7d.config(text=f"Sold (7d): {int(sold_7d)}")

    def add_return(self):
        sel, qty = self.ret_prod_var.get(), self.ret_qty_var.get()
        if not sel or qty <= 0: return
        _, name, price, cat = self.data_manager.get_product_details_by_str(sel)

        found = False
        for i in self.returns_cart:
            if (i['name'] == name or i.get('base_product') == name) and i['price'] == price:
                i['qty'] += qty
                i['subtotal'] = i['price'] * i['qty']
                found = True
                break
        if not found:
            self.returns_cart.append({
                "code": "", "name": name, "price": price, "qty": qty, 
                "subtotal": price * qty, "category": cat,
                "orig_price": price # Store orig_price for promo logic
            })
        self.refresh_returns()

    def refresh_returns(self):
        for i in self.ret_tree.get_children(): self.ret_tree.delete(i)
        tot = 0
        for idx, i in enumerate(self.returns_cart):
            # Show a clearer indicator that this cell is clickable for adjustments
            # Clearer indicator for clickable cell
            p_tag = " [ \u25BC ] " 
            tag = ('idx_' + str(idx),)
            if i.get('is_variant'): tag += ('variant',)
            
            self.ret_tree.insert("", "end", values=(
                i['category'], i['name'], f"{i['price']:.2f}", i['qty'], f"{i['subtotal']:.2f}", p_tag
            ), tags=tag)
            tot += i['subtotal']
        self.lbl_ret_total.config(text=f"Total Refund: {tot:.2f}")

    def on_ret_tree_click(self, event):
        region = self.ret_tree.identify_region(event.x, event.y)
        if region == "cell":
            column = self.ret_tree.identify_column(event.x)
            if column == "#6": # Promo Column
                item_id = self.ret_tree.identify_row(event.y)
                if item_id:
                    self.show_ret_promo_options(item_id)

    def show_ret_promo_options(self, item_id):
        tags = self.ret_tree.item(item_id, 'tags')
        if not tags: return
        cart_idx = -1
        for t in tags:
            if t.startswith('idx_'):
                cart_idx = int(t.split('_')[1])
                break
        
        if cart_idx == -1 or cart_idx >= len(self.returns_cart): return
        
        menu = tk.Menu(self.root, tearoff=0)
        menu.add_command(label=f"Manual Price", command=lambda: self.ask_ret_manual_price(cart_idx))
        menu.add_command(label=f"X% Discount", command=lambda: self.ask_ret_discount_promo(cart_idx))
        menu.add_command(label=f"Buy Y Take Z", command=lambda: self.ask_ret_bytz_promo(cart_idx))
        menu.add_separator()
        menu.add_command(label="Remove Adj/Promo", command=lambda: self.remove_ret_promo(cart_idx))
        menu.post(self.root.winfo_pointerx(), self.root.winfo_pointery())

    def ask_ret_manual_price(self, idx):
        item = self.returns_cart[idx]
        orig_name = item.get('base_product', item['name'])
        orig_price = item.get('orig_price', item['price'])
        
        new_price = simpledialog.askfloat("Manual Price", f"Enter Manual Refund Price for {orig_name}:", 
                                          initialvalue=item['price'], minvalue=0.0, parent=self.root)
        if new_price is None: return
        
        new_name = f"(Price Adj){orig_name}"
        
        item.update({
            'name': new_name, 'price': new_price, 'subtotal': new_price * item['qty'],
            'is_variant': True, 'base_product': orig_name, 'orig_price': orig_price
        })
        self.refresh_returns()

    def ask_ret_discount_promo(self, idx):
        item = self.returns_cart[idx]
        percent = simpledialog.askinteger("Return Discount", "Enter Discount Percentage (1-99):", minvalue=1, maxvalue=99, parent=self.root)
        if percent is None: return
        
        orig_name = item.get('base_product', item['name'])
        orig_price = item.get('orig_price', item['price'])
        
        new_price = orig_price * (1 - (percent / 100.0))
        new_name = f"({percent}%OFF){orig_name}"
        
        item.update({
            'name': new_name, 'price': new_price, 'subtotal': new_price * item['qty'],
            'is_variant': True, 'base_product': orig_name, 'orig_price': orig_price
        })
        self.refresh_returns()

    def ask_ret_bytz_promo(self, idx):
        item = self.returns_cart[idx]
        base_name = item.get('base_product', item['name'])
        y = simpledialog.askinteger("Buy Y", "Enter 'Y' (Buy Y):", minvalue=1, parent=self.root)
        if y is None: return
        z = simpledialog.askinteger("Take Z", "Enter 'Z' (Take Z):", minvalue=1, parent=self.root)
        if z is None: return
        
        # BYTZ logic for returns
        total_qty = 0
        cat, orig_price = "General", 0.0
        new_cart = []
        for it in self.returns_cart:
            if it.get('base_product', it['name']) == base_name:
                total_qty += it['qty']
                cat, orig_price = it['category'], it.get('orig_price', it['price'])
            else:
                new_cart.append(it)
        
        set_size = y + z
        num_sets, remainder = total_qty // set_size, total_qty % set_size
        free_qty, paid_qty = num_sets * z, (num_sets * y) + remainder
        
        if paid_qty > 0:
            new_cart.append({
                "code": "", "name": base_name, "price": orig_price, "qty": paid_qty, 
                "subtotal": orig_price * paid_qty, "category": cat, "base_product": base_name, "orig_price": orig_price
            })
        if free_qty > 0:
            new_cart.append({
                "code": "", "name": f"(B{y}T{z}){base_name}", "price": 0.0, "qty": free_qty, 
                "subtotal": 0.0, "category": cat, "is_variant": True, "base_product": base_name, "orig_price": orig_price
            })
        self.returns_cart = new_cart
        self.refresh_returns()

    def remove_ret_promo(self, idx):
        item = self.returns_cart[idx]
        base_name = item.get('base_product', item['name'])
        orig_price, cat = item.get('orig_price', item['price']), item['category']
        total_qty = sum(it['qty'] for it in self.returns_cart if it.get('base_product', it['name']) == base_name)
        
        new_cart = [it for it in self.returns_cart if it.get('base_product', it['name']) != base_name]
        new_cart.append({
            "code": "", "name": base_name, "price": orig_price, "qty": total_qty, 
            "subtotal": orig_price * total_qty, "category": cat, "base_product": base_name, "orig_price": orig_price
        })
        self.returns_cart = new_cart
        self.refresh_returns()

    def del_returns_line(self):
        if not self.ret_tree.selection(): return
        item_id = self.ret_tree.selection()[0]
        idx = self.ret_tree.index(item_id)
        self.returns_cart.pop(idx)
        self.refresh_returns()

    def clear_returns(self):
        self.returns_cart = []
        self.refresh_returns()

    def commit_returns(self):
        if not self.returns_cart: return
        tot = sum([i['subtotal'] for i in self.returns_cart])
        if not messagebox.askyesno("Confirm Return", f"Process return for {tot:.2f}?\nItems will be added back to inventory."):
            return

        now = get_current_time()
        date_str = now.strftime('%Y-%m-%d %H:%M:%S')
        fname = f"Return_{now.strftime('%Y%m%d-%H%M%S')}.pdf"

        # Generate Return Receipt (similar to sales but labeled RETURN)
        success = self.report_manager.generate_thermal_sales_receipt(
            os.path.join(CORRECTION_FOLDER, fname),
            "RETURN RECEIPT", date_str, self.returns_cart,
            tot, 0.0, user=self.session_user # Cash returned = tot
        )

        if success:
            self.data_manager.add_transaction("returns", fname, self.returns_cart, date_str, user=self.session_user)
            self.clear_returns()
            messagebox.showinfo("Success", f"Return Processed. Receipt: {fname}")

    # --- DAMAGED TAB ---
    def setup_damaged_tab(self):
        self.tab_damaged.config(style="Inventory.TFrame")
        self.setup_lws_sidebar(self.tab_damaged, "damaged")

        main_content = ttk.Frame(self.tab_damaged, style="Inventory.TFrame")
        main_content.pack(side="left", fill="both", expand=True)

        f = ttk.LabelFrame(main_content, text="Mark Damaged Items", style="Inventory.TLabelframe")
        f.pack(fill="x", padx=5, pady=5)

        top_bar = ttk.Frame(f, style="Inventory.TFrame")
        top_bar.pack(fill="x", padx=5, pady=5)

        self.dmg_prod_var = tk.StringVar()
        self.dmg_dropdown = ttk.Combobox(top_bar, textvariable=self.dmg_prod_var, width=45)
        self.dmg_dropdown['values'] = self.get_dropdown_values()
        self.setup_searchable_combobox(self.dmg_dropdown)
        self.dmg_dropdown.pack(side="left", padx=5)

        ttk.Label(top_bar, text="Qty:", style="Inventory.TLabel").pack(side="left")
        self.dmg_qty_var = tk.IntVar(value=1)
        ttk.Entry(top_bar, textvariable=self.dmg_qty_var, width=5).pack(side="left", padx=5)
        ttk.Button(top_bar, text="Add", command=self.add_damaged).pack(side="left", padx=10)

        self.lbl_dmg_stock = ttk.Label(top_bar, text="", foreground="red", font=("Segoe UI", 9, "bold"), style="Inventory.TLabel")
        self.lbl_dmg_stock.pack(side="left", padx=10)

        self.dmg_dropdown.bind("<<ComboboxSelected>>", self.on_dmg_sel)

        tree_frame = ttk.Frame(main_content, style="Inventory.TFrame")
        tree_frame.pack(fill="both", expand=True, padx=5, pady=5)

        scrollbar = ttk.Scrollbar(tree_frame)
        scrollbar.pack(side="right", fill="y")

        self.dmg_tree = ttk.Treeview(tree_frame, columns=("cat", "name", "qty"), show="headings",
                                     yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.dmg_tree.yview)
        self.dmg_tree.heading("cat", text="Category")
        self.dmg_tree.heading("name", text="Product")
        self.dmg_tree.heading("qty", text="Qty")
        self.dmg_tree.pack(fill="both", expand=True)

        b = ttk.Frame(main_content, style="Inventory.TFrame")
        b.pack(fill="x", padx=5, pady=10)
        ttk.Button(b, text="CONFIRM DAMAGES", command=self.commit_damaged, style="Accent.TButton").pack(side="right", ipadx=10)
        ttk.Button(b, text="Clear", command=self.clear_damaged).pack(side="right", padx=5)
        ttk.Button(b, text="Del Line", command=self.del_damaged_line).pack(side="right", padx=5)

    def on_dmg_sel(self, e=None):
        sel = self.dmg_prod_var.get()
        if not sel: self.lbl_dmg_stock.config(text=""); return

        _, name, _, _ = self.data_manager.get_product_details_by_str(sel)
        real_inv = self.data_manager.get_stock_level(name)
        
        # Subtract items already in damaged_cart
        in_cart = sum(i['qty'] for i in self.damaged_cart if i['name'] == name)
        
        self.lbl_dmg_stock.config(text=f"Stock: {int(real_inv - in_cart)}")

    def add_damaged(self):
        sel, qty = self.dmg_prod_var.get(), self.dmg_qty_var.get()
        if not sel or qty <= 0: return
        _, name, price, cat = self.data_manager.get_product_details_by_str(sel)

        real_inv = self.data_manager.get_stock_level(name)
        in_cart = sum(i['qty'] for i in self.damaged_cart if i['name'] == name)
        
        if qty > (real_inv - in_cart):
            messagebox.showwarning("Insufficient Stock", f"Cannot declare more than available stock ({int(real_inv - in_cart)}).")
            return

        found = False
        for i in self.damaged_cart:
            if i['name'] == name:
                i['qty'] += qty
                found = True
                break
        if not found:
            self.damaged_cart.append({"code": "", "name": name, "price": price, "qty": qty, "category": cat})
        self.refresh_damaged()
        self.on_dmg_sel() # Update stock label

    def refresh_damaged(self):
        for i in self.dmg_tree.get_children(): self.dmg_tree.delete(i)
        for i in self.damaged_cart:
            self.dmg_tree.insert("", "end", values=(i['category'], i['name'], i['qty']))

    def del_damaged_line(self):
        if not self.dmg_tree.selection(): return
        item_id = self.dmg_tree.selection()[0]
        idx = self.dmg_tree.index(item_id)
        self.damaged_cart.pop(idx)
        self.refresh_damaged()

    def clear_damaged(self):
        self.damaged_cart = []
        self.refresh_damaged()

    def commit_damaged(self):
        if not self.damaged_cart: return
        if not messagebox.askyesno("Confirm Damages", "Remove these items from inventory as DAMAGED?"):
            return

        now = get_current_time()
        date_str = now.strftime('%Y-%m-%d %H:%M:%S')
        fname = f"Damaged_{now.strftime('%Y%m%d-%H%M%S')}.pdf"

        # Calculate new stocks for receipt
        p_items = []
        for i in self.damaged_cart:
            curr_stock = self.data_manager.get_stock_level(i['name'])
            new_stock = curr_stock - i['qty']
            x = i.copy()
            x['new_stock'] = new_stock
            p_items.append(x)

        # Generate Damaged Receipt (similar to inventory but labeled DAMAGED)
        success = self.report_manager.generate_thermal_inventory_receipt(
            os.path.join(CORRECTION_FOLDER, fname),
            "DAMAGED ITEMS", date_str, p_items, user=self.session_user
        )

        if success:
            self.data_manager.add_transaction("damaged", fname, self.damaged_cart, date_str, user=self.session_user)
            self.clear_damaged()
            messagebox.showinfo("Success", f"Damages Recorded. Receipt: {fname}")

    # --- CORRECTION TAB ---
    def setup_correction_tab(self):
        paned = ttk.PanedWindow(self.tab_correction, orient="horizontal")
        paned.pack(fill="both", expand=True, padx=5, pady=5)

        frame_list = ttk.LabelFrame(paned, text="Step 1: Choose Receipt (Today)")
        paned.add(frame_list, weight=1)

        c_filter = ttk.Frame(frame_list)
        c_filter.pack(fill="x", padx=5, pady=5)
        ttk.Label(c_filter, text="Type:").pack(side="left")
        self.corr_type_var = tk.StringVar(value="sales")
        ttk.OptionMenu(c_filter, self.corr_type_var, "sales", "sales", "inventory",
                       command=lambda _: self.refresh_correction_list()).pack(side="left")
        ttk.Button(c_filter, text="Refresh", command=self.refresh_correction_list).pack(side="left", padx=5)

        self.corr_list_tree = ttk.Treeview(frame_list, columns=("time", "file"), show="headings")
        self.corr_list_tree.heading("time", text="Time")
        self.corr_list_tree.column("time", width=100)
        self.corr_list_tree.heading("file", text="Filename")
        self.corr_list_tree.pack(fill="both", expand=True, padx=5, pady=5)
        ttk.Button(frame_list, text="CHOOSE >>", command=self.load_receipt_for_correction).pack(fill="x", padx=5, pady=5)

        frame_editor = ttk.LabelFrame(paned, text="Step 2: Correct Quantities")
        paned.add(frame_editor, weight=2)

        self.lbl_corr_target = ttk.Label(frame_editor, text="No receipt selected", foreground="blue",
                                         font=("Segoe UI", 10, "bold"))
        self.lbl_corr_target.pack(padx=5, pady=5)

        self.f_corr_cash = ttk.Frame(frame_editor)
        self.f_corr_cash.pack(fill="x", padx=5, pady=2)
        self.lbl_corr_cash = ttk.Label(self.f_corr_cash, text="Cash Tendered: N/A", font=("Segoe UI", 9))
        self.lbl_corr_cash.pack(side="left")
        self.btn_corr_cash = ttk.Button(self.f_corr_cash, text="Edit Cash", command=self.ask_correction_cash, state="disabled")
        self.btn_corr_cash.pack(side="left", padx=10)
        self.correction_cash_tendered = None


        self.corr_edit_tree = ttk.Treeview(frame_editor, columns=("name", "qty_orig", "qty_adj"), show="headings")
        self.corr_edit_tree.heading("name", text="Product")
        self.corr_edit_tree.heading("qty_orig", text="Orig Qty")
        self.corr_edit_tree.column("qty_orig", width=60)
        self.corr_edit_tree.heading("qty_adj", text="Adjustment (+/-)")
        self.corr_edit_tree.column("qty_adj", width=100)
        self.corr_edit_tree.pack(fill="both", expand=True, padx=5, pady=5)
        self.corr_edit_tree.bind("<Double-1>", self.ask_correction_val)

        ttk.Label(frame_editor, text="Double click 'Adjustment' to edit. Negative (-) removes items.",
                  font=("Segoe UI", 8)).pack()
        ttk.Button(frame_editor, text="FINALIZE CORRECTION", style="Danger.TButton",
                   command=self.finalize_correction).pack(fill="x", padx=20, pady=10)
        self.selected_transaction = None

    def refresh_correction_list(self):
        for i in self.corr_list_tree.get_children(): self.corr_list_tree.delete(i)
        target_type = self.corr_type_var.get()
        now_str = get_current_time().strftime("%Y-%m-%d")

        for trans in self.data_manager.ledger:
            if trans.get('type') == target_type:
                ts = trans.get('timestamp', '')
                if ts.startswith(now_str):
                    time_part = ts.split(' ')[1] if ' ' in ts else ts
                    self.corr_list_tree.insert("", "end", values=(time_part, trans.get('filename')),
                                               tags=(json.dumps(trans),))

    def load_receipt_for_correction(self):
        sel = self.corr_list_tree.selection()
        if not sel: return

        trans_str = self.corr_list_tree.item(sel[0], 'tags')[0]
        self.selected_transaction = json.loads(trans_str)
        self.lbl_corr_target.config(text=f"Editing: {self.selected_transaction['filename']}")

        if self.selected_transaction.get('type') == 'sales':
            cash = self.selected_transaction.get('cash_tendered', 0.0)
            self.correction_cash_tendered = cash
            self.lbl_corr_cash.config(text=f"Cash Tendered: {cash:.2f}")
            self.btn_corr_cash.config(state="normal")
        else:
            self.correction_cash_tendered = None
            self.lbl_corr_cash.config(text="Cash Tendered: N/A")
            self.btn_corr_cash.config(state="disabled")


        for i in self.corr_edit_tree.get_children(): self.corr_edit_tree.delete(i)
        self.correction_cart = []
        for item in self.selected_transaction.get('items', []):
            c_item = item.copy()
            c_item['adjustment'] = 0
            self.correction_cart.append(c_item)
            self.corr_edit_tree.insert("", "end", values=(item['name'], item['qty'], 0))

    def ask_correction_val(self, event):
        if not self.selected_transaction: return
        sel = self.corr_edit_tree.selection()
        if not sel: return
        idx = self.corr_edit_tree.index(sel[0])
        item = self.correction_cart[idx]

        new_val = simpledialog.askinteger("Correction",
                                          f"Enter Adjustment for {item['name']}\n(Negative to reduce, Positive to add):",
                                          initialvalue=0, parent=self.root)
        if new_val is not None:
            self.correction_cart[idx]['adjustment'] = new_val
            self.corr_edit_tree.item(sel[0], values=(item['name'], item['qty'], new_val))

    def ask_correction_cash(self):
        if self.correction_cash_tendered is None: return
        new_val = simpledialog.askfloat("Correction", "Enter New Cash Tendered Amount:", initialvalue=self.correction_cash_tendered, parent=self.root)
        if new_val is not None:
            self.correction_cash_tendered = new_val
            self.lbl_corr_cash.config(text=f"Cash Tendered: {new_val:.2f} (Modified)")

    def finalize_correction(self):
        if not self.selected_transaction: return
        adjustments = [i for i in self.correction_cart if i['adjustment'] != 0]
        if not adjustments: messagebox.showinfo("Info", "No adjustments made."); return
        if not messagebox.askyesno("Confirm", "This will modify the database. Proceed?"): return

        # Remove old correction PDF if exists for this reference (simplification)
        ref_file = self.selected_transaction['filename']
        # Note: logic to remove old correction entry in ledger is complex, keeping simple append for safety

        now = get_current_time()
        date_str = now.strftime('%Y-%m-%d %H:%M:%S')
        fname = f"Cor_{now.strftime('%Y%m%d-%H%M%S')}.pdf"

        pdf_items = []
        ledger_adjustment_items = []

        for item in self.correction_cart:
            orig = item['qty']
            adj = item['adjustment']
            final = orig + adj
            pdf_item = {"code": "", "name": item['name'], "price": item['price'], "qty": adj, "qty_orig": orig,
                        "qty_final": final, "category": item.get('category', 'Uncategorized')}
            pdf_items.append(pdf_item)
            if adj != 0:
                ledger_item = item.copy()
                ledger_item['qty'] = adj
                ledger_adjustment_items.append(ledger_item)

        extra = f"Ref: {ref_file}"
        if self.selected_transaction['type'] == 'sales' and self.correction_cash_tendered is not None:
            extra += f" | New Cash: {self.correction_cash_tendered:.2f}"

        success = self.report_manager.generate_grouped_pdf(
            os.path.join(CORRECTION_FOLDER, fname),
            "CORRECTION RECEIPT", date_str, pdf_items,
            ["Item", "Orig", "Adj", "Final"],
            [1.0, 4.5, 5.5, 6.5], is_summary=False, extra_info=extra
        )

        if success:
            kwargs = {}
            if self.selected_transaction['type'] == 'sales':
                # Calculate adjusted total
                orig_total = sum(i['price'] * i['qty'] for i in self.selected_transaction.get('items', []))
                adj_total = sum(i['price'] * i['adjustment'] for i in self.correction_cart)
                new_total = orig_total + adj_total
                
                new_cash = self.correction_cash_tendered if self.correction_cash_tendered is not None else self.selected_transaction.get('cash_tendered', 0.0)
                new_change = new_cash - new_total
                
                kwargs['cash_tendered'] = new_cash
                kwargs['change'] = new_change
                
            self.data_manager.add_transaction("correction", fname, ledger_adjustment_items,
                                              date_str, self.selected_transaction['type'], ref_file, user=self.session_user, **kwargs)

            for i in self.corr_edit_tree.get_children(): self.corr_edit_tree.delete(i)
            self.lbl_corr_target.config(text="No receipt selected")
            self.selected_transaction = None
            messagebox.showinfo("Success", f"Correction Saved: {fname}")

    # --- RECEIPTS TAB ---
    def setup_receipts_tab(self):
        paned = ttk.PanedWindow(self.tab_receipts, orient="horizontal")
        paned.pack(fill="both", expand=True, padx=5, pady=5)

        # Left panel: Date & Filter
        left_frame = ttk.LabelFrame(paned, text="Search Receipts")
        paned.add(left_frame, weight=1)

        filter_row = ttk.Frame(left_frame)
        filter_row.pack(fill="x", padx=5, pady=5)
        
        self.rec_chk_custom_date_var = tk.BooleanVar(value=False)
        self.rec_chk_custom_date = ttk.Checkbutton(filter_row, text="Specific Date", variable=self.rec_chk_custom_date_var, command=self.toggle_rec_custom_date)
        self.rec_chk_custom_date.pack(side="left", padx=5)

        self.rec_frame_custom_date = ttk.Frame(filter_row)
        self.rec_frame_custom_date.pack(side="left")

        current_year = get_current_time().year
        self.rec_cmb_year = ttk.Combobox(self.rec_frame_custom_date, values=[y for y in range(current_year - 5, current_year + 2)], width=5, state="disabled")
        self.rec_cmb_year.set(current_year)
        self.rec_cmb_year.pack(side="left", padx=1)

        self.rec_cmb_month = ttk.Combobox(self.rec_frame_custom_date, values=[str(m).zfill(2) for m in range(1, 13)], width=3, state="disabled")
        self.rec_cmb_month.set(str(get_current_time().month).zfill(2))
        self.rec_cmb_month.pack(side="left", padx=1)

        self.rec_cmb_day = ttk.Combobox(self.rec_frame_custom_date, values=[str(d).zfill(2) for d in range(1, 32)], width=3, state="disabled")
        self.rec_cmb_day.set(str(get_current_time().day).zfill(2))
        self.rec_cmb_day.pack(side="left", padx=1)

        ttk.Label(left_frame, text="Type:").pack(anchor="w", padx=5, pady=(5,0))
        self.rec_type_var = tk.StringVar(value="All")
        ttk.OptionMenu(left_frame, self.rec_type_var, "All", "All", "sales", "inventory", "correction").pack(anchor="w", padx=5)
        
        ttk.Button(left_frame, text="Refresh", command=self.refresh_receipts_list).pack(pady=5, padx=5, anchor="w")

        self.rec_list_tree = ttk.Treeview(left_frame, columns=("time", "type", "file"), show="headings")
        self.rec_list_tree.heading("time", text="Time")
        self.rec_list_tree.heading("type", text="Type")
        self.rec_list_tree.heading("file", text="Filename")
        self.rec_list_tree.column("time", width=80)
        self.rec_list_tree.column("type", width=80)
        self.rec_list_tree.column("file", width=180)
        
        rec_scrollbar = ttk.Scrollbar(left_frame, orient="vertical", command=self.rec_list_tree.yview)
        self.rec_list_tree.configure(yscrollcommand=rec_scrollbar.set)
        
        rec_scrollbar.pack(side="right", fill="y")
        self.rec_list_tree.pack(fill="both", expand=True, padx=5, pady=5)
        self.rec_list_tree.bind("<<TreeviewSelect>>", self.on_receipt_select)

        # Right panel: Preview
        right_frame = ttk.LabelFrame(paned, text="Receipt Preview")
        paned.add(right_frame, weight=2)
        
        self.rec_preview_text = tk.Text(right_frame, font=("Courier New", 10), state="disabled", wrap="none")
        ysb = ttk.Scrollbar(right_frame, orient="vertical", command=self.rec_preview_text.yview)
        xsb = ttk.Scrollbar(right_frame, orient="horizontal", command=self.rec_preview_text.xview)
        self.rec_preview_text.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)
        
        ysb.pack(side="right", fill="y")
        xsb.pack(side="bottom", fill="x")
        self.rec_preview_text.pack(fill="both", expand=True, padx=5, pady=5)
        
    def toggle_rec_custom_date(self):
        state = "readonly" if self.rec_chk_custom_date_var.get() else "disabled"
        self.rec_cmb_year.config(state=state)
        self.rec_cmb_month.config(state=state)
        self.rec_cmb_day.config(state=state)
        
    def refresh_receipts_list(self):
        for i in self.rec_list_tree.get_children(): self.rec_list_tree.delete(i)
        
        if self.rec_chk_custom_date_var.get():
            try:
                y = int(self.rec_cmb_year.get())
                m = int(self.rec_cmb_month.get())
                d = int(self.rec_cmb_day.get())
                target_date = datetime.datetime(y, m, d).strftime("%Y-%m-%d")
            except ValueError:
                target_date = get_current_time().strftime("%Y-%m-%d")
        else:
            target_date = get_current_time().strftime("%Y-%m-%d")
            
        target_type = self.rec_type_var.get().lower()
        
        for trans in self.data_manager.ledger:
            t_type = trans.get('type', '')
            if target_type != 'all' and t_type != target_type: continue
            
            ts = trans.get('timestamp', '')
            if ts.startswith(target_date):
                time_part = ts.split(' ')[1] if ' ' in ts else ts
                self.rec_list_tree.insert("", "end", values=(time_part, t_type.capitalize(), trans.get('filename')), tags=(json.dumps(trans),))
                
    def on_receipt_select(self, event):
        sel = self.rec_list_tree.selection()
        if not sel: return
        
        trans_str = self.rec_list_tree.item(sel[0], 'tags')[0]
        trans = json.loads(trans_str)
        
        self.rec_preview_text.config(state="normal")
        self.rec_preview_text.delete("1.0", tk.END)
        
        # Build text preview mimicking 57mm thermal receipt
        width = 40  # 40 chars config based approximately on 57mm formatting in courier
        
        lines = []
        def add_centered(text): lines.append(text.center(width))
        def add_left_right(left, right): 
            space = width - len(left) - len(right)
            if space < 1: space = 1
            lines.append(left + " " * space + right)
        def add_divider(): lines.append("-" * width)
        
        add_centered(str(self.data_manager.business_name))
        add_centered(APP_TITLE)
        
        t_type = trans.get('type', '')
        if t_type == 'sales': add_centered("SALES RECEIPT")
        elif t_type == 'inventory': add_centered("INVENTORY RECEIPT")
        elif t_type == 'correction': add_centered("CORRECTION RECEIPT")
        else: add_centered("RECEIPT")
        
        lines.append("")
        lines.append(f"Date: {trans.get('timestamp', '')}")
        lines.append(f"File: {trans.get('filename', '')}")
        if t_type == 'correction':
            lines.append(f"Ref:  {trans.get('ref_filename', '')}")
            
        add_divider()
        
        if t_type == 'inventory':
            add_left_right("ITEM", "NEW STK")
        else:
            add_left_right("ITEM", "TOTAL")
            
        add_divider()
        
        total_amt = 0.0
        total_qty = 0
        total_added = 0
        
        for item in trans.get('items', []):
            name = str(item.get('name', 'Unknown'))
            if self.data_manager and name in self.data_manager.display_name_map:
                name = str(self.data_manager.display_name_map[name])
                
            if len(name) > 25:
                d_s = str(name)
                name = d_s[:12] + "..." + d_s[-10:]
                
            qty = float(item.get('qty', 0))
            price = float(item.get('price', 0.0))
            
            lines.append(name)
            
            if t_type == 'inventory':
                new_stock = float(item.get('new_stock', 0))
                add_left_right(f"  Added: {int(qty)}", str(int(new_stock)))
                total_added += int(qty)
            elif t_type == 'correction':
                orig = float(item.get('qty_orig', 0))
                adj = float(item.get('qty', 0))
                add_left_right(f"  Adj: {int(adj):+d}  Price: {price:.2f}", "")
            else:
                subtotal = float(item.get('subtotal', price * qty))
                add_left_right(f"  {int(qty)} x {price:.2f}", f"{subtotal:.2f}")
                total_amt += subtotal
                total_qty += int(qty)
                
        add_divider()
        
        if t_type == 'inventory':
            add_left_right("TOTAL ADDED:", str(int(total_added)))
        elif t_type == 'sales':
            total_discount = trans.get('total_discount', 0.0)
            if total_discount > 0:
                add_left_right("TOTAL DISCOUNT:", f"{total_discount:.2f}")
            add_left_right("TOTAL DUE:", f"{total_amt:.2f}")
            cash = trans.get('cash_tendered', 0.0)
            change = trans.get('change', 0.0)
            
            if cash > 0 or change != 0:
                add_left_right("CASH TENDERED:", f"{cash:.2f}")
                add_left_right("CHANGE:", f"{change:.2f}")
            
            lines.append("")
            lines.append(f"Total Items: {int(total_qty)}")
            
        elif t_type == 'correction':
            # Check if there is cash modifications
            if trans.get('ref_type') == 'sales':
                if 'cash_tendered' in trans and 'change' in trans:
                    add_left_right("NEW CASH:", f"{trans.get('cash_tendered', 0.0):.2f}")
                    add_left_right("NEW CHANGE:", f"{trans.get('change', 0.0):.2f}")
                    
        lines.append("")
        add_centered("*** END OF RECEIPT ***")
        
        self.rec_preview_text.insert(tk.END, "\n".join(lines))
        self.rec_preview_text.config(state="disabled")

    # --- SUMMARY TAB ---
    def setup_summary_tab(self):
        f = ttk.Frame(self.tab_summary)
        f.pack(fill="x", padx=5, pady=5)

        ttk.Label(f, text="Period:").pack(side="left")
        self.report_type = tk.StringVar(value="Daily")
        ttk.OptionMenu(f, self.report_type, "Daily", "Daily", "Weekly", "Monthly", "All Time", command=self.on_report_type_change).pack(side="left", padx=5)

        self.chk_custom_date_var = tk.BooleanVar(value=False)
        self.chk_custom_date = ttk.Checkbutton(f, text="OTHER DATE", variable=self.chk_custom_date_var,
                                               command=self.toggle_custom_date)
        self.chk_custom_date.pack(side="left", padx=10)

        self.frame_custom_date = ttk.Frame(f)
        self.frame_custom_date.pack(side="left")

        current_year = get_current_time().year
        self.cmb_year = ttk.Combobox(self.frame_custom_date,
                                     values=[y for y in range(current_year - 5, current_year + 2)], width=5,
                                     state="disabled")
        self.cmb_year.set(current_year)
        self.cmb_year.pack(side="left", padx=1)

        self.cmb_month = ttk.Combobox(self.frame_custom_date, values=[str(m).zfill(2) for m in range(1, 13)], width=3,
                                      state="disabled")
        self.cmb_month.set(str(get_current_time().month).zfill(2))
        self.cmb_month.pack(side="left", padx=1)

        self.cmb_day = ttk.Combobox(self.frame_custom_date, values=[str(d).zfill(2) for d in range(1, 32)], width=3,
                                    state="disabled")
        self.cmb_day.set(str(get_current_time().day).zfill(2))
        self.cmb_day.pack(side="left", padx=1)

        ttk.Button(f, text="Refresh View", command=self.gen_view).pack(side="left", padx=10)
        self.btn_gen_pdf = ttk.Button(f, text="Gen PDF", command=self.gen_pdf)
        self.btn_gen_pdf.pack(side="left", padx=5)

        opt_f = ttk.Frame(self.tab_summary)
        opt_f.pack(fill="x", padx=5, pady=0)
        
        self.chk_items_per_trans_var = tk.BooleanVar(value=False)
        self.chk_items_per_trans = ttk.Checkbutton(opt_f, text="View as Items per Transaction (Daily Only)", variable=self.chk_items_per_trans_var, command=self.gen_view)
        self.chk_items_per_trans.pack(side="left", padx=5)
        # It defaults to normal since report_type is Daily
        self.chk_items_per_trans.config(state="normal")
        
        tree_frame = ttk.Frame(self.tab_summary)
        tree_frame.pack(fill="both", expand=True, padx=5, pady=5)
        scrollbar = ttk.Scrollbar(tree_frame)
        scrollbar.pack(side="right", fill="y")
        self.sum_tree = ttk.Treeview(tree_frame, columns=("cat", "name", "price", "in", "out", "ret", "dmg", "rem", "sale"),
                                     show="headings", yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.sum_tree.yview)

        self.sum_tree.heading("cat", text="Cat")
        self.sum_tree.heading("name", text="Product")
        self.sum_tree.heading("price", text="Price")
        self.sum_tree.heading("in", text="In")
        self.sum_tree.heading("out", text="Out")
        self.sum_tree.heading("ret", text="Returns")
        self.sum_tree.heading("dmg", text="Damage")
        self.sum_tree.heading("rem", text="Stk")
        self.sum_tree.heading("sale", text="Sales")
        
        self.sum_tree.column("cat", width=60)
        self.sum_tree.column("name", width=180)
        self.sum_tree.column("price", width=50)
        for col in ["in", "out", "ret", "dmg", "rem"]: self.sum_tree.column(col, width=40)
        self.sum_tree.column("sale", width=60)
        
        self.sum_tree.pack(fill="both", expand=True)
        
        # Add Treeview style tags for the summary tree
        self.sum_tree.tag_configure('parent', background='#e0e0e0', font=('Helvetica', 9, 'bold'))
        self.sum_tree.tag_configure('info', foreground='#555555')
        
        bottom_f = ttk.Frame(self.tab_summary)
        bottom_f.pack(fill="x", padx=5, pady=2)
        
        self.lbl_sum_info = ttk.Label(bottom_f, text="Ready", font=("Segoe UI", 10, "bold"))
        self.lbl_sum_info.pack(side="left")
        
        self.lbl_sum_discount = ttk.Label(bottom_f, text="", font=("Segoe UI", 9, "bold"), foreground="#CC6600")
        self.lbl_sum_discount.pack(side="right", padx=10)
        
        self.lbl_sum_change = ttk.Label(bottom_f, text="", font=("Segoe UI", 9, "bold"), foreground="green")
        self.lbl_sum_change.pack(side="right", padx=10)
        
        self.lbl_sum_cash = ttk.Label(bottom_f, text="", font=("Segoe UI", 9, "bold"), foreground="blue")
        self.lbl_sum_cash.pack(side="right", padx=10)

    def on_report_type_change(self, *args):
        if self.report_type.get() == "Daily":
            self.chk_items_per_trans.config(state="normal")
            if hasattr(self, 'btn_gen_pdf'):
                self.btn_gen_pdf.config(state="normal")
        else:
            self.chk_items_per_trans_var.set(False)
            self.chk_items_per_trans.config(state="disabled")
            if hasattr(self, 'btn_gen_pdf'):
                self.btn_gen_pdf.config(state="disabled")
        self.gen_view()

    def toggle_custom_date(self):
        state = "readonly" if self.chk_custom_date_var.get() else "disabled"
        self.cmb_year.config(state=state)
        self.cmb_month.config(state=state)
        self.cmb_day.config(state=state)

    def get_period_dates(self):
        if self.chk_custom_date_var.get():
            try:
                y = int(self.cmb_year.get())
                m = int(self.cmb_month.get())
                d = int(self.cmb_day.get())
                anchor = datetime.datetime(y, m, d, 23, 59, 59)
            except ValueError:
                messagebox.showerror("Date Error", "Invalid Custom Date selected.")
                return None
        else:
            anchor = get_current_time().replace(microsecond=0)

        mode = self.report_type.get()
        if mode == "Daily": return anchor.replace(hour=0, minute=0, second=0), anchor
        if mode == "Weekly":
            start_of_week = (anchor - datetime.timedelta(days=anchor.weekday())).replace(hour=0, minute=0, second=0)
            return start_of_week, anchor
        if mode == "Monthly": return anchor.replace(day=1, hour=0, minute=0, second=0), anchor
        return None

    def get_sum_data(self, override_period=None):
        # Use filtered period if override_period provided, otherwise get from UI selection
        if override_period:
            period = override_period
        else:
            period = self.get_period_dates()

        # Calculate historical stock balance if a period is selected (shows stock AS OF the end of that period)
        if period:
            _, end_dt = period
            global_stats, _, _, _, _ = self.data_manager.calculate_stats((datetime.datetime(2000, 1, 1), end_dt))
        else:
            global_stats, _, _, _, _ = self.data_manager.calculate_stats(None)

        period_stats, in_c, out_c, corr_list, master_renames = self.data_manager.calculate_stats(period)

        rows = []
        all_names = set(self.data_manager.products_df['Product Name'].astype(str)) | set(global_stats.keys())

        names_in_excel = set(self.data_manager.products_df['Product Name'].astype(str))

        for name in all_names:
            name = name.strip()
            g_data = global_stats.get(name, {'in': 0, 'out': 0})
            rem_stock = g_data['in'] - g_data['out']

            _, _, curr_price, config_cat = self.data_manager.get_product_details_by_str(f"{name}")
            p_data = period_stats.get(name, {
                'in': 0, 'out': 0, 'returns': 0, 'damaged': 0,
                'sales_lines': [], 'in_lines': [], 'return_lines': [], 'damaged_lines': []
            })

            prices_map = {}
            # Aggregate by variant name and price point AND category
            for line in p_data.get('sales_lines', []):
                line_name = line.get('name', name)
                line_cat = line.get('category', config_cat)
                key = (line_name, line['price'], line_cat)
                if key not in prices_map: prices_map[key] = {'in': 0, 'out': 0, 'returns': 0, 'damaged': 0, 'sales': 0, 'returns_amt': 0}
                prices_map[key]['out'] += line['qty']
                prices_map[key]['sales'] += line.get('amt', 0.0)

            for line in p_data.get('in_lines', []):
                line_name = line.get('name', name)
                line_cat = line.get('category', config_cat)
                key = (line_name, line['price'], line_cat)
                if key not in prices_map: prices_map[key] = {'in': 0, 'out': 0, 'returns': 0, 'damaged': 0, 'sales': 0, 'returns_amt': 0}
                prices_map[key]['in'] += line['qty']

            for line in p_data.get('return_lines', []):
                line_name = line.get('name', name)
                line_cat = line.get('category', config_cat)
                key = (line_name, line['price'], line_cat)
                if key not in prices_map: prices_map[key] = {'in': 0, 'out': 0, 'returns': 0, 'damaged': 0, 'sales': 0, 'returns_amt': 0}
                prices_map[key]['returns'] += line['qty']
                ret_a = line.get('amt', price * line['qty'])
                prices_map[key]['returns_amt'] += ret_a
                prices_map[key]['sales'] -= ret_a # Returns decrease sales

            for line in p_data.get('damaged_lines', []):
                line_name = line.get('name', name)
                line_cat = line.get('category', config_cat)
                key = (line_name, line['price'], line_cat)
                if key not in prices_map: prices_map[key] = {'in': 0, 'out': 0, 'returns': 0, 'damaged': 0, 'sales': 0, 'returns_amt': 0}
                prices_map[key]['damaged'] += line['qty']


            if not prices_map: 
                final_cat = config_cat
                if config_cat == "Phased Out" and name in global_stats:
                    final_cat = global_stats[name].get('latest_cat', "Phased Out")
                prices_map[(name, curr_price, final_cat)] = {'in': 0, 'out': 0, 'returns': 0, 'damaged': 0, 'sales': 0, 'returns_amt': 0}

            for (variant_name, price, variant_cat), data in prices_map.items():
                is_variant = variant_name != name
                show_rem = rem_stock if price == curr_price and not is_variant else 0

                # Filter Logic
                is_all_time = (self.report_type.get() == "All Time") and (override_period is None)
                
                # Identify if product is currently phased out in products.xlsx
                is_phased_out_now = (name not in names_in_excel)

                if not is_all_time:
                    if data['in'] == 0 and data['out'] == 0 and data['returns'] == 0 and data['damaged'] == 0: continue
                elif data['in'] == 0 and data['out'] == 0 and data['returns'] == 0 and data['damaged'] == 0 and show_rem == 0 and is_phased_out_now:
                    continue

                display_name = variant_name
                if is_phased_out_now:
                    if not display_name.endswith("(Old)"):
                        display_name += " (Old)"
                        
                if is_variant:
                    display_name = "  " + display_name

                orig_p, _, _, _ = self.data_manager.get_product_details_by_str(variant_name)
                
                rows.append({
                    'code': "", 'category': variant_cat, 'name': display_name, 'price': price,
                    'orig_price': orig_p if orig_p is not None else price,
                    'in': data['in'], 'out': data['out'], 'returns': data['returns'], 
                    'returns_amt': data['returns_amt'], 'damaged': data['damaged'],
                    'remaining': show_rem, 'sales': data['sales'],
                    'base_name': name, 'is_variant': is_variant
                })

        final_rows = []
        def sort_rows(r):
            cat = r['category']
            base = r['base_name']
            is_var = r['is_variant']
            return (cat, base, is_var, r['name'])

        rows = sorted(rows, key=sort_rows)
        for r in rows:
            is_active = (r['in'] > 0 or r['out'] > 0 or r['returns'] > 0 or r['damaged'] > 0 or r['remaining'] > 0 or r['base_name'] in names_in_excel)
            if is_active: final_rows.append(r)

        return final_rows, in_c, out_c, corr_list, master_renames

    def get_sum_data_rich(self, override_period=None):
        """Returns 11 values: data, tot, p_txt, in_c, out_c, corr_list, is_daily, total_cash, total_change, per_trans, m_renames"""
        data, in_c, out_c, corr_list, m_renames = self.get_sum_data(override_period)
        
        p_txt = self.report_type.get()
        period = override_period if override_period else self.get_period_dates()
        
        # Determine if this should be treated as a daily-style report (with transaction details, cash, etc.)
        # True if user selected Daily in UI, OR if an override period for exactly 1 day is provided.
        is_daily = (p_txt == "Daily" and override_period is None)
        if override_period:
            s, e = override_period
            if (e - s).days < 1.5: # Roughly one day or less
                is_daily = True
        
        if p_txt != "All Time" and not override_period:
            if period:
                s, e = period
                if s.date() == e.date():
                    p_txt = s.strftime("%a, %d %b %y")
                else:
                    p_txt = f"{s.strftime('%a, %d %b %y')} to {e.strftime('%a, %d %b %y')}"
        else:
            period = None

        tot = 0
        total_cash = 0.0
        total_change = 0.0
        per_trans = []

        if is_daily and period:
            start_dt, end_dt = period
            sales_cash = {}
            for trans in self.data_manager.ledger:
                try:
                    ts = datetime.datetime.strptime(str(trans.get('timestamp', '')), "%Y-%m-%d %H:%M:%S")
                    if start_dt <= ts <= end_dt:
                        t_type = trans.get('type')
                        fname = trans.get('filename')
                        
                        if t_type == 'sales':
                            cash_tnd = float(trans.get('cash_tendered', 0.0))
                            change_amt = float(trans.get('change', 0.0))
                            total_disc = float(trans.get('total_discount', 0.0))
                            sales_cash[fname] = {'cash': cash_tnd, 'change': change_amt, 'total_discount': total_disc}
                            time_str = ts.strftime("%I:%M %p")
                            items = []
                            trans_subtot = 0.0
                            for item in trans.get('items', []):
                                name = item.get('name', '')
                                if name in self.data_manager.display_name_map: name = self.data_manager.display_name_map[name]
                                price = float(item.get('price', 0.0))
                                qty = int(item.get('qty', 0))
                                subtotal = float(item.get('subtotal', price * qty))
                                trans_subtot += subtotal
                                is_variant = item.get('is_variant') or name.startswith('(')
                                items.append({'name': name, 'price': price, 'qty': qty, 'subtotal': subtotal, 'is_variant': is_variant})
                            per_trans.append({'type': 'sales', 'time': time_str, 'filename': fname, 'subtotal': trans_subtot, 'cash': cash_tnd, 'change': change_amt, 'total_discount': total_disc, 'items': items})
                        elif t_type == 'inventory':
                            time_str = ts.strftime("%I:%M %p")
                            items = []
                            trans_qty = 0
                            for item in trans.get('items', []):
                                name = item.get('name', '')
                                if name in self.data_manager.display_name_map: name = self.data_manager.display_name_map[name]
                                qty = int(item.get('qty', 0))
                                trans_qty += qty
                                items.append({'name': name, 'qty': qty})
                            per_trans.append({'type': 'inventory', 'time': time_str, 'filename': fname, 'total_qty': trans_qty, 'items': items})
                        elif t_type == 'returns':
                            time_str = ts.strftime("%I:%M %p")
                            items = []
                            trans_subtot = 0.0
                            for item in trans.get('items', []):
                                name = item.get('name', '')
                                if name in self.data_manager.display_name_map: name = self.data_manager.display_name_map[name]
                                price = float(item.get('price', 0.0))
                                qty = int(item.get('qty', 0))
                                subtotal = float(item.get('subtotal', price * qty))
                                trans_subtot += subtotal
                                items.append({'name': name, 'price': price, 'qty': qty, 'subtotal': subtotal})
                            per_trans.append({'type': 'returns', 'time': time_str, 'filename': fname, 'subtotal': trans_subtot, 'items': items})
                        elif t_type == 'damaged':
                            time_str = ts.strftime("%I:%M %p")
                            items = []
                            trans_qty = 0
                            for item in trans.get('items', []):
                                name = item.get('name', '')
                                if name in self.data_manager.display_name_map: name = self.data_manager.display_name_map[name]
                                qty = int(item.get('qty', 0))
                                trans_qty += qty
                                items.append({'name': name, 'qty': qty})
                            per_trans.append({'type': 'damaged', 'time': time_str, 'filename': fname, 'total_qty': trans_qty, 'items': items})
                        elif t_type == 'correction' and trans.get('ref_type') == 'sales':
                            ref = trans.get('ref_filename')
                            if ref in sales_cash:
                                if 'cash_tendered' in trans and 'change' in trans:
                                    sales_cash[ref]['cash'] = float(trans.get('cash_tendered', 0.0))
                                    sales_cash[ref]['change'] = float(trans.get('change', 0.0))
                                    if 'total_discount' in trans: sales_cash[ref]['total_discount'] = float(trans.get('total_discount', 0.0))
                                    for pt in per_trans:
                                        if pt.get('filename') == ref:
                                            pt['cash'] = sales_cash[ref]['cash']
                                            pt['change'] = sales_cash[ref]['change']
                                            if 'total_discount' in sales_cash[ref]: pt['total_discount'] = sales_cash[ref]['total_discount']
                                            break
                except Exception: pass
            total_cash = sum(x['cash'] for x in sales_cash.values())
            total_change = sum(x['change'] for x in sales_cash.values())
            tot = sum(p.get('subtotal', 0) for p in per_trans if p.get('type') == 'sales')
            tot -= sum(p.get('subtotal', 0) for p in per_trans if p.get('type') == 'returns')
        else:
            tot = sum(s['sales'] for s in data)

        return data, tot, p_txt, in_c, out_c, corr_list, is_daily, total_cash, total_change, per_trans, m_renames

    def gen_view(self, override_period=None):
        data, tot, p_txt, in_c, out_c, corr_list, is_daily, total_cash, total_change, per_trans, m_renames = self.get_sum_data_rich(override_period)
        for i in self.sum_tree.get_children(): self.sum_tree.delete(i)

        # The metadata calculation is now handled by get_sum_data_rich
        pass

        if self.chk_items_per_trans_var.get() and is_daily:
            self.sum_tree.heading("cat", text="Time")
            self.sum_tree.heading("name", text="Transaction / Item")
            self.sum_tree.heading("price", text="Price")
            self.sum_tree.heading("in", text="-")
            self.sum_tree.heading("out", text="Qty")
            self.sum_tree.heading("ret", text="-")
            self.sum_tree.heading("dmg", text="-")
            self.sum_tree.heading("rem", text="-")
            self.sum_tree.heading("sale", text="Subtot")

            
            for pt in per_trans:
                if pt['type'] == 'sales':
                    header_text = f"SALES - {pt['filename']}"
                    parent_id = self.sum_tree.insert("", "end",
                                         values=(pt['time'], header_text, "", "", "", "", "", "", f"{pt['subtotal']:.2f}"),
                                         tags=('parent',))
                    for item in pt['items']:
                        indent = "    " if item.get('is_variant') else "  "
                        self.sum_tree.insert(parent_id, "end",
                                             values=("", f"{indent}{item['name']}", f"{item['price']:.2f}", "", str(item['qty']), "", "", "", f"{item['subtotal']:.2f}"))
                    if pt.get('total_discount', 0) > 0:
                        self.sum_tree.insert(parent_id, "end",
                                             values=("", "  TOTAL DISCOUNT", "", "", "", "", "", "", f"{pt['total_discount']:.2f}"), tags=('info',))
                    self.sum_tree.insert(parent_id, "end",
                                         values=("", "  TOTAL DUE", "", "", "", "", "", "", f"{pt['subtotal']:.2f}"), tags=('info',))
                    if pt['cash'] > 0 or pt['change'] > 0:
                        self.sum_tree.insert(parent_id, "end",
                                             values=("", "  CASH TENDERED", "", "", "", "", "", "", f"{pt['cash']:.2f}"), tags=('info',))
                        self.sum_tree.insert(parent_id, "end",
                                             values=("", "  CHANGE", "", "", "", "", "", "", f"{pt['change']:.2f}"), tags=('info',))
                    tot += pt['subtotal']

                
                elif pt['type'] == 'inventory':
                    header_text = f"INVENTORY - {pt['filename']}"
                    parent_id = self.sum_tree.insert("", "end",
                                         values=(pt['time'], header_text, "", "", str(pt['total_qty']), "", "", "", ""),
                                         tags=('parent',))
                    for item in pt['items']:
                        self.sum_tree.insert(parent_id, "end",
                                             values=("", f"  {item['name']}", "", "", str(item['qty']), "", "", "", ""))

                
                elif pt['type'] == 'returns':
                    header_text = f"RETURN - {pt['filename']}"
                    parent_id = self.sum_tree.insert("", "end",
                                         values=(pt['time'], header_text, "", "", "", str(len(pt['items'])), "", "", f"-{pt['subtotal']:.2f}"),
                                         tags=('parent',))
                    for item in pt['items']:
                        self.sum_tree.insert(parent_id, "end",
                                             values=("", f"  {item['name']}", f"{item['price']:.2f}", "", "", str(item['qty']), "", "", f"-{item['subtotal']:.2f}"))
                    tot -= pt['subtotal']


                elif pt['type'] == 'damaged':
                    header_text = f"DAMAGED - {pt['filename']}"
                    parent_id = self.sum_tree.insert("", "end",
                                         values=(pt['time'], header_text, "", "", "", "", str(pt['total_qty']), "", ""),
                                         tags=('parent',))
                    for item in pt['items']:
                        self.sum_tree.insert(parent_id, "end",
                                             values=("", f"  {item['name']}", "", "", "", "", str(item['qty']), "", ""))


        else:
            self.sum_tree.heading("cat", text="Cat")
            self.sum_tree.heading("name", text="Product")
            self.sum_tree.heading("price", text="Price")
            self.sum_tree.heading("in", text="In")
            self.sum_tree.heading("out", text="Out")
            self.sum_tree.heading("ret", text="Returns")
            self.sum_tree.heading("dmg", text="Damage")
            self.sum_tree.heading("rem", text="Stk")
            self.sum_tree.heading("sale", text="Sales")

            def sort_key(x):
                cat = x['category']
                if cat == "Phased Out": cat = "zzz_Phased Out"
                return (cat, x.get('base_name', x['name']), x.get('is_variant', False), x['name'])

            data = sorted(data, key=sort_key)
            for s in data:
                rem_val = int(s['remaining']) if s['remaining'] > 0 else ""
                if s.get('is_variant', False):
                    rem_val = ""
                
                self.sum_tree.insert("", "end",
                                     values=(s['category'], s['name'], f"{s['price']:.2f}", 
                                             int(s['in']) if s['in'] != 0 else "", 
                                             int(s['out']) if s['out'] != 0 else "",
                                             int(s.get('returns', 0)) if s.get('returns', 0) != 0 else "",
                                             int(s.get('damaged', 0)) if s.get('damaged', 0) != 0 else "",
                                             rem_val, f"{s['sales']:.2f}"))
                tot += s['sales']



        self.lbl_sum_info.config(text=f"Period: {p_txt} | Sales: {tot:.2f} | Customers: {out_c}")
        
        # Calculate total discount for the period from transaction records (reliable)
        total_discount_period = 0.0
        if per_trans:
            total_discount_period = sum(float(pt.get('total_discount', 0.0)) for pt in per_trans if pt.get('type') == 'sales')
        
        if is_daily:
            self.lbl_sum_cash.config(text=f"Total Cash Tendered: {total_cash:,.2f}")
            self.lbl_sum_change.config(text=f"Total Change: {total_change:,.2f}")
            if total_discount_period > 0:
                self.lbl_sum_discount.config(text=f"Total Discount: {total_discount_period:,.2f}")
            else:
                self.lbl_sum_discount.config(text="")
        else:
            self.lbl_sum_cash.config(text="")
            self.lbl_sum_change.config(text="")
            self.lbl_sum_discount.config(text="")
            
        self.current_aggregated = data
        self.current_per_trans = per_trans
        self.current_master_renames = m_renames
            
        return data, tot, p_txt, in_c, out_c, corr_list, is_daily, total_cash, total_change, per_trans, m_renames

    def gen_pdf(self):
        is_custom_date = self.chk_custom_date_var.get()
        data, tot, p_txt, in_c, out_c, corr_list, is_daily, total_cash, total_change, per_trans, m_renames = self.gen_view()
        now = get_current_time()

        prefix = "History" if is_custom_date else "Summary"
        fname = f"{prefix}-{now.strftime('%Y%m%d-%H%M%S')}.pdf"
        full_path = os.path.join(SUMMARY_FOLDER, fname)

        is_per_trans = self.chk_items_per_trans_var.get() and is_daily
        title = "DAILY SUMMARY" if is_daily else "INVENTORY & SALES SUMMARY"

        success = self.report_manager.generate_thermal_summary_receipt(
            full_path, title, now.strftime('%Y-%m-%d %H:%M:%S'), p_txt,
            self.current_aggregated, self.current_per_trans, is_per_trans,
            tot, total_cash, total_change, in_c, out_c, corr_list,
            master_renames=m_renames
        )

        if success:
            if not is_custom_date:
                self.data_manager.summary_count += 1
                self.data_manager.save_ledger()

                messagebox.showinfo("Success", f"Summary Generated.\nReceipt: {fname}")
            else:
                messagebox.showinfo("History Generated", f"Historical PDF Generated (No Counter).\nFile: {fname}")

    def generate_daily_summary_on_close(self, target_date: Optional[datetime.datetime] = None):
        today = target_date if target_date else get_current_time()
        start_of_day = today.replace(hour=0, minute=0, second=0, microsecond=0)
        end_of_day = today.replace(hour=23, minute=59, second=59, microsecond=999999)
        period = (start_of_day, end_of_day)
        data, tot, p_txt, in_c, out_c, corr_list, is_daily, total_cash, total_change, per_trans, m_renames = self.get_sum_data_rich(override_period=period)
        
        fname = f"Daily_Summary_{today.strftime('%Y%m%d')}.pdf"
        full_path = os.path.join(SUMMARY_FOLDER, fname)
        
        def sort_key(x):
            cat = x['category']
            if cat == "Phased Out": cat = "zzz_Phased Out"
            return (cat, x['name'])
        data.sort(key=sort_key)

        is_per_trans = False
            
        success = self.report_manager.generate_thermal_summary_receipt(
            full_path, "DAILY SUMMARY", today.strftime('%Y-%m-%d %H:%M:%S'), today.strftime("%a, %d %b %y"),
            data, per_trans, is_per_trans,
            tot, total_cash, total_change, in_c, out_c, corr_list,
            master_renames=m_renames
        )
        if success:
            print(f"Auto Daily Summary generated: {fname}")

    def get_catchup_start_time(self) -> Optional[datetime.datetime]:
        """Finds the timestamp of the oldest unsent Summary/History receipt."""
        last_sync_str = self.data_manager.config.get("last_email_sync", "")
        last_sync_dt = None
        if last_sync_str:
            try:
                last_sync_dt = datetime.datetime.strptime(last_sync_str, "%Y-%m-%d %H:%M:%S")
            except:
                pass # Invalid or empty, treat as never synced (or start from beginning?)

        # Regex to match filenames like Summary-YYYYMMDD-HHMMSS.pdf
        # Note: History- files are usually manual and maybe shouldn't count?
        # User said "summarizes all Summary Receipts generated but were not sent".
        # Let's stick to "Summary-*" files which are the official automated ones.

        candidates = []
        try:
            for f in os.listdir(SUMMARY_FOLDER):
                if f.startswith("Summary-") and f.endswith(".pdf"):
                    # Extract date
                    try:
                        part = f.replace("Summary-", "").replace(".pdf", "")
                        dt = datetime.datetime.strptime(part, "%Y%m%d-%H%M%S")

                        if last_sync_dt is None or dt > last_sync_dt:
                            candidates.append(dt)
                    except:
                        pass
        except Exception as e:
            print(f"Error scanning summary folder: {e}")

        if not candidates:
            return None

        return min(candidates)

    def get_catchup_intervals(self, start: datetime.datetime, end: datetime.datetime) -> List[Tuple[datetime.datetime, datetime.datetime]]:
        total_duration = end - start
        if total_duration.total_seconds() < 60:
            return [] # Too short to bother

        segment = total_duration / 3

        i1_end = start + segment
        i2_end = i1_end + segment

        return [
            (start, i1_end),
            (i1_end, i2_end),
            (i2_end, end)
        ]

    def update_email_sync_timestamp(self):
        now_str = get_current_time().strftime("%Y-%m-%d %H:%M:%S")
        self.data_manager.config["last_email_sync"] = now_str
        self.data_manager.save_config()

    # --- SALES REPORT TAB ---
    def setup_sales_report_tab(self):
        """Sets up the Sales Report tab with month/year selectors, data table, and export buttons."""
        top_f = ttk.Frame(self.tab_sales_report)
        top_f.pack(fill="x", padx=5, pady=5)

        ttk.Label(top_f, text="Month:", font=("Segoe UI", 10, "bold")).pack(side="left", padx=(0, 2))
        now = get_current_time()
        self.sr_month_var = tk.StringVar(value=str(now.month).zfill(2))
        self.sr_month_combo = ttk.Combobox(top_f, textvariable=self.sr_month_var,
                                           values=[str(m).zfill(2) for m in range(1, 13)],
                                           width=4, state="readonly")
        self.sr_month_combo.pack(side="left", padx=2)

        ttk.Label(top_f, text="Year:", font=("Segoe UI", 10, "bold")).pack(side="left", padx=(10, 2))
        current_year = now.year
        self.sr_year_var = tk.StringVar(value=str(current_year))
        self.sr_year_combo = ttk.Combobox(top_f, textvariable=self.sr_year_var,
                                          values=[str(y) for y in range(current_year - 5, current_year + 2)],
                                          width=6, state="readonly")
        self.sr_year_combo.pack(side="left", padx=2)

        ttk.Button(top_f, text="Generate Report", command=self.refresh_sales_report,
                   style="Accent.TButton").pack(side="left", padx=15, ipadx=8)

        export_f = ttk.Frame(top_f)
        export_f.pack(side="right")
        ttk.Button(export_f, text="Export & Send", command=self.export_sales_report,
                   style="Accent.TButton").pack(side="right", padx=3, ipadx=6)

        # Treeview – columns are built dynamically based on categories
        tree_frame = ttk.Frame(self.tab_sales_report)
        tree_frame.pack(fill="both", expand=True, padx=5, pady=5)

        scroll_y = ttk.Scrollbar(tree_frame, orient="vertical")
        scroll_y.pack(side="right", fill="y")
        scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal")
        scroll_x.pack(side="bottom", fill="x")

        self.sr_tree = ttk.Treeview(tree_frame, show="headings",
                                    yscrollcommand=scroll_y.set,
                                    xscrollcommand=scroll_x.set)
        scroll_y.config(command=self.sr_tree.yview)
        scroll_x.config(command=self.sr_tree.xview)
        self.sr_tree.pack(fill="both", expand=True)

        self.sr_tree.tag_configure('total_row', background='#C8E6C9', font=('Segoe UI', 10, 'bold'))
        self.sr_tree.tag_configure('normal', font=('Segoe UI', 9))

        # Bottom info
        bottom_f = ttk.Frame(self.tab_sales_report)
        bottom_f.pack(fill="x", padx=5, pady=2)
        self.lbl_sr_info = ttk.Label(bottom_f, text="Select a month and year, then click Generate Report.",
                                     font=("Segoe UI", 10))
        self.lbl_sr_info.pack(side="left")

        # Cache for export
        self._sr_report_data = None
        self._sr_categories = []
        self._sr_month_label = ""

    def _compute_sales_report_data(self):
        """Scans the ledger and aggregates daily sales data for the selected month/year.
        Returns (categories_sorted, daily_rows_dict, month_totals_dict, month_label_str).
        Each daily_rows_dict[date_str] = {
            'total_sales': float, 'total_discount': float, 'customers': int,
            'cat_sales': {category: float, ...}
        }
        """
        try:
            sel_month = int(self.sr_month_var.get())
            sel_year = int(self.sr_year_var.get())
        except ValueError:
            messagebox.showerror("Error", "Invalid month or year selected.")
            return None

        import calendar
        month_name = calendar.month_name[sel_month]
        month_label = f"{month_name} {sel_year}"

        # Determine the date range
        _, last_day = calendar.monthrange(sel_year, sel_month)
        start_dt = datetime.datetime(sel_year, sel_month, 1, 0, 0, 0)
        end_dt = datetime.datetime(sel_year, sel_month, last_day, 23, 59, 59)

        # Collect all categories from products and from ledger sales in this period
        all_categories = set()
        daily_data = {}  # date_str -> {'total_sales', 'total_discount', 'customers', 'cat_sales': {cat: amt}}

        for trans in self.data_manager.ledger:
            try:
                t_type = trans.get('type')
                if t_type not in ('sales', 'correction', 'returns', 'damaged'):
                    continue

                ts_str = trans.get('timestamp', '')
                try:
                    ts = datetime.datetime.strptime(ts_str, "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    continue

                if not (start_dt <= ts <= end_dt):
                    continue

                date_key = ts.strftime("%Y-%m-%d")

                if date_key not in daily_data:
                    daily_data[date_key] = {
                        'total_sales': 0.0,
                        'total_discount': 0.0,
                        'total_returns': 0.0,
                        'total_damaged': 0,
                        'customers': 0,
                        'cat_sales': {}
                    }


                day = daily_data[date_key]

                if t_type == 'sales':
                    day['customers'] += 1
                    total_disc = float(trans.get('total_discount', 0.0))
                    day['total_discount'] += total_disc

                    for item in trans.get('items', []):
                        cat = item.get('category', 'Uncategorized')
                        subtotal = float(item.get('subtotal', float(item.get('price', 0)) * int(item.get('qty', 0))))
                        all_categories.add(cat)
                        day['cat_sales'][cat] = day['cat_sales'].get(cat, 0.0) + subtotal
                        day['total_sales'] += subtotal

                elif t_type == 'correction' and trans.get('ref_type') == 'sales':
                    # Corrections on sales adjust totals
                    total_disc = float(trans.get('total_discount', 0.0))
                    day['total_discount'] += total_disc

                    for item in trans.get('items', []):
                        cat = item.get('category', 'Uncategorized')
                        price = float(item.get('price', 0))
                        qty = int(item.get('qty', 0))
                        subtotal = price * qty
                        all_categories.add(cat)
                        day['cat_sales'][cat] = day['cat_sales'].get(cat, 0.0) + subtotal
                        day['total_sales'] += subtotal

                elif t_type == 'returns':
                    total_ret = 0.0
                    for item in trans.get('items', []):
                        paid_amt = float(item.get('subtotal', float(item.get('price', 0)) * int(item.get('qty', 0))))
                        total_ret += paid_amt
                        
                        # Subtract from category too to keep category sum = total sales
                        cat = item.get('category', 'Uncategorized')
                        day['cat_sales'][cat] = day['cat_sales'].get(cat, 0.0) - paid_amt
                        
                    day['total_returns'] += total_ret
                    day['total_sales'] -= total_ret

                elif t_type == 'damaged':
                    total_dmg = 0
                    for item in trans.get('items', []):
                        total_dmg += int(item.get('qty', 0))
                    day['total_damaged'] += total_dmg


            except Exception:
                continue

        # Sort categories (exclude Phased Out — no sales possible from phased-out products)
        all_categories.discard("Phased Out")
        categories = sorted(all_categories)

        return categories, daily_data, month_label

    def refresh_sales_report(self):
        """Refreshes the Sales Report treeview with aggregated data."""
        result = self._compute_sales_report_data()
        if result is None:
            return

        categories, daily_data, month_label = result

        # Store for export
        self._sr_categories = categories
        self._sr_report_data = daily_data
        self._sr_month_label = month_label

        # Build treeview columns dynamically
        col_ids = ["date"] + [f"cat_{i}" for i in range(len(categories))] + ["total_sales", "total_returns", "total_damaged", "total_discount", "customers"]
        self.sr_tree["columns"] = col_ids


        # Configure columns
        self.sr_tree.heading("date", text="Date")
        self.sr_tree.column("date", width=100, anchor="center")

        for i, cat in enumerate(categories):
            col_id = f"cat_{i}"
            self.sr_tree.heading(col_id, text=cat)
            self.sr_tree.column(col_id, width=100, anchor="e")

        self.sr_tree.heading("total_sales", text="Net Sales")
        self.sr_tree.column("total_sales", width=120, anchor="e")

        self.sr_tree.heading("total_returns", text="Returns")
        self.sr_tree.column("total_returns", width=100, anchor="e")

        self.sr_tree.heading("total_damaged", text="Damage")
        self.sr_tree.column("total_damaged", width=80, anchor="center")

        self.sr_tree.heading("total_discount", text="Discount")
        self.sr_tree.column("total_discount", width=130, anchor="e")

        self.sr_tree.heading("customers", text="Customers")
        self.sr_tree.column("customers", width=80, anchor="center")


        # Clear existing rows
        for item in self.sr_tree.get_children():
            self.sr_tree.delete(item)

        if not daily_data:
            self.lbl_sr_info.config(text=f"{month_label}: No sales data found.")
            return

        # Sort dates
        sorted_dates = sorted(daily_data.keys())

        # Monthly totals accumulators
        month_total_sales = 0.0
        month_total_returns = 0.0
        month_total_damaged = 0
        month_total_discount = 0.0
        month_total_customers = 0
        month_cat_totals = {cat: 0.0 for cat in categories}


        for date_str in sorted_dates:
            day = daily_data[date_str]

            # Format date for display (e.g. "Mar 01, Mon")
            try:
                dt = datetime.datetime.strptime(date_str, "%Y-%m-%d")
                display_date = dt.strftime("%b %d, %a")
            except ValueError:
                display_date = date_str

            row_values = [display_date]

            for i, cat in enumerate(categories):
                cat_val = day['cat_sales'].get(cat, 0.0)
                month_cat_totals[cat] += cat_val
                row_values.append(f"{cat_val:,.2f}" if cat_val > 0 else "-")

            row_values.append(f"{day['total_sales']:,.2f}")
            row_values.append(f"{day['total_returns']:,.2f}" if day['total_returns'] > 0 else "-")
            row_values.append(str(day['total_damaged']) if day['total_damaged'] > 0 else "-")
            row_values.append(f"{day['total_discount']:,.2f}" if day['total_discount'] > 0 else "-")
            row_values.append(str(day['customers']))

            month_total_sales += day['total_sales']
            month_total_returns += day['total_returns']
            month_total_damaged += day['total_damaged']
            month_total_discount += day['total_discount']
            month_total_customers += day['customers']

            self.sr_tree.insert("", "end", values=row_values, tags=('normal',))


        # Insert Monthly Totals row
        totals_row = ["MONTHLY TOTAL"]
        for cat in categories:
            val = month_cat_totals[cat]
            totals_row.append(f"{val:,.2f}" if val > 0 else "-")
        totals_row.append(f"{month_total_sales:,.2f}")
        totals_row.append(f"{month_total_returns:,.2f}" if month_total_returns > 0 else "-")
        totals_row.append(str(month_total_damaged) if month_total_damaged > 0 else "-")
        totals_row.append(f"{month_total_discount:,.2f}" if month_total_discount > 0 else "-")
        totals_row.append(str(month_total_customers))

        self.sr_tree.insert("", "end", values=totals_row, tags=('total_row',))


        self.lbl_sr_info.config(
            text=f"{month_label} | Total Sales: {month_total_sales:,.2f} | "
                 f"Total Discount: {month_total_discount:,.2f} | Customers: {month_total_customers}"
        )

    def export_sales_report(self):
        """Generates both XLSX and PDF sales reports, then sends them via email."""
        if not self._sr_report_data:
            messagebox.showwarning("No Data", "Please generate a report first before exporting.")
            return

        categories = self._sr_categories
        daily_data = self._sr_report_data
        month_label = self._sr_month_label

        if not os.path.exists(SALESREPORT_FOLDER):
            os.makedirs(SALESREPORT_FOLDER)

        safe_label = month_label.replace(" ", "_")
        xlsx_fname = f"SalesReport_{safe_label}.xlsx"
        pdf_fname = f"SalesReport_{safe_label}.pdf"
        xlsx_path = os.path.join(SALESREPORT_FOLDER, xlsx_fname)
        pdf_path = os.path.join(SALESREPORT_FOLDER, pdf_fname)

        # --- Generate XLSX ---
        xlsx_ok = self._generate_sales_report_xlsx(categories, daily_data, month_label, xlsx_path)
        if not xlsx_ok:
            return

        # --- Generate PDF ---
        pdf_ok = self._generate_sales_report_pdf(categories, daily_data, month_label, pdf_path)
        if not pdf_ok:
            return

        # --- Increment email counter and save ---
        self.data_manager.summary_count += 1
        self.data_manager.save_ledger()

        # --- Send via email ---
        recipient = self.data_manager.config.get("recipient_email", "").strip()
        if not recipient:
            recipient = SENDER_EMAIL

        self.email_manager.trigger_summary_email(
            recipient=recipient,
            summary_pdf_path=pdf_path,
            ledger_path=LEDGER_FILE,
            business_name=self.data_manager.business_name,
            count=self.data_manager.summary_count,
            user=self.session_user,
            extra_body=f"Sales Report for {month_label}.\n",
            extra_attachments=[xlsx_path]
        )

        messagebox.showinfo(
            "Export & Send Success",
            f"Sales Report exported and sent via email.\n\n"
            f"XLSX: {xlsx_path}\n"
            f"PDF: {pdf_path}\n"
            f"Email Counter: {self.data_manager.summary_count:04d}"
        )

    def _generate_sales_report_xlsx(self, categories, daily_data, month_label, full_path):
        """Internal: generates the XLSX file. Returns True on success."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror("Error", "openpyxl library is not available.")
            return False

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Report"

        # --- Styles ---
        header_font = Font(name="Segoe UI", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        total_font = Font(name="Segoe UI", bold=True, size=10)
        total_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        data_font = Font(name="Segoe UI", size=10)
        currency_fmt = '#,##0.00'
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # --- Title Row ---
        title_text = f"Sales Report - {month_label} - {self.data_manager.business_name}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6 + len(categories))
        title_cell = ws.cell(row=1, column=1, value=title_text)
        title_cell.font = Font(name="Segoe UI", bold=True, size=14, color="1B5E20")
        title_cell.alignment = Alignment(horizontal="center")

        # --- Header Row ---
        headers = ["Date"] + categories + ["Net Sales", "Returns", "Damage", "Discount", "Customers"]
        header_row = 3
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # --- Data Rows ---
        sorted_dates = sorted(daily_data.keys())
        month_total_sales = 0.0
        month_total_returns = 0.0
        month_total_damaged = 0
        month_total_discount = 0.0
        month_total_customers = 0
        month_cat_totals = {cat: 0.0 for cat in categories}

        for row_offset, date_str in enumerate(sorted_dates):
            day = daily_data[date_str]
            row_num = header_row + 1 + row_offset

            try:
                dt = datetime.datetime.strptime(date_str, "%Y-%m-%d")
                display_date = dt.strftime("%b %d, %a")
            except ValueError:
                display_date = date_str

            cell = ws.cell(row=row_num, column=1, value=display_date)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

            for i, cat in enumerate(categories):
                cat_val = day['cat_sales'].get(cat, 0.0)
                month_cat_totals[cat] += cat_val
                cell = ws.cell(row=row_num, column=2 + i, value=cat_val if abs(cat_val) > 0.001 else None)
                cell.font = data_font
                cell.number_format = currency_fmt
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="right")

            cell = ws.cell(row=row_num, column=2 + len(categories), value=day['total_sales'])
            cell.font = data_font
            cell.number_format = currency_fmt
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right")

            cell = ws.cell(row=row_num, column=3 + len(categories), value=day['total_returns'] if day['total_returns'] > 0 else None)
            cell.font = data_font
            cell.number_format = currency_fmt
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right")

            cell = ws.cell(row=row_num, column=4 + len(categories), value=day['total_damaged'] if day['total_damaged'] > 0 else None)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

            disc_val = day['total_discount'] if day['total_discount'] > 0 else None
            cell = ws.cell(row=row_num, column=5 + len(categories), value=disc_val)
            cell.font = data_font
            cell.number_format = currency_fmt
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right")

            cell = ws.cell(row=row_num, column=6 + len(categories), value=day['customers'])
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

            month_total_sales += day['total_sales']
            month_total_returns += day['total_returns']
            month_total_damaged += day['total_damaged']
            month_total_discount += day['total_discount']
            month_total_customers += day['customers']

        # --- Totals Row ---
        totals_row_num = header_row + 1 + len(sorted_dates)
        cell = ws.cell(row=totals_row_num, column=1, value="MONTHLY TOTAL")
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

        for i, cat in enumerate(categories):
            val = month_cat_totals[cat]
            cell = ws.cell(row=totals_row_num, column=2 + i, value=val if abs(val) > 0.001 else None)
            cell.font = total_font
            cell.fill = total_fill
            cell.number_format = currency_fmt
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right")

        cell = ws.cell(row=totals_row_num, column=2 + len(categories), value=month_total_sales)
        cell.font = total_font
        cell.fill = total_fill
        cell.number_format = currency_fmt
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="right")

        cell = ws.cell(row=totals_row_num, column=3 + len(categories), value=month_total_returns if month_total_returns > 0 else None)
        cell.font = total_font
        cell.fill = total_fill
        cell.number_format = currency_fmt
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="right")

        cell = ws.cell(row=totals_row_num, column=4 + len(categories), value=month_total_damaged if month_total_damaged > 0 else None)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

        cell = ws.cell(row=totals_row_num, column=5 + len(categories),
                       value=month_total_discount if month_total_discount > 0 else None)
        cell.font = total_font
        cell.fill = total_fill
        cell.number_format = currency_fmt
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="right")

        cell = ws.cell(row=totals_row_num, column=6 + len(categories), value=month_total_customers)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

        # --- Column Widths ---
        ws.column_dimensions[get_column_letter(1)].width = 14
        for i in range(len(categories)):
            ws.column_dimensions[get_column_letter(2 + i)].width = 16
        ws.column_dimensions[get_column_letter(2 + len(categories))].width = 18 # Net Sales
        ws.column_dimensions[get_column_letter(3 + len(categories))].width = 15 # Returns
        ws.column_dimensions[get_column_letter(4 + len(categories))].width = 10 # Damage
        ws.column_dimensions[get_column_letter(5 + len(categories))].width = 18 # Discount
        ws.column_dimensions[get_column_letter(6 + len(categories))].width = 12 # Customers

        try:
            wb.save(full_path)
            return True
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to save XLSX: {e}")
            return False


    def _generate_sales_report_pdf(self, categories, daily_data, month_label, full_path):
        """Internal: generates the PDF file (A4 Landscape). Returns True on success."""
        try:
            canvas_mod = self.mod.canvas
            inch = self.mod.inch
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.colors import HexColor
        except ImportError:
            messagebox.showerror("Error", "reportlab library is not available.")
            return False

        try:
            page_size = landscape(A4)
            width, height = page_size
            c = canvas_mod.Canvas(full_path, pagesize=page_size)

            # Margins
            left_margin = 0.5 * inch
            right_margin = width - 0.5 * inch
            top_margin = height - 0.5 * inch
            usable_width = right_margin - left_margin

            # --- Calculate column widths ---
            num_cat = len(categories)
            date_w = 1.0 * inch
            total_sales_w = 0.9 * inch
            returns_w = 0.7 * inch
            damage_w = 0.6 * inch
            total_disc_w = 0.9 * inch
            customers_w = 0.7 * inch
            fixed_w = date_w + total_sales_w + returns_w + damage_w + total_disc_w + customers_w
            remaining_w = usable_width - fixed_w


            if num_cat > 0:
                cat_w = max(remaining_w / num_cat, 0.8 * inch)
            else:
                cat_w = 0

            # Build column positions
            col_positions = [left_margin]
            x = left_margin + date_w
            for i in range(num_cat):
                col_positions.append(x)
                x += cat_w
            col_positions.append(x)  # Net Sales
            x += total_sales_w
            col_positions.append(x)  # Returns
            x += returns_w
            col_positions.append(x)  # Damage
            x += damage_w
            col_positions.append(x)  # Total Discount
            x += total_disc_w
            col_positions.append(x)  # Customers

            headers = ["Date"] + categories + ["Net Sales", "Returns", "Damage", "Discount", "Customers"]


            def draw_page_header(y_pos):
                c.setFont("Helvetica-Bold", 14)
                c.setFillColor(HexColor("#1B5E20"))
                c.drawString(left_margin, y_pos, f"Sales Report - {month_label} - {self.data_manager.business_name}")
                y_pos -= 6
                c.setFont("Helvetica", 8)
                c.setFillColor(HexColor("#555555"))
                c.drawString(left_margin, y_pos - 10,
                             f"Generated: {get_current_time().strftime('%Y-%m-%d %H:%M:%S')}")
                y_pos -= 28

                header_h = 18
                c.setFillColor(HexColor("#1B5E20"))
                c.rect(left_margin, y_pos - header_h + 4, usable_width, header_h, fill=1, stroke=0)

                c.setFillColor(HexColor("#FFFFFF"))
                c.setFont("Helvetica-Bold", 7 if num_cat > 4 else 8)
                for i, hdr in enumerate(headers):
                    x_pos = col_positions[i] + 3
                    c.drawString(x_pos, y_pos - 10, hdr)

                y_pos -= header_h
                return y_pos

            def draw_row(y_pos, values, is_total=False):
                row_h = 16
                if is_total:
                    c.setFillColor(HexColor("#C8E6C9"))
                    c.rect(left_margin, y_pos - row_h + 4, usable_width, row_h, fill=1, stroke=0)
                    c.setFillColor(HexColor("#1B5E20"))
                    c.setFont("Helvetica-Bold", 8)
                else:
                    c.setFillColor(HexColor("#000000"))
                    c.setFont("Helvetica", 8)

                for i, val in enumerate(values):
                    x_pos = col_positions[i] + 3
                    c.drawString(x_pos, y_pos - 10, str(val))

                c.setStrokeColor(HexColor("#CCCCCC"))
                c.setLineWidth(0.3)
                c.line(left_margin, y_pos - row_h + 4, left_margin + usable_width, y_pos - row_h + 4)
                return y_pos - row_h

            # --- Build rows ---
            sorted_dates = sorted(daily_data.keys())
            data_rows = []
            month_total_sales = 0.0
            month_total_returns = 0.0
            month_total_damaged = 0
            month_total_discount = 0.0
            month_total_customers = 0
            month_cat_totals = {cat: 0.0 for cat in categories}


            for date_str in sorted_dates:
                day = daily_data[date_str]
                try:
                    dt = datetime.datetime.strptime(date_str, "%Y-%m-%d")
                    display_date = dt.strftime("%b %d, %a")
                except ValueError:
                    display_date = date_str

                row = [display_date]
                for cat in categories:
                    cat_val = day['cat_sales'].get(cat, 0.0)
                    month_cat_totals[cat] += cat_val
                    row.append(f"{cat_val:,.2f}" if abs(cat_val) > 0.001 else "-")
                
                row.append(f"{day['total_sales']:,.2f}")
                row.append(f"{day['total_returns']:,.2f}" if day['total_returns'] > 0 else "-")
                row.append(str(day['total_damaged']) if day['total_damaged'] > 0 else "-")
                row.append(f"{day['total_discount']:,.2f}" if day['total_discount'] > 0 else "-")
                row.append(str(day['customers']))

                month_total_sales += day['total_sales']
                month_total_returns += day['total_returns']
                month_total_damaged += day['total_damaged']
                month_total_discount += day['total_discount']
                month_total_customers += day['customers']
                data_rows.append(row)

            total_row = ["MONTHLY TOTAL"]
            for cat in categories:
                val = month_cat_totals[cat]
                total_row.append(f"{val:,.2f}" if abs(val) > 0.001 else "-")
            total_row.append(f"{month_total_sales:,.2f}")
            total_row.append(f"{month_total_returns:,.2f}" if month_total_returns > 0 else "-")
            total_row.append(str(month_total_damaged) if month_total_damaged > 0 else "-")
            total_row.append(f"{month_total_discount:,.2f}" if month_total_discount > 0 else "-")
            total_row.append(str(month_total_customers))


            # --- Render Pages ---
            y = top_margin
            y = draw_page_header(y)
            row_h = 16

            for row in data_rows:
                if y - row_h < 0.5 * inch:
                    c.showPage()
                    y = top_margin
                    y = draw_page_header(y)
                y = draw_row(y, row, is_total=False)

            if y - row_h < 0.5 * inch:
                c.showPage()
                y = top_margin
                y = draw_page_header(y)
            draw_row(y, total_row, is_total=True)

            c.save()
            return True
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to generate PDF: {e}")
            return False

    # --- SETTINGS TAB ---
    def setup_settings_tab(self):
        self.settings_notebook = ttk.Notebook(self.tab_settings)
        self.settings_notebook.pack(fill="both", expand=True, padx=5, pady=5)

        self.tab_settings_general = ttk.Frame(self.settings_notebook)
        self.tab_settings_web = ttk.Frame(self.settings_notebook)

        self.settings_notebook.add(self.tab_settings_general, text="General")
        self.settings_notebook.add(self.tab_settings_web, text="Web Server")

        # Create a scrollable frame for the general settings to handle many items
        canvas = tk.Canvas(self.tab_settings_general, borderwidth=0, background=PASTEL_GREEN_BG)
        scrollbar = ttk.Scrollbar(self.tab_settings_general, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Container for all sections
        f = ttk.Frame(scrollable_frame, padding=10)
        f.pack(fill="both", expand=True)

        # 1. APPLICATION SETTINGS
        app_f = ttk.LabelFrame(f, text="Application Settings")
        app_f.pack(fill="x", padx=10, pady=5)

        self.chk_startup_var = tk.BooleanVar(value=self.data_manager.config.get("startup", False))
        ttk.Checkbutton(app_f, text="Launch at Startup", variable=self.chk_startup_var, command=self.toggle_startup).pack(pady=10, padx=10, anchor="w")

        # 2. EMAIL SYNC
        email_f = ttk.LabelFrame(f, text="Email Receipt Sync")
        email_f.pack(fill="x", padx=10, pady=5)
        
        ttk.Label(email_f, text="Automatically send receipts and summaries to this address:", font=("Segoe UI", 9)).pack(anchor="w", padx=10, pady=(5,0))
        
        email_entry_frame = ttk.Frame(email_f)
        email_entry_frame.pack(anchor="w", pady=10, padx=10, fill="x")
        ttk.Label(email_entry_frame, text="Recipient Email:").pack(side="left")
        self.entry_email = ttk.Entry(email_entry_frame, width=35)
        self.entry_email.insert(0, self.data_manager.config.get("recipient_email", ""))
        self.entry_email.pack(side="left", padx=5)
        ttk.Button(email_entry_frame, text="Confirm & Test", command=self.verify_and_test_email).pack(side="left", padx=5)

        # Helper to create setting rows with descriptions
        def add_setting_row(parent, title, description, button_text, command, warning=None, danger=False):
            container = ttk.Frame(parent)
            container.pack(fill="x", padx=10, pady=8)
            
            # Left side: Text
            text_frame = ttk.Frame(container)
            text_frame.pack(side="left", fill="both", expand=True)
            
            ttk.Label(text_frame, text=title, font=("Segoe UI", 10, "bold")).pack(anchor="w")
            desc_label = ttk.Label(text_frame, text=description, font=("Segoe UI", 9), wraplength=500)
            desc_label.pack(anchor="w")
            
            # Right side: Button
            btn_style = "Danger.TButton" if danger else "TButton"
            
            def wrapped_cmd():
                confirm_title = f"Confirm Action: {title}"
                msg = f"{description}\n\n"
                if warning:
                    msg += f"WARNING: {warning}\n\n"
                msg += "Do you want to proceed?"
                
                if messagebox.askyesno(confirm_title, msg, icon='warning' if (danger or warning) else 'question'):
                    command()

            btn = ttk.Button(container, text=button_text, command=wrapped_cmd, style=btn_style, width=20)
            btn.pack(side="right", padx=10)
            return container

        # 3. BACKUP & RESTORE
        br_f = ttk.LabelFrame(f, text="Backup / Restore")
        br_f.pack(fill="x", padx=10, pady=5)

        add_setting_row(br_f, "Backup Ledger", 
                        "Creates a backup of the entire transaction ledger and products to a JSON file.",
                        "Backup (.json)", self.backup_data_json)
        
        add_setting_row(br_f, "Restore Ledger", 
                        "Restores data from a previously created JSON backup.",
                        "Restore (.json)", self.restore_data_json,
                        warning="This will OVERWRITE your current transactions and products! A restart will be required.")
        
        add_setting_row(br_f, "Beginning Inventory", 
                        "Generates and opens today's Beginning Inventory report based on the current ledger.",
                        "View Today's BI", self.view_today_bi)

        # 4. MAINTENANCE
        m_f = ttk.LabelFrame(f, text="System Maintenance")
        m_f.pack(fill="x", padx=10, pady=5)

        add_setting_row(m_f, "Harmonize Receipts", 
                        "Ensures all internal record counters match the actual saved receipt files. Recommended if you notice missing logs.",
                        "Harmonize Now", lambda: self.harmonize_receipts(silent=False),
                        warning="This will scan all receipt folders and may take a moment depending on the number of files.")
        
        add_setting_row(m_f, "Recover Usernames", 
                        "Recovers missing cashier/user information in the ledger by scanning existing PDF receipts.",
                        "Harmonize Usernames", self.harmonize_usernames)
        
        add_setting_row(m_f, "Current Stock Export", 
                        "Exports current inventory levels (Category, Name, Quantity) to an Excel file.",
                        "Export Stock (.xlsx)", self.export_stock_xlsx)
        
        add_setting_row(m_f, "Restore Products File", 
                        "Restore the products database from the latest automatic backup. Use if the current file is corrupted.",
                        "Restore Backup", self.regenerate_products_file,
                        warning="CAUTION: This will overwrite your current 'products.xlsx' with the latest backup. Any unsaved changes since the last backup will be lost.")

        # --- PRODUCT EDITOR ---

        add_setting_row(m_f, "Product Editor", 
                        "Rename products and update categories while preserving all historical data and stock carryover.",
                        "Open Editor", self.open_product_editor)

        po_f = ttk.LabelFrame(f, text="Phased Out Products")
        po_f.pack(fill="x", padx=10, pady=5)
        
        add_setting_row(po_f, "Clear Inventory", 
                        "Review and clear existing stock of phased-out products (products removed from products.xlsx). An email report will be sent.",
                        "Manage Items", self.manage_phased_out_products)

        # 5. DANGER ZONE
        d_f = ttk.LabelFrame(f, text="Danger Zone")
        d_f.pack(fill="x", padx=10, pady=5)
        
        add_setting_row(d_f, "Factory Reset", 
                        "WIPES all transaction data, receipts, and configuration. USE WITH EXTREME CAUTION.",
                        "DELETE ALL DATA", self.delete_all_data,
                        warning="This is IRREVERSIBLE. Only products.xlsx and the program itself will remain.",
                        danger=True)

        # Web Server Panel
        self.setup_web_server_panel(self.tab_settings_web)

    def view_today_bi(self):
        today_str = get_current_time().strftime('%Y%m%d')
        fname = f"BI-{today_str}.pdf"
        full_path = os.path.join(SUMMARY_FOLDER, fname)
        
        if not os.path.exists(full_path):
            success = self.generate_beginning_inventory_report(silent=True)
            if not success:
                return # Error already shown in generator
        
        try:
            if os.name == 'nt':
                os.startfile(full_path)
            else:
                import subprocess
                subprocess.call(['open' if sys.platform == 'darwin' else 'xdg-open', full_path])
        except Exception as e:
            messagebox.showerror("Error", f"Could not open PDF: {e}")

    def ensure_recipient_email(self) -> None:
        """
        Called on startup. If recipient_email is blank in config, automatically
        set it to SENDER_EMAIL so that BI and daily summary emails are always
        sent without requiring the user to visit Settings first.
        """
        current = self.data_manager.config.get("recipient_email", "").strip()
        if not current:
            self.data_manager.config["recipient_email"] = SENDER_EMAIL
            self.data_manager.save_config()
            # Update the entry widget in Settings if it has been built
            if hasattr(self, 'entry_email'):
                self.entry_email.delete(0, tk.END)
                self.entry_email.insert(0, SENDER_EMAIL)
            print(f"[Startup] No recipient email configured. Defaulting to sender: {SENDER_EMAIL}")

    def verify_and_test_email(self):
        email_input = self.entry_email.get().strip()
        if not email_input:
            self.data_manager.config["recipient_email"] = ""
            self.data_manager.save_config()
            messagebox.showinfo("Email Disabled", "Email field cleared.")
            return

        if not self.email_manager.validate_email_format(email_input):
            messagebox.showerror("Invalid Email", "Invalid format.")
            return

        self.data_manager.config["recipient_email"] = email_input
        self.data_manager.save_config()

        date_str = get_current_time().strftime("%Y-%m-%d %H:%M:%S")
        subject = f"{self.data_manager.business_name} - System Sync Test ({get_current_time().strftime('%Y-%m-%d')})"
        body = (
            f"Hello,\n\n"
            f"This is an automated test email to confirm that your recipient email address has been successfully configured for {self.data_manager.business_name}.\n\n"
            f"--- Details ---\n"
            f"Configured By: {self.session_user}\n"
            f"Date & Time:   {date_str}\n"
            f"---------------\n\n"
            f"You will now automatically receive inventory reports, daily summaries, and ledger backups at this address.\n\n"
            f"Best regards,\n"
            f"{APP_TITLE}"
        )
        self.email_manager.send_email_thread(email_input, subject, body, is_test=True,
                                             on_success=lambda: messagebox.showinfo("Success", "Test email sent."),
                                             on_error=lambda e: messagebox.showerror("Error", str(e)))

    def toggle_startup(self):
        startup_folder = os.path.join(os.getenv("APPDATA"), r"Microsoft\Windows\Start Menu\Programs\Startup")
        bat_path = os.path.join(startup_folder, "POS_System_Auto.bat")

        if self.chk_startup_var.get():
            try:
                exe_path = sys.executable if getattr(sys, 'frozen', False) else os.path.abspath(__file__)
                with open(bat_path, "w") as bat:
                    bat.write(f'start "" "{exe_path}"' if getattr(sys, 'frozen', False) else f'python "{exe_path}"')
                self.data_manager.config["startup"] = True
                self.data_manager.save_config()
                messagebox.showinfo("Startup", "Enabled.")
            except Exception as e:
                self.chk_startup_var.set(False)
                messagebox.showerror("Error", str(e))
        else:
            if os.path.exists(bat_path): os.remove(bat_path)
            self.data_manager.config["startup"] = False
            self.data_manager.save_config()
            messagebox.showinfo("Startup", "Disabled.")

    def open_product_editor(self):
        """Opens the external product editor module."""
        try:
            from product_editor import ProductEditor
            editor = ProductEditor(self)
            editor.show_editor_window()
        except ImportError as e:
            messagebox.showerror("Error", f"Product Editor module (product_editor.py) failed to load: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"Could not open Product Editor: {e}")

    def manage_phased_out_products(self):
        phased_out = []
        names_in_excel = set(self.data_manager.products_df['Product Name'].astype(str))
        for name, st in self.data_manager.stock_cache.items():
            stock = st['in'] - st['out']
            if name not in names_in_excel and stock != 0:
                phased_out.append({"name": name, "stock": stock})
                
        if not phased_out:
            messagebox.showinfo("Info", "No phased out products with existing stock found.")
            return

        win = tk.Toplevel(self.root)
        win.title("Clear Phased Out Products")
        win.geometry("500x400")
        
        ttk.Label(win, text="Select phased out products to clear their stock:", font=("Segoe UI", 10, "bold")).pack(pady=10)
        
        canvas = tk.Canvas(win)
        scrollbar = ttk.Scrollbar(win, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="top", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        vars_dict = {}
        for item in phased_out:
            var = tk.BooleanVar(value=False)
            vars_dict[item['name']] = (var, item['stock'])
            cb = ttk.Checkbutton(scrollable_frame, text=f"{item['name']} (Stock: {item['stock']})", variable=var)
            cb.pack(anchor="w", padx=10, pady=2)
            
        def on_confirm():
            to_clear = [name for name, (var, qty) in vars_dict.items() if var.get()]
            if not to_clear:
                messagebox.showinfo("Info", "No products selected.")
                return
                
            if messagebox.askyesno("Confirm", f"Are you sure you want to clear the stock of {len(to_clear)} phased out product(s)?"):
                action_items = []
                for name in to_clear:
                    stock = vars_dict[name][1]
                    action_items.append({
                        "name": name,
                        "qty": -stock,
                        "price": 0,
                        "category": "Phased Out"
                    })
                    
                now = get_current_time()
                date_str = now.strftime('%Y-%m-%d %H:%M:%S')
                fname = f"PhasedOutClear_{now.strftime('%Y%m%d-%H%M%S')}"
                
                self.data_manager.add_transaction("inventory", fname, action_items, date_str, user=self.session_user)
                self.data_manager.refresh_stock_cache()
                # Phased out items don't appear in standard UI dropdowns, so no UI refresh needed here 
                
                email_text = (
                    f"Hello,\n\n"
                    f"This is an automated notification regarding phased-out product inventory for {self.data_manager.business_name}.\n\n"
                    f"--- Details ---\n"
                    f"Action By:   {self.session_user}\n"
                    f"Date & Time: {date_str}\n"
                    f"---------------\n\n"
                    f"The following phased-out products have had their remaining stock cleared from the system:\n\n"
                )
                for name in to_clear:
                    email_text += f"- {name} (Cleared stock: {vars_dict[name][1]})\n"
                
                email_text += f"\nBest regards,\n{APP_TITLE}"
                    
                recipient = self.data_manager.config.get("recipient_email", "").strip() or SENDER_EMAIL
                subject = f"{self.data_manager.business_name} - Phased Out Stock Cleared ({now.strftime('%Y-%m-%d')})"
                
                def on_fail(err):
                    record = {
                        "timestamp": date_str,
                        "subject": subject,
                        "body": email_text,
                        "files": []
                    }
                    self.data_manager.offline_receipts.append(record)
                    self.data_manager.save_ledger()

                self.email_manager.send_email_thread(
                    recipient, subject, email_text, [],
                    on_error=on_fail
                )
                
                messagebox.showinfo("Success", f"Cleared {len(to_clear)} phased out product(s).")
                win.destroy()
                
        btn_frame = ttk.Frame(win)
        btn_frame.pack(fill="x", pady=10)
        
        ttk.Button(btn_frame, text="Confirm", command=on_confirm).pack(side="right", padx=10)
        ttk.Button(btn_frame, text="Cancel", command=win.destroy).pack(side="right")

    def backup_data_json(self):
        if not self.data_manager.ledger:
            messagebox.showinfo("Backup", "No data to backup.")
            return
        save_path = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON Database", "*.json")])
        if save_path:
            try:
                products_data = self.data_manager.get_product_list()
                data = {
                    "transactions": self.data_manager.ledger,
                    "summary_count": self.data_manager.summary_count,
                    "products_master": products_data,
                    "shortcuts_asked": self.data_manager.shortcuts_asked
                }
                with open(save_path, 'w') as f:
                    json.dump(data, f, indent=2)
                messagebox.showinfo("Backup", "Done.")
            except Exception as e:
                messagebox.showerror("Error", f"Backup failed: {e}")

    def restore_data_json(self):
        path = filedialog.askopenfilename(filetypes=[("JSON Database", "*.json")])
        if not path: return
        if not messagebox.askyesno("Confirm", "Overwrite data and REGENERATE receipts?"): return
        try:
            with open(path, 'r') as f:
                backup_data = json.load(f)

            # Restore logic coordinated via DataManager
            products_master = []
            if isinstance(backup_data, list):
                self.data_manager.ledger = backup_data
                self.data_manager.summary_count = 0
            elif isinstance(backup_data, dict):
                self.data_manager.ledger = backup_data.get("transactions", [])
                self.data_manager.summary_count = backup_data.get("summary_count", 0)
                self.data_manager.shortcuts_asked = backup_data.get("shortcuts_asked", False)

                # Restore products if present (basic regen)
                products_master = backup_data.get("products_master", [])
                
                # --- NEW LOGIC: Fallback to product_history if products_master missing ---
                if not products_master:
                    history = backup_data.get("product_history", [])
                    if history and isinstance(history, list) and len(history) > 0:
                        products_master = history[-1].get("items", [])

            # --- NEW LOGIC: Reconstruct from transactions if still missing ---
            if not products_master and self.data_manager.ledger:
                extracted_products = {}
                for entry in self.data_manager.ledger:
                    for item in entry.get("items", []):
                        name = item.get("name")
                        if name:
                            extracted_products[name] = {
                                "Business Name": getattr(self.data_manager, 'business_name', "MY BUSINESS"),
                                "Product Category": item.get("category", "UNCATEGORIZED"),
                                "Product Name": name,
                                "Price": item.get("price", 0.0)
                            }
                products_master = list(extracted_products.values())

            if products_master and hasattr(self, 'mod') and self.mod.pd is not None:
                try:
                    new_df = self.mod.pd.DataFrame(products_master)
                    new_df.to_excel(DATA_FILE, index=False)
                    self.data_manager.load_products()
                    # Update UI
                    self.inv_dropdown['values'] = self.get_dropdown_values()
                    self.pos_dropdown['values'] = self.get_dropdown_values()
                except Exception:
                    pass # Non-critical

            self.harmonize_receipts(silent=True)
            self.data_manager.save_ledger()
            self.data_manager.refresh_stock_cache()
            
            # Set flag for BI and Summary generation upon restart
            self.data_manager.config['restored_flag'] = True
            self.data_manager.save_config()
            
            messagebox.showinfo("Restored", "Data restored successfully.\n\nThe application will now close. Please manually reopen it to finalize report synchronization.")
            self.root.destroy()

        except Exception as e:
            messagebox.showerror("Error", f"Failed: {e}")

    def harmonize_receipts(self, silent: bool = False):
        if not silent:
            if not messagebox.askyesno("Confirm", "This will DELETE and REGENERATE all PDF receipts from the database.\nContinue?"):
                return

        try:
            # Regenerate Receipts
            for folder in [INVENTORY_FOLDER, RECEIPT_FOLDER, CORRECTION_FOLDER]:
                if os.path.exists(folder): shutil.rmtree(folder); os.makedirs(folder)

            for entry in self.data_manager.ledger:
                fname = entry.get('filename')
                date_str = entry.get('timestamp')
                items = entry.get('items', [])
                orig_user = entry.get('user', self.session_user)

                if entry['type'] == "inventory":
                    self.report_manager.generate_grouped_pdf(
                        os.path.join(INVENTORY_FOLDER, fname), "INVENTORY RECEIPT", date_str,
                        items, ["Item", "Price", "Qty Added", "New Stock"], [1.0, 4.5, 5.5, 6.5],
                        subtotal_indices=[2], is_inventory=True, user=orig_user
                    )
                elif entry['type'] == "sales":
                    self.report_manager.generate_grouped_pdf(
                        os.path.join(RECEIPT_FOLDER, fname), "SALES RECEIPT", date_str, items,
                        ["Item", "Price", "Qty", "Total"], [1.0, 4.5, 5.5, 6.5],
                        subtotal_indices=[2, 3], user=orig_user
                    )
                elif entry['type'] == "correction":
                    pdf_items = []
                    for it in items:
                        pdf_items.append({
                            "code": "", "name": it['name'], "price": it['price'],
                            "qty_orig": 0, "qty": it['qty'], "qty_final": it['qty'],
                            "category": it.get('category')
                        })
                    self.report_manager.generate_grouped_pdf(
                        os.path.join(CORRECTION_FOLDER, fname), "CORRECTION RECEIPT", date_str,
                        pdf_items, ["Item", "Orig", "Adj", "Final"], [1.0, 4.5, 5.5, 6.5], user=orig_user
                    )

            if not silent:
                messagebox.showinfo("Success", "All receipts harmonized with ledger.")

        except Exception as e:
            if not silent:
                messagebox.showerror("Error", f"Harmonization Failed: {e}")
            else:
                raise e

    def regenerate_products_file(self):
        history = self.data_manager.product_history
        if not history or len(history) < 2:
            messagebox.showinfo("History", "No previous product versions available.")
            return

        # Show versions (excluding the very last one which is current)
        # candidates are history[:-1] if we assume the last one is always current.
        # But wait, history[-1] is the one loaded *now*.
        # If the user changed the file externally and opened the app, history[-1] matches the file.
        # The user wants "past 3 versions prior to the current version".

        # If we have [A, B, C, D(current)], we want user to pick from C, B, A.
        # So we slice history[:-1] and reverse it to show newest first.
        candidates = history[:-1]
        if not candidates:
             messagebox.showinfo("History", "No *past* versions available.")
             return

        # Create a simple dialog
        win = tk.Toplevel(self.root)
        win.title("Select Version")
        win.geometry("400x300")

        ttk.Label(win, text="Select a past version to restore:", font=("Segoe UI", 10, "bold")).pack(pady=10)

        tree = ttk.Treeview(win, columns=("ts", "count"), show="headings")
        tree.heading("ts", text="Timestamp")
        tree.heading("count", text="Items")
        tree.column("ts", width=200)
        tree.column("count", width=80)
        tree.pack(fill="both", expand=True, padx=10, pady=5)

        # Sort candidates newest first
        sorted_candidates = sorted(candidates, key=lambda x: x['timestamp'], reverse=True)

        # Limit to 3 as requested
        sorted_candidates = sorted_candidates[:3]

        for ver in sorted_candidates:
            tree.insert("", "end", values=(ver['timestamp'], len(ver['items'])), tags=(json.dumps(ver),))

        def restore():
            sel = tree.selection()
            if not sel: return

            data_str = tree.item(sel[0], 'tags')[0]
            version_data = json.loads(data_str)
            items = version_data.get('items', [])

            if not items:
                messagebox.showerror("Error", "Selected version has no items.")
                return

            if messagebox.askyesno("Confirm", f"Restore products from {version_data['timestamp']}?\nThis will overwrite products.xlsx."):
                try:
                    df = self.mod.pd.DataFrame(items)
                    # Filter out internal columns (starting with underscore)
                    clean_df = df[[c for c in df.columns if not str(c).startswith('_')]]
                    clean_df.to_excel(DATA_FILE, index=False)
                    self.data_manager.load_products()

                    # Update UI Dropdowns
                    self.inv_dropdown['values'] = self.get_dropdown_values()
                    self.pos_dropdown['values'] = self.get_dropdown_values()

                    messagebox.showinfo("Success", "Products restored and reloaded.")
                    win.destroy()
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to restore: {e}")

        ttk.Button(win, text="Restore Selected", command=restore).pack(pady=10)

    def run_load_test(self):
        pwd = simpledialog.askstring("Load Test", "Enter Password:", show="*")
        if pwd != "migs": messagebox.showerror("Error", "Incorrect Password"); return
        if not messagebox.askyesno("WARNING",
                                   "This will DELETE ALL DATA and generate dummy data for the last 30 days.\n\nAre you sure?"): return

        self.data_manager.ledger = []
        self.data_manager.summary_count = 0
        for folder in [INVENTORY_FOLDER, RECEIPT_FOLDER, CORRECTION_FOLDER]:
            if os.path.exists(folder): shutil.rmtree(folder); os.makedirs(folder)

        if self.data_manager.products_df.empty:
            messagebox.showerror("Error", "No products loaded from products.xlsx")
            return

        products = []
        for _, row in self.data_manager.products_df.iterrows():
            products.append(
                {"name": row['Product Name'], "price": float(row['Price']), "category": row['Product Category']})

        stock_tracker = {p['name']: 0 for p in products}
        start_date = get_current_time() - datetime.timedelta(days=30)

        try:
            for day_offset in range(31):
                curr_date = start_date + datetime.timedelta(days=day_offset)
                date_str_base = curr_date.strftime("%Y-%m-%d")

                # Inventory Logic
                if day_offset % 7 == 0 or day_offset == 30:
                    inv_items = []
                    for p in products:
                        current_qty = stock_tracker[p['name']]
                        weekly_demand_est = 21
                        safety_stock = random.randint(10, 20)
                        target_level = weekly_demand_est + safety_stock
                        needed = target_level - current_qty
                        if needed > 0:
                            stock_tracker[p['name']] += needed
                            inv_items.append({"code": "", "name": p['name'], "price": p['price'], "qty": needed,
                                              "category": p['category'], "new_stock": stock_tracker[p['name']]})
                    if inv_items:
                        ts = f"{date_str_base} 08:00:00"
                        fname = f"Inventory_{curr_date.strftime('%Y%m%d')}-080000.pdf"

                        self.report_manager.generate_grouped_pdf(
                            os.path.join(INVENTORY_FOLDER, fname), "INVENTORY RECEIPT", ts,
                            inv_items, ["Item", "Price", "Qty Added", "New Stock"],
                            [1.0, 4.5, 5.5, 6.5], subtotal_indices=[2], is_inventory=True
                        )
                        self.data_manager.ledger.append(
                            {"type": "inventory", "timestamp": ts, "filename": fname, "items": inv_items})

                # Sales Logic
                num_sales = random.randint(5, 10)
                for s_i in range(num_sales):
                    sales_items = []
                    num_lines = random.randint(1, 5)
                    attempts = 0
                    while len(sales_items) < num_lines and attempts < 20:
                        attempts += 1
                        p = random.choice(products)
                        if any(x['name'] == p['name'] for x in sales_items): continue

                        qty = random.randint(1, 3)
                        if stock_tracker[p['name']] >= qty:
                            stock_tracker[p['name']] -= qty
                            sub = p['price'] * qty
                            sales_items.append(
                                {"code": "", "name": p['name'], "price": p['price'], "qty": qty, "subtotal": sub,
                                 "category": p['category']})

                    if sales_items:
                        hour = 9 + (s_i % 9)
                        minute = random.randint(0, 59)
                        ts = f"{date_str_base} {hour:02d}:{minute:02d}:{random.randint(10, 59)}"
                        fname = f"{curr_date.strftime('%Y%m%d')}-{hour:02d}{minute:02d}{random.randint(10, 59)}.pdf"

                        self.report_manager.generate_grouped_pdf(
                            os.path.join(RECEIPT_FOLDER, fname), "SALES RECEIPT", ts, sales_items,
                            ["Item", "Price", "Qty", "Total"], [1.0, 4.5, 5.5, 6.5],
                            subtotal_indices=[2, 3]
                        )
                        self.data_manager.ledger.append({"type": "sales", "timestamp": ts, "filename": fname, "items": sales_items})

            self.data_manager.save_ledger()
            self.data_manager.refresh_stock_cache()
            messagebox.showinfo("Load Test", "Simulation Complete.\nData overwritten.")

        except Exception as e:
            messagebox.showerror("Load Test Error", f"Simulation failed: {e}")

    def delete_all_data(self):
        confirm = messagebox.askyesno("Delete All Data", "Are you sure you want to delete all transaction data?\n\nThis will delete the entire AppData folder and all receipt folders.\nOnly the application and products.xlsx will remain.", icon='warning')
        if not confirm:
            return

        # Check internet connection
        if not self.is_online():
            messagebox.showerror("No Internet", "An active internet connection is required to ensure a secure backup is emailed before data wipes.")
            return

        confirm2 = simpledialog.askstring("Confirm Deletion", "Type 'DELETE' to confirm deletion of all data:")
        if confirm2 != "DELETE":
            messagebox.showinfo("Cancelled", "Data deletion cancelled.")
            return

        try:
            # 4. Security Backup (Requirement 1 & 2)
            now = get_current_time()
            self.generate_daily_summary_on_close() # Ensure today's summary is generated
            
            # Increment and save count
            self.data_manager.summary_count += 1
            self.data_manager.save_ledger()
            
            summary_fname = f"Daily_Summary_{now.strftime('%Y%m%d')}.pdf"
            summary_path = os.path.join(SUMMARY_FOLDER, summary_fname)
            
            import tempfile
            import shutil
            temp_dir = tempfile.mkdtemp()
            temp_ledger = os.path.join(temp_dir, "ledger.json")
            temp_summary = os.path.join(temp_dir, summary_fname)
            
            shutil.copy2(LEDGER_FILE, temp_ledger)
            summary_exists = os.path.exists(summary_path)
            if summary_exists:
                shutil.copy2(summary_path, temp_summary)
            
            recipient = self.data_manager.config.get("recipient_email", "").strip()
            
            email_done = threading.Event()
            email_error = [None]
            
            def on_success():
                email_done.set()
                
            def on_error(msg):
                email_error[0] = msg
                email_done.set()

            self.email_manager.trigger_summary_email(
                recipient=recipient,
                summary_pdf_path=temp_summary if summary_exists else "",
                ledger_path=temp_ledger,
                business_name=self.data_manager.business_name,
                count=self.data_manager.summary_count,
                user=self.session_user,
                extra_body="--- SECURITY BACKUP BEFORE FACTORY RESET ---\nThis is an automated backup triggered by a 'DELETE ALL DATA' request.\n",
                on_success=on_success,
                on_error=on_error
            )
            
            # Progress window
            wait_win = tk.Toplevel(self.root)
            wait_win.title("Security Backup")
            wait_win.geometry("300x120")
            wait_win.resizable(False, False)
            if self.root.winfo_viewable():
                wait_win.transient(self.root)
            wait_win.grab_set()
            
            ttk.Label(wait_win, text="Sending secure backup email...\nPlease do not close the software.", font=('Helvetica', 10, 'bold'), padding=20, wraplength=250, justify="center").pack()
            wait_win.update()
            
            start_wait = time.time()
            while not email_done.is_set() and time.time() - start_wait < 60: # 60s timeout
                self.root.update()
                time.sleep(0.1)
                
            wait_win.destroy()
            
            if not email_done.is_set():
                messagebox.showerror("Backup Failed", "Timed out waiting for backup email. Deletion aborted for security.")
                return
            
            if email_error[0]:
                messagebox.showerror("Backup Failed", f"Failed to send backup email: {email_error[0]}\n\nDeletion aborted for security.")
                return

            # Cleanup temp files
            try: shutil.rmtree(temp_dir, ignore_errors=True)
            except: pass

            # 5. Data Deletion (Original Logic)
            if os.path.exists(APP_DATA_DIR):
                shutil.rmtree(APP_DATA_DIR, ignore_errors=True)

            # 2. Cleanup current directory
            # Keep only the executable/script and products.xlsx
            if getattr(sys, 'frozen', False):
                exe_name = os.path.basename(sys.executable)
            else:
                exe_name = os.path.basename(__file__)
            
            keep_files = {exe_name, DATA_FILE}
            
            cwd = os.getcwd()
            for item in os.listdir(cwd):
                if item in keep_files:
                    continue
                
                item_path = os.path.join(cwd, item)
                try:
                    if os.path.isdir(item_path):
                        shutil.rmtree(item_path, ignore_errors=True)
                    else:
                        os.remove(item_path)
                except:
                    pass # Best effort cleanup for locked files

            # 3. Notification (Optional, but good for UX)
            # messagebox.showinfo("Success", "All data has been deleted. The application will now close.")
            
            # 4. Force Close immediately to avoid summary generation on exit
            os._exit(0)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete all data: {e}")

    def export_stock_xlsx(self):
        # Category, Product Name, Stock
        rows = []
        known_names = set(p['Product Name'] for p in self.data_manager.get_product_list())

        for name in list(self.data_manager.stock_cache.keys()):
            stk = self.data_manager.get_stock_level(name)
            if name not in known_names and stk <= 0:
                continue

            details = self.data_manager.get_product_details_by_str(name)
            cat = details[3] if details else "Uncategorized"
            if cat == "Phased Out":
                # Ensure it clearly says Old
                if not name.endswith("(Old)"):
                    name += " (Old)"

            rows.append({
                "Category": cat,
                "Product Name": name,
                "Stock": stk
            })
        
        if not self.data_manager.mod.pd:
            messagebox.showerror("Error", "Pandas library not loaded.")
            return
            
        df = self.data_manager.mod.pd.DataFrame(rows)
        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if save_path:
            try:
                df.to_excel(save_path, index=False)
                messagebox.showinfo("Success", f"Stock exported to:\n{save_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Export failed: {e}")

    def import_stock_xlsx(self):
        path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        if not path: return
        
        messagebox.showinfo("Import", "Please note that imported products will be added on top of existing stocks upon committing.")
        
        try:
            if not self.data_manager.mod.pd:
                messagebox.showerror("Error", "Pandas library not loaded.")
                return
            df = self.data_manager.mod.pd.read_excel(path)
            # Validate columns
            required = ["Category", "Product Name", "Stock"]
            if not all(col in df.columns for col in required):
                messagebox.showerror("Error", f"Invalid format. File must contain columns: {', '.join(required)}")
                return
            
            imported_count = 0
            errors = []
            
            for _, row in df.iterrows():
                name = str(row['Product Name']).strip()
                qty_val = row['Stock']
                try:
                    qty = int(qty_val)
                except:
                    continue
                
                # Validate against current products.xlsx
                if name not in self.data_manager.name_lookup_cache:
                    errors.append(name)
                    continue
                
                # Add to inventory cart
                it = self.data_manager.name_lookup_cache[name]
                self.inventory_cart.append({
                    "code": "",
                    "name": name,
                    "price": float(it.get('Price', 0)),
                    "qty": qty,
                    "category": it.get('Product Category', 'GENERAL')
                })
                imported_count += 1
            
            self.refresh_inv()
            
            if errors:
                error_msg = f"Imported {imported_count} items.\n\nThe following products were NOT found in products.xlsx and were skipped:\n" + "\n".join(errors[:10])
                if len(errors) > 10: error_msg += f"\n...and {len(errors)-10} more."
                messagebox.showwarning("Import Partial", error_msg)
            else:
                messagebox.showinfo("Success", f"Successfully imported {imported_count} items to the inventory cart.")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to import: {e}")

    def harmonize_usernames(self):
        if not self.data_manager.mod.PdfReader:
            messagebox.showerror("Error", "pypdf library not found. Cannot read usernames from PDFs.")
            return
            
        updated = 0
        
        for entry in self.data_manager.ledger:
            if not entry.get('user'):
                fname = entry.get('filename')
                if not fname: continue
                
                t_type = entry.get('type')
                
                # Determine folder
                folder = ""
                if t_type == "sales": folder = RECEIPT_FOLDER
                elif t_type == "inventory": folder = INVENTORY_FOLDER
                elif t_type == "correction": folder = CORRECTION_FOLDER
                elif t_type == "summary": folder = SUMMARY_FOLDER
                
                path = os.path.join(folder, fname)
                if os.path.exists(path):
                    try:
                        reader = self.data_manager.mod.PdfReader(path)
                        text = ""
                        for page in reader.pages:
                            text += page.extract_text()
                        
                        found_user = None
                        if "Cashier: " in text:
                            found_user = text.split("Cashier: ")[1].split("\n")[0].strip()
                        elif "User: " in text:
                            found_user = text.split("User: ")[1].split("\n")[0].strip()
                        elif "by " in text:
                            parts = text.split("by ")
                            if len(parts) > 1:
                                found_user = parts[1].split("\n")[0].split(" ")[0].strip() # Take first word mostly
                        
                        if found_user:
                            # Clean up common PDF artifacts from split
                            found_user = found_user.split(" ")[0].strip()
                            if found_user:
                                entry['user'] = found_user
                                updated += 1
                    except Exception as e:
                        print(f"Error reading {path}: {e}")
            
        if updated > 0:
            self.data_manager.save_ledger()
            messagebox.showinfo("Success", f"Harmonized {updated} usernames in ledger.json.")
        else:
            messagebox.showinfo("Info", "No missing usernames found or could be recovered from PDFs.")

    # --- WEB SERVER GUI HELPERS ---
    def setup_web_server_panel(self, parent_frame):
        frame = ttk.Frame(parent_frame)
        frame.pack(fill="both", expand=True, padx=20, pady=20)

        left_panel = ttk.LabelFrame(frame, text="Connection Info")
        left_panel.pack(side="left", fill="both", expand=True, padx=10)

        self.lbl_url = ttk.Label(left_panel, text="Server URL:", font=("Segoe UI", 12))
        self.lbl_url.pack(pady=(20, 5))

        self.entry_url = ttk.Entry(left_panel, font=("Segoe UI", 10, "bold"), justify="center", width=40)
        self.entry_url.pack(pady=5)

        self.qr_lbl = tk.Label(left_panel, bg="white")
        self.qr_lbl.pack(pady=20, padx=20)

        ttk.Button(left_panel, text="Generate New QR (Revoke Old)", command=self.rotate_token).pack(pady=10)
        ttk.Label(left_panel, text="Scanning a new QR invalidates previous links.", foreground="red",
                  font=("Segoe UI", 9)).pack()

        right_panel = ttk.LabelFrame(frame, text="Connected Devices (Session)")
        right_panel.pack(side="right", fill="both", expand=True, padx=10)

        self.device_tree = ttk.Treeview(right_panel, columns=("ip", "count"), show="headings")
        self.device_tree.heading("ip", text="Device IP")
        self.device_tree.heading("count", text="Transactions")
        self.device_tree.column("ip", width=150)
        self.device_tree.column("count", width=100)
        self.device_tree.pack(fill="both", expand=True, padx=5, pady=5)

        self.qr_lbl.config(text="Click 'Generate New QR' to start LWS")

    def is_online(self):
        try:
            socket.create_connection(("8.8.8.8", 53), timeout=3)
            return True
        except OSError:
            return False

    def get_local_ip(self):
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            ip = s.getsockname()[0]
            s.close()
            return ip
        except:
            return "127.0.0.1"

    def find_free_port(self):
        port = 5000
        while port < 5100:
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                if s.connect_ex(('localhost', port)) != 0:
                    return port
                port += 1
        return 5000

    def start_web_server(self):
        if self.web_server_running: return

        # Dependency injection for the thread
        self.web_thread = WebServerThread(
            self.mod, self.web_queue, self.web_port,
            lambda: self.data_manager,
            lambda: self.session_token
        )
        self.web_thread.start()
        self.web_server_running = True
        self.show_remote_sidebars()

    def rotate_token(self):
        self.session_token = secrets.token_hex(4)
        if not self.web_server_running:
            self.start_web_server()
        self.generate_qr()
        self.refresh_connected_devices_table()

    def generate_qr(self):
        url = f"http://{self.local_ip}:{self.web_port}/?token={self.session_token}"
        self.entry_url.config(state="normal")
        self.entry_url.delete(0, tk.END)
        self.entry_url.insert(0, url)
        self.entry_url.config(state="readonly")

        qr = self.mod.qrcode.QRCode(box_size=10, border=4)
        qr.add_data(url)
        qr.make(fit=True)
        img = qr.make_image(fill="black", back_color="white")

        self.tk_qr = ImageTk.PhotoImage(img.resize((250, 250)))
        self.qr_lbl.config(image=self.tk_qr)

    def refresh_connected_devices_table(self):
        for i in self.device_tree.get_children():
            self.device_tree.delete(i)
        for ip, count in self.connected_devices.items():
            self.device_tree.insert("", "end", values=(ip, count))

    # --- LWS REQUEST HANDLING ---
    def setup_lws_sidebar(self, parent_frame, mode):
        sidebar_frame = ttk.Frame(parent_frame, width=250)
        sidebar_frame.pack(side="right", fill="y", padx=5, pady=5)
        sidebar_frame.pack_forget()
        sidebar_frame.pack_propagate(False)

        self.lws_sidebars[mode] = sidebar_frame
        lbl_title = ttk.Label(sidebar_frame, text="Remote Requests", font=("Segoe UI", 10, "bold"))
        lbl_title.pack(pady=5)

        tree_frame = ttk.Frame(sidebar_frame)
        tree_frame.pack(fill="both", expand=True)

        cols = ("time", "ip", "summary")
        tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
        tree.heading("time", text="Time")
        tree.heading("ip", text="IP")
        tree.heading("summary", text="Total")
        tree.column("time", width=50)
        tree.column("ip", width=80)
        tree.column("summary", width=80)
        tree.pack(fill="both", expand=True)

        if mode == 'inventory':
            self.inv_req_tree = tree
        else:
            self.pos_req_tree = tree

        btn_frame = ttk.Frame(sidebar_frame)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="Add to Cart", width=12,
                   command=lambda: self.action_remote_request(mode, "add_to_cart")).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="✖", width=4,
                   command=lambda: self.action_remote_request(mode, "reject")).pack(side="left", padx=5)

    def show_remote_sidebars(self):
        for mode, frame in self.lws_sidebars.items():
            frame.pack(side="right", fill="y", padx=5, pady=5)

    def process_web_queue(self):
        try:
            while True:
                task = self.web_queue.get_nowait()
                if task['type'] == 'web_transaction':
                    self.handle_remote_transaction(task['data'], task['ip'])
        except queue.Empty:
            pass
        self.root.after(500, self.process_web_queue)

    def handle_remote_transaction(self, data, ip):
        if ip not in self.connected_devices:
            self.connected_devices[ip] = 0
        self.connected_devices[ip] += 1
        self.refresh_connected_devices_table()

        mode = data.get('mode')
        items = data.get('items', [])

        if not items: return
        now = get_current_time()

        request_id = secrets.token_hex(4)
        request_data = {
            "id": request_id,
            "ip": ip,
            "mode": mode,
            "timestamp": now,
            "items": items
        }

        # Double Check Stock for Safety
        if mode == 'sales':
             for i in items:
                 name = i['name']
                 req_qty = int(i['qty'])
                 avail = self.data_manager.get_stock_level(name)
                 if req_qty > avail:
                     print(f"Rejected invalid stock request from {ip}")
                     return

        self.remote_requests.append(request_data)
        self.refresh_remote_sidebars()

    def refresh_remote_sidebars(self):
        # Refresh Inventory Sidebar
        for i in self.inv_req_tree.get_children(): self.inv_req_tree.delete(i)
        for req in self.remote_requests:
            if req['mode'] == 'inventory':
                time_str = req['timestamp'].strftime('%H:%M')
                total_items = sum(int(x.get('qty', 0)) for x in req['items'])
                item_summary = f"{total_items} items"
                self.inv_req_tree.insert("", "end", values=(time_str, req['ip'], item_summary), tags=(req['id'],))

        # Refresh POS Sidebar
        for i in self.pos_req_tree.get_children(): self.pos_req_tree.delete(i)
        for req in self.remote_requests:
            if req['mode'] == 'sales':
                time_str = req['timestamp'].strftime('%H:%M')
                total_amt = sum(float(x.get('price', 0)) * int(x.get('qty', 0)) for x in req['items'])
                item_summary = f"{total_amt:.2f}"
                self.pos_req_tree.insert("", "end", values=(time_str, req['ip'], item_summary), tags=(req['id'],))

    def action_remote_request(self, mode, action):
        tree = self.inv_req_tree if mode == 'inventory' else self.pos_req_tree
        sel = tree.selection()
        if not sel: return

        req_id = tree.item(sel[0], 'tags')[0]
        req = next((r for r in self.remote_requests if r['id'] == req_id), None)
        if not req: return

        if action == "reject":
            if messagebox.askyesno("Reject", "Reject this request?"):
                self.remote_requests.remove(req)
                self.refresh_remote_sidebars()
        elif action == "add_to_cart":
            self.load_remote_request_to_cart(req)

    def load_remote_request_to_cart(self, req):
        mode = req['mode']
        items = req['items']

        processed_items = []
        for i in items:
            processed_items.append({
                "code": "",
                "name": i['name'],
                "price": float(i['price']),
                "qty": int(i['qty']),
                "subtotal": float(i.get('price', 0)) * int(i.get('qty', 0)),
                "category": i.get('category', 'General')
            })

        if mode == 'sales':
            warnings = []
            for i in processed_items:
                avail = self.data_manager.get_stock_level(i['name'])
                if i['qty'] > avail:
                    warnings.append(f"{i['name']} (Req: {i['qty']}, Avail: {int(avail)})")

            if warnings:
                msg = "Insufficient stock for:\n" + "\n".join(warnings) + "\n\nAdd to cart anyway?"
                if not messagebox.askyesno("Stock Warning", msg):
                    return

            self.sales_cart = processed_items
            self.refresh_pos()
            self.on_pos_sel(None)
            self.notebook.select(self.tab_pos)

        elif mode == 'inventory':
            self.inventory_cart = processed_items
            self.refresh_inv()
            self.notebook.select(self.tab_inventory)

        self.remote_requests.remove(req)
        self.refresh_remote_sidebars()

    # --- AUTO TASKS ---
    def check_shortcuts(self):
        if not self.data_manager.shortcuts_asked:
            if messagebox.askyesno("Desktop Shortcuts", "Create shortcuts for App and Summary Folder on Desktop?"):
                self.create_shortcuts()
            self.data_manager.shortcuts_asked = True
            self.data_manager.save_ledger()

    def create_shortcuts(self):
        try:
            desktop = os.path.normpath(os.path.join(os.environ['USERPROFILE'], 'Desktop'))
            exe_path = sys.executable if getattr(sys, 'frozen', False) else os.path.abspath(__file__)
            exe_dir = os.path.dirname(exe_path)
            link_name = f"{APP_TITLE}.lnk"
            self.create_shortcut_vbs(exe_path, os.path.join(desktop, link_name), working_dir=exe_dir)

            folder_path = os.path.abspath(SUMMARY_FOLDER)
            folder_link_name = "Summary Receipts.lnk"
            self.create_shortcut_vbs(folder_path, os.path.join(desktop, folder_link_name))
            messagebox.showinfo("Shortcuts", "Shortcuts created on Desktop.")
        except Exception as e:
            messagebox.showerror("Shortcut Error", f"Could not create shortcuts: {e}")

    def create_shortcut_vbs(self, target, link_path, working_dir=None):
        work_dir_line = f'oLink.WorkingDirectory = "{working_dir}"' if working_dir else ""
        vbs_content = f"""
            Set oWS = WScript.CreateObject("WScript.Shell")
            Set oLink = oWS.CreateShortcut("{link_path}")
            oLink.TargetPath = "{target}"
            {work_dir_line}
            oLink.Save
        """
        vbs_path = os.path.join(os.environ["TEMP"], "create_shortcut.vbs")
        with open(vbs_path, "w") as f:
            f.write(vbs_content)
        os.system(f'cscript //Nologo "{vbs_path}"')
        try:
            os.remove(vbs_path)
        except:
            pass

    def check_beginning_inventory_reminder(self):
        today_str = get_current_time().strftime("%Y-%m-%d")
        last_bi_date = self.data_manager.last_bi_date
        if last_bi_date != today_str:
            self.generate_beginning_inventory_report()

    def generate_beginning_inventory_report(self, silent: bool = False):
        today = get_current_time()
        today_str = today.strftime("%Y-%m-%d")

        # Capture current state (Beginning Inventory)
        items = []
        # Get all known products
        all_products = self.data_manager.get_product_list()

        # Include any product that has stock in cache even if not in product list (unlikely but possible if phased out)
        cached_names = set(self.data_manager.stock_cache.keys())
        known_names = set(p['Product Name'] for p in all_products)

        for p in all_products:
            name = p['Product Name']
            cat = p['Product Category']
            qty = self.data_manager.get_stock_level(name)
            items.append({
                "name": name,
                "category": cat,
                "qty": qty,
                "price": 0, # Not used in BI
                "subtotal": 0 # Not used in BI
            })

        # Add items that are in stock but not in the product list (orphaned/phased out)
        for name in cached_names:
            if name not in known_names:
                qty = self.data_manager.get_stock_level(name)
                items.append({
                    "name": f"{name} (Old)",
                    "category": "Phased Out",
                    "qty": qty,
                    "price": 0,
                    "subtotal": 0
                })

        # FILTER: Show only items with POSITIVE stock (qty > 0)
        items = [i for i in items if i['qty'] > 0]

        fname = f"BI-{today.strftime('%Y%m%d')}.pdf"
        full_path = os.path.join(SUMMARY_FOLDER, fname)

        success = self.report_manager.generate_thermal_beginning_inventory_receipt(
            full_path, "BEGINNING INVENTORY",
            today.strftime('%Y-%m-%d %H:%M:%S'), items
        )

        if success:
            self.data_manager.summary_count += 1
            self.data_manager.last_bi_date = today_str
            self.data_manager.save_ledger()

            if silent:
                print(f"Beginning Inventory generated (Silent/Regenerated): {fname}")
                return True

            recipient = self.data_manager.config.get("recipient_email", "").strip()
            
            yesterday = today - datetime.timedelta(days=1)
            yesterday_str = yesterday.strftime('%Y%m%d')
            old_summary_path = os.path.join(SUMMARY_FOLDER, f"Daily_Summary_{yesterday_str}.pdf")
            
            attachments = [full_path, LEDGER_FILE]
            if os.path.exists(old_summary_path):
                attachments.append(old_summary_path)

            if not recipient:
                recipient = SENDER_EMAIL

            subject = f"[{self.data_manager.summary_count:04d}] {self.data_manager.business_name} - Beginning Inventory & Daily Summary ({today_str})"
            body = (
                f"Hello,\n\n"
                f"This is the automated Beginning Inventory report and data synchronization for {self.data_manager.business_name}.\n\n"
                f"--- Details ---\n"
                f"Active User:   {self.session_user}\n"
                f"Email Counter: {self.data_manager.summary_count:04d}\n"
                f"Date & Time:   {get_current_time().strftime('%Y-%m-%d %H:%M:%S')}\n"
                f"---------------\n\n"
                f"Attached to this email, please find:\n"
                f"- Today's Beginning Inventory PDF\n"
                f"- The most recent Ledger Database Backup (ledger.json)\n"
                f"- Yesterday's Daily Summary PDF (if available)\n\n"
                f"Best regards,\n"
                f"{APP_TITLE}"
            )

            def on_email_error(err):
                print(f"Beginning Inventory email failed, storing for catchup: {err}")
                record = {
                    "timestamp": today.strftime('%Y-%m-%d %H:%M:%S'),
                    "files": attachments,
                    "counter": self.data_manager.summary_count,
                    "subject": subject
                }
                if not hasattr(self.data_manager, 'offline_receipts'):
                    self.data_manager.offline_receipts = []
                self.data_manager.offline_receipts.append(record)
                self.data_manager.save_ledger()

            self.email_manager.send_email_thread(
                recipient, subject, body,
                attachments,
                on_error=on_email_error
            )

            # Optional: Silent or unobtrusive notification
            # self.root.after(500, lambda: messagebox.showinfo("Info", "Beginning Inventory generated."))
            # User said "automatically". Usually implies silent, but I'll print to console or status if I had one.
            print(f"Beginning Inventory generated: {fname}")


# --- SPLASH SCREEN ---
class SplashScreen(tk.Toplevel):
    def __init__(self, root, img_path, business_name, app_title):
        super().__init__(root)
        self.overrideredirect(True)
        width, height = 400, 250
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        x = (screen_width // 2) - (width // 2)
        y = (screen_height // 2) - (height // 2)
        self.geometry(f"{width}x{height}+{int(x)}+{int(y)}")
        self.configure(bg="#2b2b2b")

        if img_path and os.path.exists(img_path):
            try:
                pil_img = Image.open(img_path).resize((180, 130), Image.Resampling.LANCZOS)
                self.img = ImageTk.PhotoImage(pil_img)
                tk.Label(self, image=self.img, bg="#2b2b2b").pack(pady=10)
            except:
                pass

        display_text = f"{business_name}\n{app_title}"
        tk.Label(self, text=display_text, font=("Segoe UI", 12, "bold"), fg="white", bg="#2b2b2b",
                 justify="center").pack(pady=5)
        self.status_lbl = tk.Label(self, text="Initializing...", font=("Segoe UI", 9), fg="lightgray", bg="#2b2b2b")
        self.status_lbl.pack(side="bottom", pady=15)
        self.update()

    def update_status(self, text):
        self.status_lbl.config(text=text)
        self.update()


# --- BOOTSTRAP ---
def launch_app():
    # Close PyInstaller Splash
    try:
        import pyi_splash
        pyi_splash.update_text("Starting Application...")
        pyi_splash.close()
    except ImportError:
        pass

    root = tk.Tk()
    root.withdraw()

    # Load basic config for splash
    cfg = {"splash_img": "", "cached_business_name": "MMD POS System"}
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, 'r') as f:
                loaded_cfg = json.load(f)
                cfg.update(loaded_cfg)
        except:
            pass

    splash = SplashScreen(root, cfg.get("splash_img", ""), cfg.get("cached_business_name", ""), APP_TITLE)

    def loader():
        # Lazy Loading heavy modules
        try:
            splash.update_status("Loading Data Engine (pandas)...")
            import pandas as pd
            AppModules.pd = pd

            splash.update_status("Loading PDF Engine (reportlab)...")
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.units import inch
            AppModules.canvas = canvas
            AppModules.letter = letter
            AppModules.inch = inch

            splash.update_status("Loading Utils...")
            from pypdf import PdfReader
            import ntplib
            AppModules.PdfReader = PdfReader
            AppModules.ntplib = ntplib

            splash.update_status("Loading Web Server...")
            from flask import Flask, request, jsonify, render_template_string
            import qrcode
            AppModules.Flask = Flask
            AppModules.request = request
            AppModules.jsonify = jsonify
            AppModules.render_template_string = render_template_string
            AppModules.qrcode = qrcode

            splash.update_status("Loading Email Modules...")
            import smtplib
            import ssl
            from email.mime.text import MIMEText
            from email.mime.multipart import MIMEMultipart
            from email.mime.base import MIMEBase
            from email import encoders
            AppModules.smtplib = smtplib
            AppModules.ssl = ssl
            AppModules.MIMEText = MIMEText
            AppModules.MIMEMultipart = MIMEMultipart
            AppModules.MIMEBase = MIMEBase
            AppModules.encoders = encoders

        except Exception as e:
            messagebox.showerror("Critical Error", f"Missing Libraries: {e}")
            root.destroy()
            return

        splash.update_status("Starting Interface...")

        def login_logic():
            splash.destroy()
            user = simpledialog.askstring("Login", "User:", parent=root)
            if user:
                root.deiconify()
                # Inject modules into main system
                POSSystem(root, user, splash=None, modules=AppModules)
            else:
                root.destroy()

        root.after(500, login_logic)

    root.after(100, loader)
    root.mainloop()

# --- MOBILE HTML TEMPLATE ---
MOBILE_HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>POS Remote</title>
    <style>
        body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif; background: #f2f2f2; margin: 0; padding: 0; transition: background 0.3s; }

        /* THEMES */
        body.sales-theme .header { background: #333; }
        body.sales-theme .btn-add { background: #007bff; }
        body.sales-theme .mode-btn.active { background: #333; color: white; border-bottom: 4px solid #007bff; }

        body.stock-theme { background: #fff3e0; }
        body.stock-theme .header { background: #e65100; }
        body.stock-theme .btn-add { background: #ef6c00; }
        body.stock-theme .mode-btn.active { background: #e65100; color: white; border-bottom: 4px solid white; }

        .header { color: white; padding: 15px; text-align: center; font-weight: bold; font-size: 1.2em; box-shadow: 0 2px 4px rgba(0,0,0,0.2); }
        .container { padding: 15px; }
        .card { background: white; padding: 15px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); margin-bottom: 15px; }
        select, input { width: 100%; padding: 12px; margin: 5px 0; border: 1px solid #ddd; border-radius: 4px; box-sizing: border-box; font-size: 16px; }
        button { width: 100%; padding: 12px; border: none; border-radius: 4px; font-size: 16px; font-weight: bold; cursor: pointer; margin-top: 10px; color: white; }

        .btn-success { background: #28a745; }
        .btn-danger { background: #dc3545; }

        .cart-item { display: flex; justify-content: space-between; padding: 10px 0; border-bottom: 1px solid #eee; }
        .mode-switch { display: flex; background: #ddd; }
        .mode-btn { flex: 1; padding: 15px; text-align: center; cursor: pointer; font-weight:bold; color: #555; transition: all 0.3s; }

        .stock-tag { font-size: 0.8em; color: #666; display: block; margin-bottom: 5px; text-align: right; }
        .error-overlay { position: fixed; top:0; left:0; width:100%; height:100%; background:white; color:red; display:flex; align-items:center; justify-content:center; font-size:1.5em; text-align:center; z-index:999; display:none;}
    </style>
</head>
<body class="sales-theme">
    <div id="auth-error" class="error-overlay">
        Session Expired or Invalid.<br>Please scan the QR code again.
    </div>

    <div class="header" id="header-title">POS REMOTE: SALES</div>

    <div class="mode-switch">
        <div class="mode-btn active" id="btn-sales" onclick="setMode('sales')">SALES</div>
        <div class="mode-btn" id="btn-stock" onclick="setMode('inventory')">STOCK IN</div>
    </div>

    <div class="container">
        <div class="card">
            <label>Select Product:</label>
            <select id="product-select" onchange="updateStockDisplay()"></select>
            <span id="stock-display" class="stock-tag">Checking stock...</span>

            <label>Quantity:</label>
            <input type="number" id="qty" value="1" min="1">

            <button class="btn-add" onclick="addToCart()">ADD TO CART</button>
        </div>

        <div class="card">
            <h3>Cart (<span id="mode-label">Sales</span>)</h3>
            <div id="cart-list"></div>
            <hr>
            <div style="text-align: right; font-weight: bold; font-size: 1.2em;">Total: <span id="total-amt">0.00</span></div>
            <button class="btn-success" onclick="submitTransaction()">REQUEST</button>
            <button class="btn-danger" style="margin-top:5px; background: #666;" onclick="clearCart()">CLEAR</button>
        </div>
    </div>

    <script>
        const AUTH_TOKEN = "{{ token }}";

        let products = [];
        let cart = [];
        let currentMode = 'sales';

        // Load Products with Token and Stock
        fetch('/get_products?token=' + AUTH_TOKEN)
        .then(r => {
            if (r.status === 403) throw new Error("Unauthorized");
            return r.json();
        })
        .then(data => {
            products = data.products; // Now includes 'stock' key
            const sel = document.getElementById('product-select');
            products.forEach(p => {
                let opt = document.createElement('option');
                opt.value = p.name;
                opt.text = p.name + ' (' + p.price + ')';
                sel.add(opt);
            });
            updateStockDisplay();
        })
        .catch(e => {
            document.getElementById('auth-error').style.display = 'flex';
        });

        function setMode(mode) {
            currentMode = mode;
            // Update UI Theme
            document.body.className = mode === 'sales' ? 'sales-theme' : 'stock-theme';

            document.getElementById('btn-sales').className = mode === 'sales' ? 'mode-btn active' : 'mode-btn';
            document.getElementById('btn-stock').className = mode === 'inventory' ? 'mode-btn active' : 'mode-btn';

            document.getElementById('mode-label').innerText = mode === 'sales' ? 'Sales' : 'Stock In';
            document.getElementById('header-title').innerText = mode === 'sales' ? 'POS REMOTE: SALES' : 'POS REMOTE: STOCK IN';

            cart = [];
            renderCart();
            updateStockDisplay();
        }

        function updateStockDisplay() {
            const name = document.getElementById('product-select').value;
            const prod = products.find(p => p.name === name);
            const display = document.getElementById('stock-display');

            if (prod) {
                if (currentMode === 'sales') {
                    // Check local cart to subtract from display
                    let inCart = 0;
                    let existing = cart.find(i => i.name === name);
                    if(existing) inCart = existing.qty;

                    let avail = prod.stock - inCart;
                    display.innerText = "Available Stock: " + avail;
                    display.style.color = avail < 5 ? "red" : "#666";
                } else {
                    display.innerText = "Current Stock: " + prod.stock + " (Adding Mode)";
                    display.style.color = "green";
                }
            }
        }

        function addToCart() {
            const name = document.getElementById('product-select').value;
            const qty = parseInt(document.getElementById('qty').value);
            if(qty < 1 || isNaN(qty)) return;

            const prod = products.find(p => p.name === name);
            if(!prod) return;

            // --- STOCK CHECK (Client Side) ---
            if (currentMode === 'sales') {
                let currentCartQty = 0;
                let existingItem = cart.find(i => i.name === name);
                if (existingItem) currentCartQty = existingItem.qty;

                if ((currentCartQty + qty) > prod.stock) {
                    alert("Insufficient Stock! You have " + prod.stock + " available.");
                    return;
                }
            }

            let existing = cart.find(i => i.name === name);
            if(existing) {
                existing.qty += qty;
                existing.subtotal = existing.qty * existing.price;
            } else {
                cart.push({
                    name: prod.name,
                    price: prod.price,
                    category: prod.category,
                    qty: qty,
                    subtotal: prod.price * qty
                });
            }
            renderCart();
            updateStockDisplay();
        }

        function renderCart() {
            const list = document.getElementById('cart-list');
            list.innerHTML = '';
            let total = 0;
            cart.forEach((item, index) => {
                total += item.subtotal;
                let div = document.createElement('div');
                div.className = 'cart-item';
                div.innerHTML = `<span>${item.name} x${item.qty}</span> <span>${item.subtotal.toFixed(2)}</span> <span style='color:red; cursor:pointer; margin-left:10px;' onclick='remItem(${index})'>[x]</span>`;
                list.appendChild(div);
            });
            document.getElementById('total-amt').innerText = total.toFixed(2);
        }

        function remItem(idx) {
            cart.splice(idx, 1);
            renderCart();
            updateStockDisplay();
        }

        function clearCart() {
            cart = [];
            renderCart();
            updateStockDisplay();
        }

        function submitTransaction() {
            if(cart.length === 0) return alert("Cart is empty");

            // Re-validate all items before sending (Client side double check)
            if (currentMode === 'sales') {
                for (let item of cart) {
                    let prod = products.find(p => p.name === item.name);
                    if (prod && item.qty > prod.stock) {
                        alert("Stock changed! " + item.name + " has insufficient quantity.");
                        return;
                    }
                }
            }

            if(!confirm("Submit this request to PC?")) return;

            // Submit with Token in URL
            fetch('/submit_transaction?token=' + AUTH_TOKEN, {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({mode: currentMode, items: cart})
            })
            .then(r => {
                if (r.status === 403) throw new Error("Unauthorized");
                return r.json();
            })
            .then(res => {
                if(res.status === 'success') {
                    alert("Success! Request sent to PC.");
                    cart = [];
                    renderCart();
                    // Refresh product list to get new stock levels
                    location.reload();
                } else {
                    alert("Error: " + res.message);
                }
            })
            .catch(e => {
                if (e.message === "Unauthorized") document.getElementById('auth-error').style.display = 'flex';
                else alert("Connection Error");
            });
        }
    </script>
</body>
</html>
"""

if __name__ == "__main__":
    launch_app()